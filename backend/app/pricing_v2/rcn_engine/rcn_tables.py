"""Base RCN reference tables and size-scaling logic.

Contains the category-specific base RCN values, HP/weight scaling,
category normalization, and spec modifier computation.

Source: V1 rcn_v2/calculator.py (table and lookup portions).
"""
from __future__ import annotations

import os
import threading
from dataclasses import dataclass
from typing import Any

# ── POLICY CONSTANTS ──────────────────────────────────────────────────

DEFAULT_BASE_RCN = 100_000.0
DEFAULT_UNKNOWN_RCN_CONFIDENCE = 0.30
HP_SCALING_EXPONENT = 0.6
WEIGHT_SCALING_EXPONENT = 0.85
WEIGHT_SCALING_BASE_LBS = 10_000.0

# ── BASE RCN TABLES ──────────────────────────────────────────────────

ROTATING_BASE_RCN: dict[str, dict[str, dict[str, float]]] = {
    "compressor_package": {
        "reciprocating": {"base_rcn": 100_000, "base_hp": 100},
        "screw": {"base_rcn": 80_000, "base_hp": 100},
        "centrifugal": {"base_rcn": 160_000, "base_hp": 500},
    },
    "pump": {
        "centrifugal": {"base_rcn": 22_000, "base_hp": 50},
        "positive_displacement": {"base_rcn": 38_000, "base_hp": 50},
        "progressive_cavity": {"base_rcn": 18_000, "base_hp": 25},
    },
    "generator": {
        "diesel": {"base_rcn": 60_000, "base_hp": 100},
        "natural_gas": {"base_rcn": 75_000, "base_hp": 100},
        "dual_fuel": {"base_rcn": 85_000, "base_hp": 100},
    },
    "blower": {
        "rotary_lobe": {"base_rcn": 32_000, "base_hp": 50},
        "centrifugal": {"base_rcn": 48_000, "base_hp": 100},
    },
}

STATIC_BASE_RCN: dict[str, float] = {
    "separator": 130_000,
    "tank": 50_000,
    "treater": 200_000,
    "vessel": 110_000,
    "heater": 85_000,
}

HEAVY_EQUIP_BASE_RCN: dict[str, float] = {
    "loader": 250_000,
    "excavator": 300_000,
    "dozer": 350_000,
    "grader": 280_000,
    "backhoe": 120_000,
}

TRUCK_BASE_RCN: dict[str, float] = {
    "class_8": 280_000,
    "class_6_7": 150_000,
    "vocational": 200_000,
}

CATEGORY_BASE_MAP: dict[str, str] = {
    "compressors": "compressor_package",
    "compressor": "compressor_package",
    "compressor_package": "compressor_package",
    "separators": "separator",
    "separator": "separator",
    "vessel": "vessel",
    "tanks": "tank",
    "tank": "tank",
    "production": "treater",
    "treater": "treater",
    "heater": "heater",
    "pumps": "pump",
    "pump": "pump",
    "generators": "generator",
    "generator": "generator",
    "blower": "blower",
    "electrical": "e_house",
    "e_house": "e_house",
    "mcc": "e_house",
    "switchgear": "e_house",
    "vfd": "e_house",
    "loaders": "loader",
    "loader": "loader",
    "excavators": "excavator",
    "excavator": "excavator",
    "dozer": "dozer",
    "grader": "grader",
    "backhoe": "backhoe",
    "heavy_construction": "loader",
    "trucks": "truck",
    "truck": "truck",
    "trailers": "truck",
    "trailer": "truck",
}

DEFAULT_ROTATING_TYPE: dict[str, str] = {
    "compressor_package": "reciprocating",
    "pump": "centrifugal",
    "generator": "diesel",
    "blower": "rotary_lobe",
}


@dataclass
class RCNInput:
    """Validated calculator input. Plain dataclass — no Pydantic."""
    current_year: int | None = None
    year: int | None = None
    assumed_age_years: int | None = None
    hours: float | None = None
    horsepower: float | None = None
    weight_lbs: float | None = None
    truck_class: str = "class_8"
    equipment_type: str = ""
    condition: str | None = None
    is_nace_compliant: bool = False
    years_h2s_exposure: float | None = None
    wti_price: float | None = None
    region: str | None = None
    location: str | None = None
    material: str | None = None
    drive_type: str | None = None
    comparable_count: int = 0
    comparable_cv: float | None = None
    data_age_days: int | None = None
    spec_modifiers: dict[str, float] | list[float] | float | None = None
    # Tier 2.5: tank volume in BBL for $/BBL seed-bracket lookup. Populated
    # by the bulk runner via regex extraction from title/description.
    volume_bbl: float | None = None


# ── TANK $/BBL SEED LADDER (lazy-loaded from xlsx, cached at module scope) ─

_TANK_SEED_LOCK = threading.Lock()
_TANK_SEED_CACHE: list[tuple[float, float]] | None = None  # [(bbl, mid_rcn), ...] sorted ascending.

# Two candidate paths searched in order. The Railway deploy uploads only
# `backend/`, so `backend/seeds/...` is what prod uses. Local dev / tests
# typically have the original at the repo-root `seeds/...`. RCN_SEED_PATH
# env var, if set, overrides both.
_TANK_SEED_PATH_CANDIDATES = (
    os.path.join(os.path.dirname(__file__), "..", "..", "..", "seeds", "rcn_price_reference_seed_v2.xlsx"),
    os.path.join(os.path.dirname(__file__), "..", "..", "..", "..", "seeds", "rcn_price_reference_seed_v2.xlsx"),
)


def _resolve_seed_path() -> str | None:
    """Return the first candidate seed path that exists, or None."""
    env = os.environ.get("RCN_SEED_PATH")
    if env and os.path.isfile(env):
        return env
    for candidate in _TANK_SEED_PATH_CANDIDATES:
        if os.path.isfile(candidate):
            return candidate
    return None


def _load_tank_seed_brackets() -> list[tuple[float, float]]:
    """Read the Static Equipment sheet's tank rows: returns sorted [(bbl, mid_rcn)] tuples.

    The seed lives in `seeds/rcn_price_reference_seed_v2.xlsx::Static Equipment`.
    Rows look like:
        ('tank:prod:100bbl:std', 'prod_tank', 'Various', 'Production Tank',
         'standard', '100 BBL', '100 BBL', 'BBL', 8000, 12000, 16000, '$/BBL', ...)
    Index 0 = id, 5 = spec_low, 8/9/10 = rcn low/mid/high.

    Returns an empty list on any read failure — caller falls back to flat $50k.
    """
    seed_path = _resolve_seed_path()
    if seed_path is None:
        return []
    try:
        import openpyxl  # local import — keeps engine import cheap if seed never queried
        wb = openpyxl.load_workbook(seed_path, read_only=True, data_only=True)
        if "Static Equipment" not in wb.sheetnames:
            return []
        ws = wb["Static Equipment"]
        brackets: list[tuple[float, float]] = []
        for row in ws.iter_rows(values_only=True):
            if not row or not row[0]:
                continue
            key = str(row[0]).strip().lower()
            if not key.startswith("tank:prod:"):
                continue
            # Spec column 5 = "100 BBL" / "200 BBL" / etc.
            spec = str(row[5] or "").strip()
            mid_rcn = row[9]
            if not spec or mid_rcn is None:
                continue
            # Extract numeric BBL from spec text.
            digits = "".join(c for c in spec if c.isdigit() or c == ".")
            if not digits:
                continue
            try:
                bbl = float(digits)
                mid = float(mid_rcn)
            except (TypeError, ValueError):
                continue
            brackets.append((bbl, mid))
        wb.close()
        brackets.sort(key=lambda t: t[0])
        return brackets
    except Exception:
        # Any I/O / parsing failure → empty ladder → engine falls back to flat $50k.
        return []


def _get_tank_seed_brackets() -> list[tuple[float, float]]:
    """Thread-safe lazy accessor for the tank $/BBL ladder."""
    global _TANK_SEED_CACHE
    if _TANK_SEED_CACHE is not None:
        return _TANK_SEED_CACHE
    with _TANK_SEED_LOCK:
        if _TANK_SEED_CACHE is None:
            _TANK_SEED_CACHE = _load_tank_seed_brackets()
    return _TANK_SEED_CACHE


def _reset_tank_seed_cache_for_tests() -> None:
    """Test hook — clear the cache so a test can override RCN_SEED_PATH."""
    global _TANK_SEED_CACHE
    with _TANK_SEED_LOCK:
        _TANK_SEED_CACHE = None


def _tank_rcn_from_bbl(volume_bbl: float, brackets: list[tuple[float, float]]) -> float | None:
    """Return the bracket-anchored RCN for a given tank volume.

    Lookup rule:
      - At or below the smallest bracket → smallest bracket's mid RCN.
      - At or above the largest bracket → linear extrapolation at the tail $/BBL rate
        (mid_largest / bbl_largest), so a 5000 BBL doesn't snap to the 750 BBL price.
      - Between two brackets → linear interpolation on (bbl, rcn).

    Returns None if the bracket ladder is empty (caller falls back to flat scalar).
    """
    if not brackets or volume_bbl <= 0:
        return None
    if volume_bbl <= brackets[0][0]:
        return brackets[0][1]
    if volume_bbl >= brackets[-1][0]:
        tail_bbl, tail_rcn = brackets[-1]
        per_bbl = tail_rcn / tail_bbl
        return per_bbl * volume_bbl
    # Linear interp between adjacent brackets.
    for i in range(len(brackets) - 1):
        lo_bbl, lo_rcn = brackets[i]
        hi_bbl, hi_rcn = brackets[i + 1]
        if lo_bbl <= volume_bbl <= hi_bbl:
            t = (volume_bbl - lo_bbl) / (hi_bbl - lo_bbl)
            return lo_rcn + t * (hi_rcn - lo_rcn)
    return None  # unreachable given the bounds above


def normalize_category(category: str) -> str:
    """Normalize a category string to its base-table key."""
    normalized = category.strip().lower().replace("-", "_").replace(" ", "_")
    return CATEGORY_BASE_MAP.get(normalized, normalized)


def compute_spec_modifiers_factor(
    modifiers: dict[str, float] | list[float] | float | None,
) -> float:
    if modifiers is None:
        return 1.0
    factor = 1.0
    if isinstance(modifiers, dict):
        for value in modifiers.values():
            if value > 0:
                factor *= value
        return factor
    if isinstance(modifiers, list):
        for value in modifiers:
            if value > 0:
                factor *= value
        return factor
    return modifiers if modifiers > 0 else 1.0


def compute_base_rcn(
    category_key: str,
    specs: RCNInput,
) -> tuple[float, float, bool]:
    """Return (scaled_base_rcn, rcn_source_quality, has_size_param)."""
    horsepower = specs.horsepower
    weight_lbs = specs.weight_lbs
    truck_class = (specs.truck_class or "class_8").strip().lower().replace(" ", "_")
    equipment_type = (specs.equipment_type or "").strip().lower().replace(" ", "_")

    if category_key in ROTATING_BASE_RCN:
        type_map = ROTATING_BASE_RCN[category_key]
        resolved_type = equipment_type if equipment_type in type_map else DEFAULT_ROTATING_TYPE[category_key]
        params = type_map[resolved_type]
        base_rcn = float(params["base_rcn"])
        base_hp = float(params["base_hp"])
        if horsepower is not None and horsepower > 0:
            scaling = (horsepower / base_hp) ** HP_SCALING_EXPONENT
            quality = 0.80 if equipment_type in type_map else 0.70
            return base_rcn * scaling, quality, True
        quality = 0.60 if equipment_type in type_map else 0.50
        return base_rcn, quality, False

    if category_key in STATIC_BASE_RCN:
        base_rcn = float(STATIC_BASE_RCN[category_key])
        # Tier 2.5: tanks with parseable BBL route through the seed's $/BBL ladder
        # (not the flat $50k scalar). Falls back to flat scalar when volume is
        # missing OR the seed couldn't be loaded.
        if category_key == "tank" and specs.volume_bbl is not None and specs.volume_bbl > 0:
            bracket_rcn = _tank_rcn_from_bbl(specs.volume_bbl, _get_tank_seed_brackets())
            if bracket_rcn is not None and bracket_rcn > 0:
                return float(bracket_rcn), 0.75, True
        if weight_lbs is not None and weight_lbs > 0:
            scaling = (weight_lbs / WEIGHT_SCALING_BASE_LBS) ** WEIGHT_SCALING_EXPONENT
            return base_rcn * scaling, 0.70, True
        return base_rcn, 0.60, False

    if category_key in HEAVY_EQUIP_BASE_RCN:
        return float(HEAVY_EQUIP_BASE_RCN[category_key]), 0.70, False

    if category_key == "truck":
        return float(TRUCK_BASE_RCN.get(truck_class, TRUCK_BASE_RCN["class_8"])), 0.70, False

    return DEFAULT_BASE_RCN, DEFAULT_UNKNOWN_RCN_CONFIDENCE, False
