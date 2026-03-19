from __future__ import annotations
import os
import openpyxl
from sqlalchemy import text
from app.db.session import get_session


async def search_comparables(keywords: list[str], category: str | None = None,
                             price_min: float = 0, price_max: float = 99999999,
                             max_results: int = 20) -> str:
    if not keywords:
        return "No keywords provided."
    where = " OR ".join(f"title ILIKE :kw{i}" for i in range(len(keywords)))
    params = {f"kw{i}": f"%{kw}%" for i, kw in enumerate(keywords)}
    sql = f"SELECT title, asking_price, currency, source, location, year, hours, url FROM listings WHERE ({where}) AND asking_price IS NOT NULL AND asking_price BETWEEN :pmin AND :pmax"
    params["pmin"], params["pmax"] = price_min, price_max
    if category:
        sql += " AND category_normalized ILIKE :cat"
        params["cat"] = f"%{category}%"
    sql += f" ORDER BY asking_price DESC LIMIT :lim"
    params["lim"] = max_results
    async with get_session() as session:
        result = await session.execute(text(sql), params)
        rows = result.fetchall()
    if not rows:
        return f"No comparables found for {keywords}. Normal for high-spec equipment."
    lines = [f"Found {len(rows)} comparables:"]
    for r in rows:
        yr = f"{r.year} " if r.year else ""
        hrs = f" ({r.hours}hrs)" if r.hours else ""
        loc = f" — {r.location}" if r.location else ""
        lines.append(f"  {yr}{r.title} | ${r.asking_price:,.0f} {r.currency or 'CAD'}{hrs}{loc} | {r.source} | {r.url or 'no link'}")
    return "\n".join(lines)


async def get_category_stats(category: str) -> str:
    sql = """SELECT COUNT(*) as total, COUNT(CASE WHEN asking_price > 0 THEN 1 END) as with_price,
             AVG(CASE WHEN asking_price > 0 THEN asking_price END) as avg_price,
             MIN(CASE WHEN asking_price > 0 THEN asking_price END) as min_price,
             MAX(CASE WHEN asking_price > 0 THEN asking_price END) as max_price
             FROM listings WHERE category_normalized ILIKE :cat"""
    async with get_session() as session:
        result = await session.execute(text(sql), {"cat": f"%{category}%"})
        r = result.fetchone()
    if not r or r.total == 0:
        return f"No listings found for category '{category}'."
    return (f"Category '{category}': {r.total} listings ({r.with_price} with price)\n"
            f"  Avg: ${r.avg_price:,.0f}  Min: ${r.min_price:,.0f}  Max: ${r.max_price:,.0f}")


_RCN_ROWS: list[dict] | None = None
_XLSX_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "..", "seeds", "rcn_price_reference_seed_v2.xlsx")
_DATA_TABS = ["Compressors", "Drivers", "Generators", "Static Equipment", "Pumps & Pumpjacks"]
_EVIDENCE_TAB = "PDF Pricing Evidence Intake"

_ESCALATION = {2020: 1.40, 2021: 1.30, 2022: 1.16, 2023: 1.08, 2024: 1.04, 2025: 1.02, 2026: 1.00}
_USD_TO_CAD = 1.44


def _load_rcn_data() -> list[dict]:
    global _RCN_ROWS
    if _RCN_ROWS is not None:
        return _RCN_ROWS
    wb = openpyxl.load_workbook(_XLSX_PATH, read_only=True, data_only=True)
    rows = []

    # ── Reference tabs ────────────────────────────────
    for tab in _DATA_TABS:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        headers = None
        for row in ws.iter_rows(values_only=True):
            if row and str(row[0] or "").strip().lower() == "reference key":
                headers = [str(c or "").strip().lower().replace(" ", "_").replace("(", "").replace(")", "").replace("/", "_") for c in row]
                continue
            if headers is None or not row or not row[0]:
                continue
            # Skip section-header rows (only first cell has text, rest are None)
            if all(c is None for c in row[1:6]):
                continue
            d = dict(zip(headers, row))
            d["_tab"] = tab
            d["_source"] = "reference_table"
            # Build searchable text from key columns
            searchable = " ".join(str(d.get(k) or "") for k in
                ["reference_key", "manufacturer", "model_frame", "model", "type_model", "configuration", "notes"])
            d["_search"] = searchable.lower()
            rows.append(d)

    # ── Appraisal evidence tab ────────────────────────
    if _EVIDENCE_TAB in wb.sheetnames:
        ws = wb[_EVIDENCE_TAB]
        ev_headers = None
        for row in ws.iter_rows(values_only=True):
            vals = list(row)
            if ev_headers is None:
                ev_headers = [str(c or "").strip().lower().replace(" ", "_") for c in vals]
                continue
            if not vals or not vals[0]:
                continue
            d = dict(zip(ev_headers, vals))
            # Only load RCN rows
            if str(d.get("price_type") or "").upper() != "RCN":
                continue
            price = d.get("price_value")
            if not price or not isinstance(price, (int, float)) or price <= 0:
                continue

            # Currency conversion
            rcn_cad = float(price)
            if str(d.get("currency") or "").upper() == "USD":
                rcn_cad *= _USD_TO_CAD

            # Escalation by effective_date year
            eff_date = d.get("effective_date")
            year = None
            if hasattr(eff_date, "year"):
                year = eff_date.year
            elif isinstance(eff_date, str) and len(eff_date) >= 4:
                try:
                    year = int(eff_date[:4])
                except ValueError:
                    pass
            if year and year in _ESCALATION:
                rcn_cad *= _ESCALATION[year]

            rcn_mid = round(rcn_cad)
            rcn_low = round(rcn_cad * 0.85)
            rcn_high = round(rcn_cad * 1.15)

            # Build searchable text
            searchable = " ".join(str(d.get(k) or "") for k in
                ["manufacturer", "model_frame", "description", "equipment_category", "notes"])

            ref_key = f"ev:{str(d.get('asset_id') or '')}:{str(d.get('manufacturer') or '')[:10]}"
            source_label = str(d.get("source_file") or "")[:60]
            appraiser = str(d.get("appraiser") or "")
            eff_str = str(eff_date or "")[:10]

            rows.append({
                "reference_key": ref_key,
                "manufacturer": d.get("manufacturer"),
                "model_frame": d.get("model_frame"),
                "description": d.get("description"),
                "equipment_category": d.get("equipment_category"),
                "rcn_low": rcn_low,
                "rcn_mid": rcn_mid,
                "rcn_high": rcn_high,
                "confidence": d.get("confidence", "—"),
                "notes": f"Appraisal by {appraiser}, {eff_str}. Source: {source_label}",
                "effective_date": eff_str,
                "original_currency": d.get("currency"),
                "original_value": price,
                "_tab": "Appraisal Evidence",
                "_source": "appraisal_evidence",
                "_search": searchable.lower(),
            })

    wb.close()
    _RCN_ROWS = rows
    return _RCN_ROWS


def _score_match(row_search: str, terms: list[str]) -> int:
    return sum(1 for t in terms if t in row_search)


_FALLBACK_RCN = {
    "vru": {
        "match_terms": ["vru", "vapor recovery", "vapour recovery", "rotary vane"],
        "rcn_low": 200000, "rcn_mid": 240000, "rcn_high": 280000,
        "note": "Complete VRU package 30-50HP class with shelter, condensate handling, PLC, NACE piping.",
    },
    "incinerator": {
        "match_terms": ["incinerator", "enclosed combustor", "ecd"],
        "rcn_low": 50000, "rcn_mid": 75000, "rcn_high": 100000,
        "note": "Enclosed combustion device 48-60 inch.",
    },
    "flare": {
        "match_terms": ["flare stack", "flare knockout"],
        "rcn_low": 15000, "rcn_mid": 25000, "rcn_high": 40000,
        "note": "Flare knockout + stack, size dependent.",
    },
    "line_heater": {
        "match_terms": ["line heater", "gas production unit"],
        "rcn_low": 30000, "rcn_mid": 50000, "rcn_high": 80000,
        "note": "Line heater 0.5-1.5 MMBTU with separator package.",
    },
    "meter_skid": {
        "match_terms": ["meter skid", "meter run", "meter building"],
        "rcn_low": 8000, "rcn_mid": 15000, "rcn_high": 25000,
        "note": "Meter skid/run 2-4 inch ANSI 600.",
    },
    "treater": {
        "match_terms": ["treater", "heater treater", "emulsion treater"],
        "rcn_low": 80000, "rcn_mid": 150000, "rcn_high": 250000,
        "note": "Heater treater 48-72 inch.",
    },
}


def _check_fallback(query: str) -> str | None:
    q = query.lower()
    for key, d in _FALLBACK_RCN.items():
        if any(term in q for term in d["match_terms"]):
            return (f"Found 1 RCN match(es):\n\n  [fallback:{key}] (Fallback)\n"
                    f"  RCN: ${d['rcn_low']:,} — ${d['rcn_mid']:,} — ${d['rcn_high']:,} CAD\n"
                    f"  Notes: {d['note']} (fallback - not yet in reference spreadsheet)\n"
                    f"\nApply depreciation factors for age, condition, hours.")
    return None


def lookup_rcn(equipment_type: str, manufacturer: str | None = None, model: str | None = None,
               drive_type: str | None = None, stages: int | None = None, hp: int | None = None) -> str:
    query = " ".join(p.lower() for p in [equipment_type, manufacturer or "", model or ""] if p)
    # Check fallback first — specific match_terms are more reliable than weak fuzzy hits
    fb = _check_fallback(query)
    if fb:
        return fb
    data = _load_rcn_data()
    terms = [t for t in query.split() if len(t) > 1]
    scored = [(r, _score_match(r["_search"], terms)) for r in data]
    matches = sorted([(r, s) for r, s in scored if s >= max(1, len(terms) // 2)], key=lambda x: -x[1])[:3]
    if not matches:
        return ("No RCN anchor found in reference tables. Estimate from general knowledge. "
                "HP scaling: base × (target_hp / base_hp) ^ 0.6")
    lines = [f"Found {len(matches)} RCN match(es):"]
    for r, score in matches:
        ref = r.get("reference_key", "?")
        source = r.get("_source", "reference_table")
        source_label = "Reference Table" if source == "reference_table" else "Appraisal Evidence"
        low = r.get("rcn_new_low_cad") or r.get("rcn_low") or 0
        mid = r.get("rcn_new_mid_cad") or r.get("rcn_mid") or 0
        high = r.get("rcn_new_high_cad") or r.get("rcn_high") or 0
        exp = r.get("scaling_exp") or r.get("scaling_method") or "0.6"
        base_hp = r.get("base_hp", "—")
        conf = r.get("confidence", "—")
        notes = r.get("notes", "")
        lines.append(f"\n  [{ref}] ({r['_tab']} — {source_label})")
        lines.append(f"  RCN: ${low:,.0f} — ${mid:,.0f} — ${high:,.0f} CAD")
        if base_hp and base_hp != "—" and source == "reference_table":
            lines.append(f"  Base HP: {base_hp}, Scaling exponent: {exp}")
        lines.append(f"  Confidence: {conf}")
        if notes:
            lines.append(f"  Notes: {notes}")
    lines.append("\nApply depreciation factors for age, condition, hours.")
    return "\n".join(lines)


AGE_CURVES = {
    "rotating":    [(1,0.90),(3,0.80),(5,0.65),(8,0.50),(12,0.38),(15,0.28),(20,0.20),(25,0.15),(30,0.10),(999,0.08)],
    "static":      [(1,0.92),(3,0.85),(5,0.75),(8,0.65),(12,0.55),(15,0.45),(20,0.35),(25,0.25),(30,0.18),(999,0.12)],
    "pump_jack":   [(1,0.92),(3,0.85),(5,0.78),(8,0.70),(12,0.60),(15,0.50),(20,0.40),(25,0.30),(30,0.22),(999,0.15)],
    "electrical":  [(1,0.90),(3,0.82),(5,0.72),(8,0.60),(12,0.48),(15,0.38),(20,0.28),(25,0.20),(30,0.14),(999,0.10)],
}
CONDITION = {"A": 1.00, "B": 0.75, "C": 0.50, "D": 0.20}
HOURS_FACTORS = [(5000,1.10),(15000,1.00),(30000,0.85),(50000,0.70),(999999,0.55)]
SERVICE_FACTORS = {"sweet": 1.00, "sour": 1.15, "sour_high_h2s": 1.25}


def _lookup(table, value):
    for threshold, factor in table:
        if value <= threshold:
            return factor
    return table[-1][1]


def calculate_fmv(rcn: float, equipment_class: str, age_years: int, condition: str = "B",
                  hours: int | None = None, service: str = "sweet",
                  vfd_equipped: bool = False, turnkey_package: bool = False,
                  nace_rated: bool = False) -> str:
    curve = AGE_CURVES.get(equipment_class, AGE_CURVES["rotating"])
    age_f = _lookup(curve, age_years)
    cond_f = CONDITION.get(condition.upper(), 0.75)
    hrs_f = _lookup(HOURS_FACTORS, hours) if hours is not None and equipment_class == "rotating" else 1.00
    svc_f = SERVICE_FACTORS.get(service, 1.00)
    prem = (1.05 if vfd_equipped else 1.0) * (1.05 if turnkey_package else 1.0) * (1.15 if nace_rated else 1.0)
    combined = age_f * cond_f * hrs_f * svc_f * prem
    mid = rcn * combined
    low, high = mid * 0.85, mid * 1.15
    list_price, walkaway = mid * 1.12, low * 0.92
    factors = f"Age({age_years}yr)={age_f:.2f} × Cond({condition})={cond_f:.2f}"
    if hours is not None and equipment_class == "rotating":
        factors += f" × Hours({hours:,})={hrs_f:.2f}"
    factors += f" × Service({service})={svc_f:.2f}"
    if prem != 1.0:
        factors += f" × Premiums={prem:.2f}"
    return (f"FMV CALCULATION:\n"
            f"  RCN: ${rcn:,.0f}\n"
            f"  Factors: {factors}\n"
            f"  Combined factor: {combined:.4f}\n"
            f"  FMV Range: ${low:,.0f} — ${mid:,.0f} — ${high:,.0f}\n"
            f"  Recommended list price: ${list_price:,.0f}\n"
            f"  Walk-away floor: ${walkaway:,.0f}")


def check_equipment_risks(equipment_type: str, age_years: int, hours: int | None = None,
                          idle_years: int | None = None, drive_type: str | None = None,
                          plc_model: str | None = None, manufacturer: str | None = None,
                          location_country: str = "CA", identical_units: int = 1,
                          days_on_market: int | None = None, total_views: int | None = None) -> str:
    risks = []
    if idle_years and idle_years > 5:
        risks.append(f"IDLE DEGRADATION ({idle_years}yr idle): Elastomer/seal degradation likely. Pre-commissioning inspection required — budget $3,000–$8,000.")
    if idle_years and idle_years > 3 and "rotary vane" in (equipment_type or "").lower():
        risks.append("VANE MOISTURE DAMAGE: Rotary vane units idle >3yr risk vane swelling/moisture damage. Inspection mandatory before valuation.")
    if plc_model and "micrologix" in plc_model.lower():
        risks.append(f"PLC OBSOLESCENCE: {plc_model} discontinued 2017. Budget $8,000–$15,000 for controls upgrade (CompactLogix or Micro800).")
    if age_years > 10:
        risks.append(f"CONTROLS CHECK: At {age_years}yr, verify PLC/HMI platform is still supported and parts are available.")
    if location_country and location_country.upper() != "CA":
        risks.append(f"CROSS-BORDER: Equipment in {location_country}. Transport $5,000–$15,000 + ASME→ABSA re-certification $2,000–$5,000.")
    if identical_units > 10:
        risks.append(f"OVERSUPPLY: {identical_units} identical units — recommend lot sale strategy to avoid cannibalizing individual pricing.")
    elif identical_units > 3:
        risks.append(f"VOLUME DISCOUNT: {identical_units} identical units — expect 5–10% volume discount pressure.")
    if days_on_market and total_views and days_on_market > 730 and total_views > 500:
        risks.append(f"STALE LISTING: {days_on_market} days on market with {total_views} views — price resistance confirmed. Recommend 15–20% reduction.")
    mfr = (manufacturer or "").lower()
    dt = (drive_type or "").lower()
    if mfr in ["ajax", "cooper"] or dt == "integral":
        risks.append("DECLINING MARKET: Integral/slow-speed engines face shrinking buyer pool. Extended sale timeline 6–18 months.")
    if mfr in ["worthington", "energy industries", "dresser"]:
        risks.append(f"UNCOMMON FRAME: {manufacturer} — limited aftermarket parts, smaller buyer pool. Discount 10–20% vs common frames.")
    if not risks:
        return "No significant risk factors identified."
    return "RISK FACTORS:\n• " + "\n• ".join(risks)
