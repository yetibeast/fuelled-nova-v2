"""LLM-backed item extractor for the batch uploader.

Takes arbitrary file bytes (.xlsx, .csv, .eml) and returns a list of dicts
shaped like ``{"title": str, "category": str, "specs": dict}`` — the same
shape consumed by ``app.api.batch.BatchItem``. Used as a fallback when the
strict schema parser in ``batch.py`` can't recognise the file.
"""
from __future__ import annotations

import base64
import email
import json
import logging
import re
from email import policy
from io import BytesIO, StringIO

import anthropic

from app.config import ANTHROPIC_API_KEY

_log = logging.getLogger(__name__)
_client = anthropic.AsyncAnthropic(api_key=ANTHROPIC_API_KEY)
_MODEL = "claude-sonnet-4-20250514"
_MAX_CHARS = 180_000  # ~45k tokens of input; truncate beyond
_MAX_ITEMS = 2000  # hard safety cap on extracted items
_MAX_OUTPUT_TOKENS = 64_000  # Sonnet 4 ceiling; enough for ~1500 compact items


SYSTEM_PROMPT = """\
You extract equipment line items from user-uploaded files for a batch equipment-\
valuation tool. Users upload spreadsheets, emails, and inventory dumps in wildly \
different shapes (row-per-item, pivoted matrix, free text). You normalise them \
into a flat list of items to price.

Rules:
1. Each item has: title (concise natural-language like "Corlac 1000bbl oil tank w/ Profire 2100 BMS (2007)"), category (one of: tank, vessel, compressor, pump, separator, line_heater, drive_head, scrubber, regulator, treater, generator, meter_skid, building, other), specs (object with optional keys make, model, year, hp, size, location, serial, notes).
2. For a pivoted matrix (equipment-attribute rows × site/location columns), emit ONE item per populated (equipment-unit × site) cell. Roll up the attribute rows that belong to the same equipment unit (e.g. all "Tank 1" sub-rows) into that item's specs and title. Use the site/well code as specs.location.
3. Skip cells that are blank or contain only: NA, N/A, None, null, TBD, -.
4. Preserve identifying info verbatim where it exists (model numbers, serials, years, dimensions, manufacturers).
5. Output MUST be JSONL: ONE complete JSON object per line, nothing else. No prose, no code fences, no wrapping array, no explanation. Each line is `{"title":"...","category":"...","specs":{...}}` and ends with a newline. If your output is cut off mid-line, the final partial line is discarded; write the most important items first.
"""


async def extract_items(
    data: bytes,
    filename: str,
    content_type: str | None = None,
) -> list[dict]:
    """Extract equipment items from an uploaded file via Claude.

    Returns a list of plain dicts with keys ``title``, ``category``, ``specs``.
    Raises ``ValueError`` on unreadable input or malformed LLM response.
    """
    ext = (filename or "").rsplit(".", 1)[-1].lower()

    if ext in ("xlsx", "xls"):
        text_body = _xlsx_to_text(data)
    elif ext == "csv":
        text_body = data.decode("utf-8-sig", errors="replace")
    elif ext == "eml":
        text_body = _eml_to_text(data)
    else:
        raise ValueError(f"Unsupported file type for LLM extraction: .{ext}")

    if not text_body.strip():
        raise ValueError("File is empty or unreadable")

    truncated = False
    if len(text_body) > _MAX_CHARS:
        text_body = text_body[:_MAX_CHARS]
        truncated = True
        _log.info("batch_extractor: truncated %s to %d chars", filename, _MAX_CHARS)

    user_msg = (
        f"File: {filename}\n"
        + ("[NOTE: input was truncated; extract what you can see]\n" if truncated else "")
        + f"\nContents:\n{text_body}"
    )

    # Streaming is required for large max_tokens (SDK enforces this above ~10min cap).
    text_chunks: list[str] = []
    async with _client.messages.stream(
        model=_MODEL,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": user_msg}],
        max_tokens=_MAX_OUTPUT_TOKENS,
    ) as stream:
        async for chunk in stream.text_stream:
            text_chunks.append(chunk)
        final = await stream.get_final_message()

    text = "".join(text_chunks)
    stop_reason = getattr(final, "stop_reason", None)
    items = _parse_items(text, truncated_output=stop_reason == "max_tokens")

    if not items:
        raise ValueError("No items extracted from the file")
    if len(items) > _MAX_ITEMS:
        _log.warning("batch_extractor: capping %d extracted items at %d", len(items), _MAX_ITEMS)
        items = items[:_MAX_ITEMS]
    return items


# ── File-format → plain text ─────────────────────────────────────────────
def _xlsx_to_text(data: bytes) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    chunks: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        chunks.append(f"=== SHEET: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            # Strip trailing empties for compactness
            while cells and not cells[-1]:
                cells.pop()
            if cells:
                chunks.append("\t".join(cells))
        chunks.append("")
    wb.close()
    return "\n".join(chunks)


def _eml_to_text(data: bytes) -> str:
    msg = email.message_from_bytes(data, policy=policy.default)
    chunks: list[str] = []
    subject = msg.get("Subject", "")
    sender = msg.get("From", "")
    if subject or sender:
        chunks.append(f"SUBJECT: {subject}\nFROM: {sender}\n")

    for part in msg.walk():
        ctype = part.get_content_type()
        fname = part.get_filename() or ""

        if fname.lower().endswith((".xlsx", ".xls")):
            try:
                payload = part.get_payload(decode=True)
                chunks.append(f"=== ATTACHMENT: {fname} ===")
                chunks.append(_xlsx_to_text(payload))
            except Exception as e:
                _log.warning("failed to read xlsx attachment %s: %s", fname, e)
        elif fname.lower().endswith(".csv"):
            try:
                payload = part.get_payload(decode=True)
                chunks.append(f"=== ATTACHMENT: {fname} ===")
                chunks.append(payload.decode("utf-8-sig", errors="replace"))
            except Exception as e:
                _log.warning("failed to read csv attachment %s: %s", fname, e)
        elif ctype == "text/plain" and not fname:
            body = part.get_content()
            if body.strip():
                chunks.append("=== EMAIL BODY ===")
                chunks.append(body)

    return "\n".join(chunks)


# ── LLM response → items list ────────────────────────────────────────────
def _parse_items(text: str, truncated_output: bool = False) -> list[dict]:
    """Parse LLM output into items.

    Primary path: JSONL, one object per line. A partial/malformed trailing line
    is skipped (survives output truncation).

    Fallback: if the model returned a JSON array, parse that too.
    """
    text = re.sub(r"^```(?:json|jsonl)?\s*|\s*```$", "", text.strip(), flags=re.MULTILINE)
    if not text:
        return []

    items: list[dict] = []
    skipped = 0

    stripped = text.lstrip()
    if stripped.startswith("["):
        # Whole-array fallback. If Claude slipped into array mode anyway, salvage it.
        try:
            data = json.loads(stripped)
            if isinstance(data, list):
                for raw in data:
                    norm = _normalize_item(raw)
                    if norm:
                        items.append(norm)
                if items:
                    return items
        except json.JSONDecodeError:
            pass  # fall through to line-by-line salvage

    for line in text.splitlines():
        line = line.strip().rstrip(",")
        if not line or line in ("[", "]"):
            continue
        try:
            obj = json.loads(line)
        except json.JSONDecodeError:
            skipped += 1
            continue
        norm = _normalize_item(obj)
        if norm:
            items.append(norm)

    if skipped:
        level = logging.WARNING if truncated_output else logging.INFO
        _log.log(level, "batch_extractor: skipped %d malformed JSONL lines (truncated=%s)", skipped, truncated_output)
    if truncated_output:
        _log.warning("batch_extractor: LLM hit max_tokens; returning %d items parsed before truncation", len(items))

    return items


def _normalize_item(raw) -> dict | None:
    if not isinstance(raw, dict):
        return None
    title = str(raw.get("title") or "").strip()
    if not title:
        return None
    category = str(raw.get("category") or "").strip()
    specs = raw.get("specs") or {}
    if not isinstance(specs, dict):
        specs = {}
    return {"title": title, "category": category, "specs": specs}
