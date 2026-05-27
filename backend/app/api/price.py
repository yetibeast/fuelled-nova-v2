from __future__ import annotations
import base64
import csv
import io
import json
import logging
from fastapi import APIRouter, Form, File, Header, HTTPException, UploadFile
from fastapi.responses import Response
from app.pricing_v2.service import run_pricing
from app.pricing_v2.report import generate_report
from app.api.admin import _require_auth, _require_auth_identity

router = APIRouter()
_log = logging.getLogger(__name__)

# Anthropic's `document` content block only accepts application/pdf. Anything
# else sent that way returns BadRequestError → 500. Images get the `image`
# content block. Spreadsheets get converted to CSV text and prepended to the
# user message. Anything else gets rejected at the door, not silently passed
# as a "document".
MEDIA_MAP = {
    "application/pdf": "document",
    "image/png": "image",
    "image/jpeg": "image",
    "image/webp": "image",
}

SPREADSHEET_MIMES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
    "application/vnd.ms-excel",                                            # .xls
    "application/vnd.ms-excel.sheet.macroenabled.12",                      # .xlsm
}


def _xlsx_to_text(content: bytes, filename: str) -> str:
    """Render a spreadsheet upload as one CSV block per sheet, suitable for
    appending to a user message as plain text. Captures values only — formulas
    are resolved via openpyxl data_only mode.
    """
    import openpyxl  # local import — keeps top-level import fast
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    parts: list[str] = [f"[Spreadsheet attached: {filename}]"]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        buf = io.StringIO()
        writer = csv.writer(buf)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if c is None else c for c in row])
        rendered = buf.getvalue().rstrip()
        if rendered:
            parts.append(f"\n=== Sheet: {sheet_name} ===\n{rendered}")
    return "\n".join(parts)


@router.post("/price")
async def post_price(
    message: str = Form(...),
    files: list[UploadFile] = File(default=[]),
    history: str = Form(default=""),
    conversation_id: str = Form(default=""),
    authorization: str = Header(default=""),
):
    user_id, user_email = _require_auth_identity(authorization)
    attachments = []
    spreadsheet_blocks: list[str] = []
    for f in files:
        content = await f.read(10_485_761)
        if len(content) > 10_485_760:
            raise HTTPException(413, "File too large (10MB max)")
        media_type = f.content_type or "application/octet-stream"

        if media_type in SPREADSHEET_MIMES:
            try:
                spreadsheet_blocks.append(_xlsx_to_text(content, f.filename or "spreadsheet.xlsx"))
            except Exception as e:
                raise HTTPException(
                    400,
                    f"Could not read spreadsheet '{f.filename}': {e}. "
                    "Make sure the file is a valid .xlsx, .xls, or .xlsm.",
                )
            continue

        if media_type not in MEDIA_MAP:
            raise HTTPException(
                415,
                f"Unsupported file type '{media_type}' for '{f.filename}'. "
                "Supported: PDF, PNG/JPEG/WEBP images, and Excel spreadsheets.",
            )

        data = base64.b64encode(content).decode()
        att_type = MEDIA_MAP[media_type]
        attachments.append({"type": att_type, "media_type": media_type, "data": data})

    # Prepend any spreadsheet content to the user message so it lands in the
    # text content block — Anthropic's document block doesn't accept xlsx.
    if spreadsheet_blocks:
        message = "\n\n".join(spreadsheet_blocks) + "\n\n" + message

    conversation_history = None
    if history:
        try:
            conversation_history = json.loads(history)
        except json.JSONDecodeError:
            pass

    result = await run_pricing(
        message,
        attachments if attachments else None,
        conversation_history,
        user_id=user_id,
        user_email=user_email,
        conversation_id=conversation_id or None,
    )

    # Auto-capture evidence (fire-and-forget)
    try:
        from app.api.evidence import capture_evidence
        await capture_evidence(
            {
                "user_message": message,
                "structured_data": result.get("structured", {}),
                "confidence": result.get("confidence", "LOW"),
                "tools_used": result.get("tools_used", []),
            },
            authorization=authorization,
        )
    except Exception as e:
        _log.warning("Evidence capture failed: %s", e)

    return result


@router.post("/report")
async def post_report(
    structured_data: str = Form(...),
    response_text: str = Form(default=""),
    user_message: str = Form(default="Equipment Valuation"),
    authorization: str = Header(default=""),
):
    _require_auth(authorization)
    try:
        structured = json.loads(structured_data)
    except json.JSONDecodeError:
        raise HTTPException(400, "Invalid JSON in structured_data")

    docx_bytes = generate_report(
        structured=structured,
        response_text=response_text,
        user_message=user_message,
    )

    return Response(
        content=docx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": "attachment; filename=Fuelled_Valuation_Report.docx"},
    )
