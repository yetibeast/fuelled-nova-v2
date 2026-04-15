"""Fuelled pricing-coverage analytics endpoint."""
from __future__ import annotations

import asyncio
import io
import logging
import uuid
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Body, Header
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import text

from app.api.admin import _require_auth
from app.api.calibration import _require_admin
from app.db.session import get_session
from app.pricing_v2.service import run_pricing

router = APIRouter(tags=["fuelled-coverage"])
_log = logging.getLogger(__name__)

# Table-creation guard (run once per process)
_table_init = False

# ── Canonical SQL fragments ───────────────────────────────────────────────

_BASE = "source = 'fuelled' AND is_active = true"
_HAS_ASKING = "asking_price > 0"
_HAS_VALUE = "(asking_price > 0 OR fair_value > 0)"
_AI_ONLY = "(fair_value > 0 AND NOT asking_price > 0)"
_TIER = """CASE
    WHEN make IS NOT NULL AND model IS NOT NULL AND year IS NOT NULL THEN 1
    WHEN make IS NOT NULL AND year IS NOT NULL THEN 2
    WHEN make IS NOT NULL THEN 3
    ELSE 4
END"""
_COMPLETENESS = """(
    CASE WHEN make IS NOT NULL THEN 25 ELSE 0 END +
    CASE WHEN model IS NOT NULL THEN 20 ELSE 0 END +
    CASE WHEN year IS NOT NULL THEN 25 ELSE 0 END +
    CASE WHEN hours IS NOT NULL THEN 15 ELSE 0 END +
    CASE WHEN horsepower IS NOT NULL THEN 10 ELSE 0 END +
    CASE WHEN condition IS NOT NULL THEN 5 ELSE 0 END
)"""


_VALUATIONS_DDL = """
CREATE TABLE IF NOT EXISTS fuelled_valuations (
    id UUID PRIMARY KEY,
    listing_id UUID NOT NULL,
    fmv_low DOUBLE PRECISION,
    fmv_mid DOUBLE PRECISION,
    fmv_high DOUBLE PRECISION,
    confidence VARCHAR(10),
    status VARCHAR(10) NOT NULL DEFAULT 'success',
    error_reason TEXT,
    tier INTEGER,
    data_completeness INTEGER,
    tools_used TEXT,
    trace_id VARCHAR(64),
    created_at TIMESTAMPTZ DEFAULT NOW()
)
"""

# Module-level job state for batch pricing
_fuelled_job: dict | None = None


async def _ensure_table(session):
    """Create fuelled_valuations table if it doesn't exist yet."""
    global _table_init
    if _table_init:
        return
    await session.execute(text(_VALUATIONS_DDL))
    await session.commit()
    _table_init = True


# ---------------------------------------------------------------------------
# GET /admin/fuelled/coverage
# ---------------------------------------------------------------------------

@router.get("/admin/fuelled/coverage")
async def fuelled_coverage_stats(authorization: str = Header(None)):
    """Return coverage statistics for Fuelled-sourced listings."""
    _require_auth(authorization)

    async with get_session() as session:
        await _ensure_table(session)

        # Main stats
        result = await session.execute(text(f"""
            SELECT
                COUNT(*),
                COUNT(CASE WHEN {_HAS_ASKING} THEN 1 END),
                COUNT(CASE WHEN {_HAS_VALUE} THEN 1 END),
                COUNT(CASE WHEN {_AI_ONLY} THEN 1 END),
                AVG({_COMPLETENESS})
            FROM listings
            WHERE {_BASE}
        """))
        row = result.fetchone()
        total = row[0] or 0
        asking_ct = row[1] or 0
        valued_ct = row[2] or 0
        ai_only_ct = row[3] or 0
        completeness_avg = round(row[4] or 0, 1)

        # Tier breakdown (unvalued only)
        result = await session.execute(text(f"""
            SELECT {_TIER} AS tier, COUNT(*) AS cnt
            FROM listings
            WHERE {_BASE} AND NOT {_HAS_VALUE}
            GROUP BY tier
            ORDER BY tier
        """))
        tier_rows = result.fetchall()
        by_tier = {f"tier_{r[0]}": r[1] for r in tier_rows} if tier_rows else {}

        # Category breakdown (unvalued, top 10)
        result = await session.execute(text(f"""
            SELECT category, COUNT(*) AS cnt
            FROM listings
            WHERE {_BASE} AND NOT {_HAS_VALUE}
            GROUP BY category
            ORDER BY cnt DESC
            LIMIT 10
        """))
        cat_rows = result.fetchall()
        by_category = [{"category": r[0], "count": r[1]} for r in cat_rows] if cat_rows else []

    unpriced = total - valued_ct

    return {
        "total": total,
        "asking_price_count": asking_ct,
        "asking_price_pct": round(asking_ct / total * 100, 1) if total else 0,
        "valued_count": valued_ct,
        "valued_pct": round(valued_ct / total * 100, 1) if total else 0,
        "ai_only_count": ai_only_ct,
        "unpriced": unpriced,
        "by_tier": by_tier,
        "by_category": by_category,
        "completeness_avg": completeness_avg,
    }


# ── Report helpers (reusable by batch endpoint later) ────────────────────────

_FIELD_WEIGHTS = {
    "make": 25, "model": 20, "year": 25,
    "hours": 15, "horsepower": 10, "condition": 5,
}


def _completeness_and_missing(row) -> tuple[int, str]:
    """Return (score 0-100, comma-separated missing fields string)."""
    score = 0
    missing = []
    for field, weight in _FIELD_WEIGHTS.items():
        val = row[field] if isinstance(row, dict) else getattr(row, field, None)
        if val is not None:
            score += weight
        else:
            missing.append(field)
    return score, ", ".join(missing)


def _tier(row) -> int:
    """Return pricability tier 1-4 based on make/model/year presence."""
    def _get(f):
        return row[f] if isinstance(row, dict) else getattr(row, f, None)

    has_make = _get("make") is not None
    has_model = _get("model") is not None
    has_year = _get("year") is not None
    if has_make and has_model and has_year:
        return 1
    if has_make and has_year:
        return 2
    if has_make:
        return 3
    return 4


# ---------------------------------------------------------------------------
# POST /admin/fuelled/generate-report
# ---------------------------------------------------------------------------

_REPORT_COLUMNS = [
    "Title", "Category", "Make", "Model", "Year",
    "Condition", "Hours", "HP", "Data Completeness %",
    "Missing Fields", "Days Listed", "Pricability Tier", "URL",
]


@router.post("/admin/fuelled/generate-report")
async def fuelled_generate_report(authorization: str = Header(None)):
    """Generate an XLSX report of unpriced Fuelled listings."""
    _require_auth(authorization)

    async with get_session() as session:
        await _ensure_table(session)

        result = await session.execute(text(f"""
            SELECT title, category, make, model, year,
                   condition, hours, horsepower, url, first_seen
            FROM listings
            WHERE {_BASE} AND NOT {_HAS_VALUE}
            ORDER BY first_seen ASC
        """))
        rows = result.fetchall()

    now = datetime.now(timezone.utc)
    wb = Workbook()
    ws = wb.active
    ws.title = "Unpriced Fuelled Listings"

    # Header row with bold font
    bold = Font(bold=True)
    for col_idx, header in enumerate(_REPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        score, missing = _completeness_and_missing(row)
        tier = _tier(row)
        first_seen = row["first_seen"]
        days_listed = (now - first_seen).days if first_seen else None

        ws.cell(row=row_idx, column=1, value=row["title"])
        ws.cell(row=row_idx, column=2, value=row["category"])
        ws.cell(row=row_idx, column=3, value=row["make"])
        ws.cell(row=row_idx, column=4, value=row["model"])
        ws.cell(row=row_idx, column=5, value=row["year"])
        ws.cell(row=row_idx, column=6, value=row["condition"])
        ws.cell(row=row_idx, column=7, value=row["hours"])
        ws.cell(row=row_idx, column=8, value=row["horsepower"])
        ws.cell(row=row_idx, column=9, value=score)
        ws.cell(row=row_idx, column=10, value=missing)
        ws.cell(row=row_idx, column=11, value=days_listed)
        ws.cell(row=row_idx, column=12, value=tier)
        ws.cell(row=row_idx, column=13, value=row["url"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"fuelled_unpriced_{now.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Batch pricing helpers
# ---------------------------------------------------------------------------

def _build_pricing_query(row) -> str:
    """Build a directive prompt for a single listing that tells the agent NOT to ask follow-ups."""
    def _get(f):
        return row[f] if isinstance(row, dict) else getattr(row, f, None)

    parts = [
        "Provide a fair market value estimate for this equipment.",
        "Do not ask follow-up questions — use your best judgment with the data available.",
        "If critical data is missing, state your assumptions and provide a wider confidence range.",
        "",
    ]
    fields = [
        ("Equipment", _get("title")),
        ("Category", _get("category")),
        ("Manufacturer", _get("make")),
        ("Model", _get("model")),
        ("Year", _get("year")),
        ("Condition", _get("condition")),
        ("Hours", _get("hours")),
        ("Horsepower", _get("horsepower")),
    ]
    for label, val in fields:
        if val is not None:
            parts.append(f"{label}: {val}")
    return "\n".join(parts)


async def _price_fuelled_batch(items: list[dict], job: dict):
    """Background task: price each item via run_pricing, write audit rows + update listings."""
    global _fuelled_job
    for i, item in enumerate(items):
        job["current_item"] = item.get("title", "")
        query = _build_pricing_query(item)

        fmv_low = fmv_mid = fmv_high = None
        confidence = None
        tools_used_str = ""
        trace_id = None
        status = "success"
        error_reason = None

        try:
            out = await asyncio.wait_for(run_pricing(query), timeout=60)
            v = out.get("structured", {}).get("valuation", {})
            fmv_low = v.get("fmv_low", 0) or 0
            fmv_mid = v.get("fmv_mid", 0) or 0
            fmv_high = v.get("fmv_high", 0) or 0
            confidence = out.get("confidence", "LOW")
            tools_used_str = ",".join(out.get("tools_used", []))
            trace_id = out.get("trace_id")
        except asyncio.TimeoutError:
            status = "failed"
            error_reason = "Timed out after 60s"
        except Exception as e:
            status = "failed"
            error_reason = str(e)[:500]

        listing_id = item["id"]
        tier_val = _tier(item)
        completeness_val, _ = _completeness_and_missing(item)

        try:
            async with get_session() as session:
                await _ensure_table(session)

                # Insert audit row
                await session.execute(text("""
                    INSERT INTO fuelled_valuations
                        (id, listing_id, fmv_low, fmv_mid, fmv_high,
                         confidence, status, error_reason, tier,
                         data_completeness, tools_used, trace_id)
                    VALUES
                        (:id, :lid, :fmv_low, :fmv_mid, :fmv_high,
                         :conf, :status, :err, :tier,
                         :comp, :tools, :trace)
                """), {
                    "id": str(uuid.uuid4()),
                    "lid": listing_id,
                    "fmv_low": fmv_low,
                    "fmv_mid": fmv_mid,
                    "fmv_high": fmv_high,
                    "conf": confidence,
                    "status": status,
                    "err": error_reason,
                    "tier": tier_val,
                    "comp": completeness_val,
                    "tools": tools_used_str,
                    "trace": trace_id,
                })

                # Update listing fair_value if success and fmv_mid > 0
                if status == "success" and fmv_mid and fmv_mid > 0:
                    await session.execute(text("""
                        UPDATE listings SET fair_value = :fv, last_valued_at = NOW()
                        WHERE id = :lid
                    """), {"fv": fmv_mid, "lid": listing_id})

                await session.commit()

            if status == "success":
                job["succeeded"] += 1
            else:
                job["failed"] += 1
        except Exception as e:
            _log.error("DB write failed for listing %s: %s", listing_id, e)
            job["failed"] += 1

        job["completed"] = i + 1

    job["status"] = "completed"
    job["current_item"] = None
    job["finished_at"] = datetime.now(timezone.utc).isoformat()


# ---------------------------------------------------------------------------
# POST /admin/fuelled/price-batch
# ---------------------------------------------------------------------------

@router.post("/admin/fuelled/price-batch")
async def fuelled_price_batch(
    authorization: str = Header(None),
    body: Optional[dict] = Body(None),
):
    """Launch batch AI pricing for unpriced Fuelled listings. Admin only."""
    global _fuelled_job
    _require_admin(authorization)

    # Reject duplicate while a job is running
    if _fuelled_job and _fuelled_job.get("status") == "running":
        return JSONResponse(
            status_code=409,
            content={"detail": "A batch job is already running", "job_id": _fuelled_job["job_id"]},
        )

    # Parse optional filters
    tiers = (body or {}).get("tiers")
    limit = (body or {}).get("limit", 50)

    # Build tier filter clause
    tier_clause = ""
    if tiers:
        tier_list = ",".join(str(int(t)) for t in tiers)
        tier_clause = f"AND ({_TIER}) IN ({tier_list})"

    async with get_session() as session:
        await _ensure_table(session)

        result = await session.execute(text(f"""
            SELECT id, title, category, make, model, year,
                   condition, hours, horsepower
            FROM listings
            WHERE {_BASE} AND NOT {_HAS_VALUE}
            {tier_clause}
            ORDER BY ({_COMPLETENESS}) DESC
            LIMIT :lim
        """), {"lim": limit})
        rows = result.fetchall()

    if not rows:
        return {"job_id": None, "total": 0, "detail": "No unpriced items found"}

    items = []
    for r in rows:
        items.append({
            "id": r["id"], "title": r["title"], "category": r["category"],
            "make": r["make"], "model": r["model"], "year": r["year"],
            "condition": r["condition"], "hours": r["hours"],
            "horsepower": r["horsepower"],
        })

    job_id = str(uuid.uuid4())
    _fuelled_job = {
        "job_id": job_id,
        "status": "running",
        "total": len(items),
        "completed": 0,
        "succeeded": 0,
        "failed": 0,
        "current_item": None,
        "started_at": datetime.now(timezone.utc).isoformat(),
        "finished_at": None,
    }

    asyncio.create_task(_price_fuelled_batch(items, _fuelled_job))

    return {"job_id": job_id, "total": len(items)}


# ---------------------------------------------------------------------------
# GET /admin/fuelled/price-batch/status
# ---------------------------------------------------------------------------

@router.get("/admin/fuelled/price-batch/status")
async def fuelled_price_batch_status(authorization: str = Header(None)):
    """Return current batch job status. Read-only — any authenticated user."""
    _require_auth(authorization)

    if _fuelled_job is None:
        return {"status": "idle"}
    return _fuelled_job
