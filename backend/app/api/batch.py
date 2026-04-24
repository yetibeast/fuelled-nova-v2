"""Batch pricing endpoints: JSON batch, spreadsheet upload, export."""
from __future__ import annotations
import asyncio
import csv
import datetime
import json
import os
import uuid
from io import BytesIO, StringIO

import jwt
from fastapi import APIRouter, Header, HTTPException, File, UploadFile
from fastapi.responses import Response
from pydantic import BaseModel

from app.config import JWT_SECRET, LOG_DIR
from app.pricing_v2.service import run_pricing

router = APIRouter(prefix="/price", tags=["batch"])


# ── Auth helper ───────────────────────────────────────────────
def _require_auth(authorization: str | None):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing token")
    try:
        jwt.decode(authorization[7:], JWT_SECRET, algorithms=["HS256"])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


# ── Request / response models ─────────────────────────────────
class BatchItem(BaseModel):
    title: str
    category: str = ""
    specs: dict = {}


class BatchRequest(BaseModel):
    items: list[BatchItem]


# ── Core batch logic ──────────────────────────────────────────
async def _price_batch(items: list[BatchItem]) -> dict:
    results, errors = [], []
    total_fmv_low, total_fmv_high = 0, 0

    for i, item in enumerate(items):
        specs_text = ", ".join(f"{k}: {v}" for k, v in item.specs.items()) if item.specs else ""
        user_msg = f"{item.title}. Category: {item.category}. {specs_text}".strip()
        try:
            out = await asyncio.wait_for(run_pricing(user_msg), timeout=60)
            v = out.get("structured", {}).get("valuation", {})
            fmv_lo = v.get("fmv_low", 0) or 0
            fmv_hi = v.get("fmv_high", 0) or 0
            total_fmv_low += fmv_lo
            total_fmv_high += fmv_hi
            results.append({
                "title": item.title,
                "structured": out.get("structured", {}),
                "response": out.get("response", ""),
                "confidence": out.get("confidence", "LOW"),
                "tools_used": out.get("tools_used", []),
            })
        except asyncio.TimeoutError:
            errors.append({"title": item.title, "error": "Timed out after 60s"})
        except Exception as e:
            errors.append({"title": item.title, "error": str(e)})

    summary = {
        "total": len(items),
        "completed": len(results),
        "failed": len(errors),
        "total_fmv_low": total_fmv_low,
        "total_fmv_high": total_fmv_high,
    }

    # Log batch job
    os.makedirs(LOG_DIR, exist_ok=True)
    entry = {"timestamp": datetime.datetime.utcnow().isoformat() + "Z", "summary": summary}
    with open(os.path.join(LOG_DIR, "batch_log.jsonl"), "a") as f:
        f.write(json.dumps(entry) + "\n")

    return {"results": results, "errors": errors, "summary": summary}


# ── In-memory async job store ────────────────────────────────
_batch_jobs: dict[str, dict] = {}
_JOB_TTL_SECONDS = 3600  # 1 hour


def _cleanup_old_jobs():
    """Evict batch jobs older than 1 hour to prevent unbounded memory growth."""
    now = datetime.datetime.utcnow()
    expired = [
        jid for jid, job in _batch_jobs.items()
        if (now - job.get("created_at", now)).total_seconds() > _JOB_TTL_SECONDS
    ]
    for jid in expired:
        del _batch_jobs[jid]


async def _price_batch_async(job_id: str, items: list[BatchItem]):
    """Price items one by one, updating _batch_jobs[job_id] after each."""
    job = _batch_jobs[job_id]
    total_fmv_low, total_fmv_high = 0, 0
    for i, item in enumerate(items):
        job["current_item"] = item.title
        specs_text = ", ".join(f"{k}: {v}" for k, v in item.specs.items()) if item.specs else ""
        user_msg = f"{item.title}. Category: {item.category}. {specs_text}".strip()
        try:
            out = await asyncio.wait_for(run_pricing(user_msg), timeout=60)
            v = out.get("structured", {}).get("valuation", {})
            fmv_lo = v.get("fmv_low", 0) or 0
            fmv_hi = v.get("fmv_high", 0) or 0
            total_fmv_low += fmv_lo
            total_fmv_high += fmv_hi
            job["results"].append({
                "title": item.title,
                "structured": out.get("structured", {}),
                "response": out.get("response", ""),
                "confidence": out.get("confidence", "LOW"),
                "tools_used": out.get("tools_used", []),
            })
        except asyncio.TimeoutError:
            job["errors"].append({"title": item.title, "error": "Timed out after 60s"})
        except Exception as e:
            job["errors"].append({"title": item.title, "error": str(e)})
        job["completed"] = i + 1

    job["status"] = "completed"
    job["current_item"] = None
    job["summary"] = {
        "total": len(items),
        "completed": len(job["results"]),
        "failed": len(job["errors"]),
        "total_fmv_low": total_fmv_low,
        "total_fmv_high": total_fmv_high,
    }

    # Log batch job
    os.makedirs(LOG_DIR, exist_ok=True)
    entry = {"timestamp": datetime.datetime.utcnow().isoformat() + "Z", "summary": job["summary"]}
    with open(os.path.join(LOG_DIR, "batch_log.jsonl"), "a") as f:
        f.write(json.dumps(entry) + "\n")


async def _parse_file_to_items(data: bytes, filename: str) -> list[BatchItem]:
    """Parse uploaded file bytes into BatchItem list.

    Fast path: row-per-item CSV/XLSX with schema headers (title, make, model, ...).
    Fallback: hand the file to Claude to extract items from arbitrary shapes
    (pivoted matrices, emails, free-form inventories).
    """
    ext = (filename or "").rsplit(".", 1)[-1].lower()
    if ext not in ("csv", "xlsx", "xls", "eml"):
        raise HTTPException(status_code=400, detail="Unsupported file type. Use .xlsx, .csv, or .eml")

    # Fast path: try strict row-per-item schema for CSV/XLSX only.
    if ext in ("csv", "xlsx", "xls"):
        items = _try_schema_parse(data, ext)
        if items:
            return items

    # Fallback: LLM extraction for anything else (pivoted sheets, emails, unknown shapes).
    from app.pricing_v2.batch_extractor import extract_items
    try:
        raw_items = await extract_items(data, filename)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Could not extract items: {e}")

    if not raw_items:
        raise HTTPException(status_code=400, detail="No equipment items found in the file")
    return [BatchItem(**it) for it in raw_items]


def _try_schema_parse(data: bytes, ext: str) -> list[BatchItem] | None:
    """Attempt the fast row-per-item parse.

    Returns:
        - list[BatchItem] if a title column was detected and rows parsed successfully.
        - None if the file doesn't look row-per-item (no title column) — caller
          should fall back to the LLM extractor.

    Raises HTTPException(400) if a schema header is present but there are no data rows
    (clear user error; no point asking the LLM).
    """
    if ext == "csv":
        text = data.decode("utf-8-sig", errors="replace")
        all_rows = list(csv.reader(StringIO(text)))
    else:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        all_rows = [[str(c.value or "") for c in row] for row in ws.iter_rows()]
        wb.close()

    if not all_rows:
        return None
    col_map = _detect_columns(all_rows[0])
    if "title" not in col_map:
        return None
    items = _rows_to_items(all_rows[1:], col_map)
    if not items:
        raise HTTPException(status_code=400, detail="Spreadsheet has header row but no data rows")
    return items


# ── POST /api/price/batch/start  (async with polling) ───────
@router.post("/batch/start")
async def batch_start(file: UploadFile = File(...), authorization: str = Header(None)):
    _require_auth(authorization)
    data = await file.read(10_485_761)
    if len(data) > 10_485_760:
        raise HTTPException(status_code=413, detail="File too large (10MB max)")

    _cleanup_old_jobs()

    # Parse lazily in the background — the LLM fallback can take minutes on
    # large pivoted files, longer than any proxy will hold an HTTP POST open.
    job_id = str(uuid.uuid4())
    _batch_jobs[job_id] = {
        "job_id": job_id,
        "status": "parsing",
        "total": 0,
        "completed": 0,
        "current_item": "Analyzing file — this can take a few minutes for complex layouts…",
        "results": [],
        "errors": [],
        "summary": None,
        "created_at": datetime.datetime.utcnow(),
    }
    asyncio.create_task(_parse_then_price(job_id, data, file.filename or ""))
    return {"job_id": job_id}


async def _parse_then_price(job_id: str, data: bytes, filename: str):
    job = _batch_jobs[job_id]
    try:
        items = await _parse_file_to_items(data, filename)
    except HTTPException as e:
        job["status"] = "failed"
        job["error"] = e.detail
        job["current_item"] = None
        return
    except Exception as e:
        job["status"] = "failed"
        job["error"] = f"Failed to parse file: {e}"
        job["current_item"] = None
        return

    job["status"] = "running"
    job["total"] = len(items)
    job["current_item"] = None
    await _price_batch_async(job_id, items)


# ── GET /api/price/batch/{job_id}/status ─────────────────────
@router.get("/batch/{job_id}/status")
async def batch_status(job_id: str, authorization: str = Header(None)):
    _require_auth(authorization)
    if job_id not in _batch_jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    return _batch_jobs[job_id]


# ── POST /api/price/batch ─────────────────────────────────────
@router.post("/batch")
async def batch_price(body: BatchRequest, authorization: str = Header(None)):
    _require_auth(authorization)
    return await _price_batch(body.items)


# ── POST /api/price/batch/upload  (spreadsheet) ──────────────
_COL_HINTS = {
    "title": ["title", "description", "equipment", "name", "item"],
    "category": ["category", "type", "class"],
    "make": ["make", "manufacturer", "mfg"],
    "model": ["model"],
    "year": ["year", "vintage"],
    "hp": ["hp", "horsepower", "power"],
    "location": ["location", "region", "province", "state"],
    "specs": ["specs", "specifications", "details", "notes"],
}


def _detect_columns(headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    lower = [h.lower().strip() for h in headers]
    for field, hints in _COL_HINTS.items():
        for i, col in enumerate(lower):
            if any(h in col for h in hints):
                mapping[field] = i
                break
    return mapping


def _rows_to_items(rows: list[list[str]], col_map: dict[str, int]) -> list[BatchItem]:
    items: list[BatchItem] = []
    for row in rows:
        if not any(cell.strip() for cell in row):
            continue
        title = row[col_map["title"]].strip() if "title" in col_map and col_map["title"] < len(row) else ""
        if not title:
            continue
        category = row[col_map["category"]].strip() if "category" in col_map and col_map["category"] < len(row) else ""
        specs: dict = {}
        for field in ("make", "model", "year", "hp", "location"):
            if field in col_map and col_map[field] < len(row):
                val = row[col_map[field]].strip()
                if val:
                    specs[field] = val
        items.append(BatchItem(title=title, category=category, specs=specs))
    return items


@router.post("/batch/upload")
async def batch_upload(file: UploadFile = File(...), authorization: str = Header(None)):
    _require_auth(authorization)
    data = await file.read(10_485_761)
    if len(data) > 10_485_760:
        raise HTTPException(status_code=413, detail="File too large (10MB max)")
    items = await _parse_file_to_items(data, file.filename or "")
    return await _price_batch(items)


# ── POST /api/price/batch/export  (XLSX download) ────────────
class ExportRequest(BaseModel):
    results: list[dict]


@router.post("/batch/export")
async def batch_export(body: ExportRequest, authorization: str = Header(None)):
    _require_auth(authorization)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, numbers

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Valuations"

    headers = ["#", "Title", "Category", "RCN ($)", "FMV Low ($)", "FMV High ($)",
               "FMV Mid ($)", "Confidence", "Key Rationale", "Comps Found", "Risk Factors"]
    navy_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=10)
    alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    currency_fmt = '#,##0'

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = navy_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")

    cat_totals: dict[str, dict] = {}
    for ri, r in enumerate(body.results, 2):
        v = r.get("structured", {}).get("valuation", {})
        fmv_lo = v.get("fmv_low", 0) or 0
        fmv_hi = v.get("fmv_high", 0) or 0
        fmv_mid = (fmv_lo + fmv_hi) / 2 if fmv_lo else 0
        rcn = v.get("rcn", 0) or 0
        cat = v.get("category", "") or r.get("title", "")[:20]
        comps = r.get("structured", {}).get("comparables", [])
        risks = r.get("structured", {}).get("risks", [])
        rationale = (r.get("response", "") or "")[:120]

        vals = [ri - 1, r.get("title", ""), cat, rcn, fmv_lo, fmv_hi, fmv_mid,
                r.get("confidence", ""), rationale, len(comps),
                "; ".join(risks) if isinstance(risks, list) else str(risks)]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ci in (4, 5, 6, 7):
                cell.number_format = currency_fmt
            if ri % 2 == 0:
                cell.fill = alt_fill

        cat_totals.setdefault(cat, {"count": 0, "fmv_low": 0, "fmv_high": 0})
        cat_totals[cat]["count"] += 1
        cat_totals[cat]["fmv_low"] += fmv_lo
        cat_totals[cat]["fmv_high"] += fmv_hi

    # Summary tab
    ws2 = wb.create_sheet("Summary")
    sum_headers = ["Category", "Count", "Total FMV Low ($)", "Total FMV High ($)"]
    for ci, h in enumerate(sum_headers, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.fill = navy_fill
        cell.font = white_font
    for ri, (cat, vals) in enumerate(cat_totals.items(), 2):
        ws2.cell(row=ri, column=1, value=cat)
        ws2.cell(row=ri, column=2, value=vals["count"])
        ws2.cell(row=ri, column=3, value=vals["fmv_low"]).number_format = currency_fmt
        ws2.cell(row=ri, column=4, value=vals["fmv_high"]).number_format = currency_fmt

    buf = BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Fuelled_Batch_Valuations.xlsx"},
    )


# ── POST /api/price/batch/report  (portfolio .docx) ─────────
class ReportRequest(BaseModel):
    results: list[dict]
    summary: dict = {}


@router.post("/batch/report")
async def batch_report(body: ReportRequest, authorization: str = Header(None)):
    _require_auth(authorization)
    from app.pricing_v2.portfolio_report import generate_portfolio_report

    summary = body.summary or {
        "total": len(body.results),
        "completed": len(body.results),
        "failed": 0,
        "total_fmv_low": sum(
            (r.get("structured", {}).get("valuation", {}).get("fmv_low", 0) or 0)
            for r in body.results
        ),
        "total_fmv_high": sum(
            (r.get("structured", {}).get("valuation", {}).get("fmv_high", 0) or 0)
            for r in body.results
        ),
    }
    docx_bytes = generate_portfolio_report(body.results, summary)
    return Response(
        content=docx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": "attachment; filename=Fuelled_Portfolio_Report.docx"},
    )
