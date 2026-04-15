"""Tests for Fuelled coverage + report endpoints."""
from __future__ import annotations

import io

from openpyxl import load_workbook


class TestFuelledCoverage:
    """Coverage stats endpoint tests."""

    def test_requires_auth(self, client):
        """401 without token."""
        resp = client.get("/api/admin/fuelled/coverage")
        assert resp.status_code == 401

    def test_returns_coverage_stats(self, client, admin_headers):
        """All expected fields are present in the response."""
        resp = client.get("/api/admin/fuelled/coverage", headers=admin_headers)
        assert resp.status_code == 200
        data = resp.json()
        expected_keys = {
            "total", "asking_price_count", "asking_price_pct",
            "valued_count", "valued_pct", "ai_only_count",
            "unpriced", "by_tier", "by_category", "completeness_avg",
        }
        assert expected_keys.issubset(data.keys())

    def test_percentages_are_valid(self, client, admin_headers):
        """Percentages are 0-100 and valued_pct >= asking_price_pct."""
        resp = client.get("/api/admin/fuelled/coverage", headers=admin_headers)
        data = resp.json()
        assert 0 <= data["asking_price_pct"] <= 100
        assert 0 <= data["valued_pct"] <= 100
        assert data["valued_pct"] >= data["asking_price_pct"]

    def test_counts_match_seeds(self, client, admin_headers):
        """Counts align with the 6 seeded fuelled listings."""
        resp = client.get("/api/admin/fuelled/coverage", headers=admin_headers)
        data = resp.json()
        # 6 total seeded fuelled listings
        assert data["total"] == 6
        # 1 has asking_price (fu-t1-priced = 450000)
        assert data["asking_price_count"] == 1
        # 5 have no asking_price and no fair_value
        assert data["unpriced"] == 5

    def test_analyst_can_view(self, client, user_headers):
        """Non-admin users (analysts) can access this read-only endpoint."""
        resp = client.get("/api/admin/fuelled/coverage", headers=user_headers)
        assert resp.status_code == 200


class TestFuelledReport:
    """POST /admin/fuelled/generate-report XLSX download tests."""

    def test_requires_auth(self, client):
        """401 without token."""
        resp = client.post("/api/admin/fuelled/generate-report")
        assert resp.status_code == 401

    def test_returns_xlsx(self, client, admin_headers):
        """200 with spreadsheet content-type."""
        resp = client.post("/api/admin/fuelled/generate-report", headers=admin_headers)
        assert resp.status_code == 200
        assert "spreadsheet" in resp.headers["content-type"]

    def test_xlsx_has_expected_headers(self, client, admin_headers):
        """All required columns are present in the header row."""
        resp = client.post("/api/admin/fuelled/generate-report", headers=admin_headers)
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        expected = [
            "Title", "Category", "Make", "Model", "Year",
            "Condition", "Hours", "HP", "Data Completeness %",
            "Missing Fields", "Days Listed", "Pricability Tier", "URL",
        ]
        assert headers == expected

    def test_report_row_count_matches_unpriced(self, client, admin_headers):
        """5 data rows + 1 header = 6 total rows."""
        resp = client.post("/api/admin/fuelled/generate-report", headers=admin_headers)
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        # 5 unpriced listings + 1 header row = 6 rows
        assert ws.max_row == 6
