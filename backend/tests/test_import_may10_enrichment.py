"""Tests for backend/scripts/import_may10_enrichment.py.

Verifies parsing of the May 10 supply-targets workbook into the two new tables:
  - seller_contact_enrichment (supply-side, joins to listings.seller_name)
  - buyer_targets (buy-side, no listings join)

The script reads the xlsx and emits parameter dicts ready for the INSERT
... ON CONFLICT ... DO UPDATE statements. To keep the tests fast and
hermetic we exercise the parser layer directly (parse_named_contacts,
parse_ls_event_managers, parse_buyer_companies, parse_buyer_contacts) and
the row-level upsert is exercised against the in-memory MockSession via a
separate test that runs the script's main() function.
"""
from __future__ import annotations

import io

import openpyxl
import pytest


# Skip-import guard until the script lands.
import_may10_enrichment = pytest.importorskip(
    "scripts.import_may10_enrichment",
    reason="import script not yet implemented",
)


# ── Fixture helpers ──────────────────────────────────────────────────────


def _build_workbook() -> openpyxl.Workbook:
    """Build a miniature workbook that mirrors the four sheets the importer
    consumes. Three rows per sheet is enough to exercise the parsing branches
    (named contact, LS event manager, buyer company, buyer contact)."""
    wb = openpyxl.Workbook()
    # openpyxl creates a default 'Sheet' — rename it for the first sheet.
    nc = wb.active
    nc.title = "Named Contacts (Direct)"
    nc.append([
        "Company", "Name", "Title", "LinkedIn",
        "Location", "Email (likely)", "Confidence", "Outreach Notes",
    ])
    nc.append([
        "Central Valley Renewables (CVRF)", "Mariah Mandt, PE",
        "VP / GM, Bakersfield Renewable Fuels",
        "https://www.linkedin.com/in/mariah-mandt-pe-22a33895/",
        "Bakersfield CA", "mariah.mandt@bkrenewablefuels.com",
        "likely", "Site GM — owns physical asset decisions.",
    ])
    nc.append([
        "Central Valley Renewables (CVRF)", "Saket Adhatrao",
        "DCS Controls Specialist / Operations",
        "https://www.linkedin.com/in/saket-adhatrao/",
        "Bakersfield CA", "saket.adhatrao@bkrenewablefuels.com",
        "likely", "Active operations contact.",
    ])
    # Row with whitespace + None to exercise normalization.
    nc.append([
        "  Empty Co  ", "  ", None, None, None, None, None, None,
    ])

    ls = wb.create_sheet("LS Event Managers")
    ls.append([
        "Name", "Email", "Phone", "LinkedIn", "Title",
        "Region", "Tenure", "Events Covered", "Notes for Outreach", "Confidence",
    ])
    ls.append([
        "Ruth Hernandez", "ruth.hernandez@liquidityservices.com",
        "713-590-1418 (Houston)",
        "https://www.linkedin.com/in/ruth-hernandez-14866239",
        "Inside Sales Representative, Liquidity Services CAG",
        "US — Houston", "unknown",
        "ConocoPhillips Permian, Pemex Deer Park",
        "Operational POC — gatekeeper.", "likely",
    ])
    ls.append([
        "Jeff Goddard", "goddard@liquidityservices.com",
        "714-376-1422 (Orange County CA)",
        "https://www.linkedin.com/in/jeff-goddard-cea-34703857/",
        "CAG Account Manager, Liquidity Services",
        "Southern California / US West", "3+ years",
        "Three smaller anonymous events",
        "Real account manager.", "confirmed",
    ])

    bt = wb.create_sheet("Buyer Targets (Companies)")
    bt.append([
        "Vertical", "Company", "Ticker", "HQ",
        "Basin / Footprint", "Scale", "Capex Driver / Why They Buy",
        "Suppliers Page",
    ])
    bt.append([
        "US Upstream O&G", "Diamondback Energy", "FANG", "Midland TX",
        "Permian (Midland + Delaware)",
        "13-14 rigs end-2025; ~$45B mkt cap",
        "Closed Endeavor mega-deal 2024. Ramping capex.",
        "https://www.diamondbackenergy.com/about/overview",
    ])
    bt.append([
        "US Upstream O&G", "Permian Resources", "PR", "Midland + Houston TX",
        "Permian (Delaware Basin core)",
        "190.7 net wells in 2025",
        "$608M Apache bolt-on.",
        "https://permianres.com/contact/",
    ])
    # Buyer with no buyer-contact match — should still get a buyer_targets
    # row (company-level only, contact fields NULL).
    bt.append([
        "US Midstream", "Lonely Co Inc", "LONE", "Houston TX",
        "Eagle Ford", "Small private", "Single-asset operator.",
        "https://lonelyco.example/suppliers",
    ])

    bc = wb.create_sheet("Buyer Contacts")
    bc.append([
        "Company", "Name", "Title", "LinkedIn",
        "Location", "Email (likely)", "Confidence", "Outreach Notes",
    ])
    bc.append([
        "Diamondback Energy", "Gwendolyn ('Gwen') Smith",
        "Senior Procurement Specialist",
        "https://www.linkedin.com/in/gwendolyn-smith2/",
        "Midland TX", "gwendolyn.smith@diamondbackenergy.com",
        "likely", "Senior procurement IC.",
    ])
    bc.append([
        "Diamondback Energy", "Cody King", "Materials Supervisor",
        "https://www.linkedin.com/in/cody-king-977b2292/",
        "Midland TX", "cody.king@diamondbackenergy.com",
        "likely", "Runs materials coordination.",
    ])
    bc.append([
        "Permian Resources", "Steven Dodson",
        "CISO (cross-functional escalation)",
        "https://www.linkedin.com/in/stevedod/",
        "Dallas-Fort Worth TX", "steven.dodson@permianres.com",
        "likely", "ZoomInfo confirms email pattern.",
    ])
    return wb


@pytest.fixture
def workbook_path(tmp_path):
    wb = _build_workbook()
    out = tmp_path / "fixture.xlsx"
    wb.save(out)
    return out


# ── parse_named_contacts ────────────────────────────────────────────────


def test_parse_named_contacts_returns_one_row_per_contact(workbook_path):
    rows = import_may10_enrichment.parse_named_contacts(workbook_path)
    # Two real rows + one whitespace-only row should be skipped.
    assert len(rows) == 2
    names = [r["contact_name"] for r in rows]
    assert "Mariah Mandt, PE" in names
    assert "Saket Adhatrao" in names


def test_parse_named_contacts_maps_columns(workbook_path):
    rows = import_may10_enrichment.parse_named_contacts(workbook_path)
    mariah = next(r for r in rows if r["contact_name"] == "Mariah Mandt, PE")
    # seller_name == Company (this row joins to listings.seller_name)
    assert mariah["seller_name"] == "Central Valley Renewables (CVRF)"
    # source NULL → applies to all sources for that seller
    assert mariah["source"] is None
    assert mariah["contact_title"] == "VP / GM, Bakersfield Renewable Fuels"
    assert mariah["contact_email"] == "mariah.mandt@bkrenewablefuels.com"
    assert mariah["contact_linkedin"] == \
        "https://www.linkedin.com/in/mariah-mandt-pe-22a33895/"
    assert mariah["location"] == "Bakersfield CA"
    assert mariah["contact_confidence"] == "likely"
    assert "Site GM" in mariah["outreach_notes"]


def test_parse_named_contacts_skips_blank_rows(workbook_path):
    """Rows where Company is whitespace-only and Name is blank should be dropped."""
    rows = import_may10_enrichment.parse_named_contacts(workbook_path)
    for r in rows:
        assert r["seller_name"].strip() != ""


# ── parse_ls_event_managers ─────────────────────────────────────────────


def test_parse_ls_event_managers_pins_source_to_allsurplus(workbook_path):
    rows = import_may10_enrichment.parse_ls_event_managers(workbook_path)
    # Two LS managers in fixture
    assert len(rows) == 2
    for r in rows:
        assert r["source"] == "allsurplus"
        # seller_name is canonical Liquidity Services bucket
        assert r["seller_name"] == "Liquidity Services"


def test_parse_ls_event_managers_prefixes_outreach_notes_with_events(workbook_path):
    rows = import_may10_enrichment.parse_ls_event_managers(workbook_path)
    ruth = next(r for r in rows if r["contact_name"] == "Ruth Hernandez")
    # Outreach notes prefixed with "[LS Event Manager — covers events: X, Y]"
    assert ruth["outreach_notes"].startswith("[LS Event Manager — covers events:")
    assert "ConocoPhillips Permian" in ruth["outreach_notes"]
    assert "Operational POC" in ruth["outreach_notes"]


def test_parse_ls_event_managers_uses_email_phone_linkedin(workbook_path):
    rows = import_may10_enrichment.parse_ls_event_managers(workbook_path)
    jeff = next(r for r in rows if r["contact_name"] == "Jeff Goddard")
    assert jeff["contact_email"] == "goddard@liquidityservices.com"
    assert "714-376-1422" in jeff["contact_phone"]
    assert "linkedin.com/in/jeff-goddard" in jeff["contact_linkedin"]
    assert jeff["contact_confidence"] == "confirmed"


# ── parse_buyer_companies + parse_buyer_contacts ────────────────────────


def test_parse_buyer_targets_joins_contacts_to_companies(workbook_path):
    rows = import_may10_enrichment.parse_buyer_targets(workbook_path)
    diamondback = [r for r in rows if r["company"] == "Diamondback Energy"]
    # Two Diamondback contacts → two rows, both carrying company metadata.
    assert len(diamondback) == 2
    names = {r["contact_name"] for r in diamondback}
    assert names == {"Gwendolyn ('Gwen') Smith", "Cody King"}
    for r in diamondback:
        assert r["vertical"] == "US Upstream O&G"
        assert r["ticker"] == "FANG"
        assert r["hq"] == "Midland TX"
        assert "Permian" in r["basin"]
        assert "rigs end-2025" in r["scale"]
        assert "Endeavor" in r["capex_driver"]
        assert r["suppliers_page"] == \
            "https://www.diamondbackenergy.com/about/overview"


def test_parse_buyer_targets_emits_company_row_when_no_contacts(workbook_path):
    """Companies with no Buyer Contacts row still get a row (contact fields NULL)."""
    rows = import_may10_enrichment.parse_buyer_targets(workbook_path)
    lonely = [r for r in rows if r["company"] == "Lonely Co Inc"]
    assert len(lonely) == 1
    r = lonely[0]
    assert r["contact_name"] is None
    assert r["contact_email"] is None
    assert r["vertical"] == "US Midstream"
    assert r["ticker"] == "LONE"


def test_parse_buyer_targets_contact_metadata(workbook_path):
    rows = import_may10_enrichment.parse_buyer_targets(workbook_path)
    gwen = next(r for r in rows if r["contact_name"] == "Gwendolyn ('Gwen') Smith")
    assert gwen["contact_title"] == "Senior Procurement Specialist"
    assert gwen["contact_email"] == "gwendolyn.smith@diamondbackenergy.com"
    assert gwen["location"] == "Midland TX"
    assert gwen["contact_confidence"] == "likely"
    assert "procurement" in gwen["outreach_notes"].lower()


# ── No empty/None garbage rows ───────────────────────────────────────────


def test_all_parsers_handle_none_and_whitespace(workbook_path):
    """Mixed None / whitespace / real data should never produce rows where
    the joining key (seller_name for seller, company for buyer) is empty
    or all-whitespace."""
    seller_rows = (
        import_may10_enrichment.parse_named_contacts(workbook_path)
        + import_may10_enrichment.parse_ls_event_managers(workbook_path)
    )
    for r in seller_rows:
        assert r["seller_name"] and r["seller_name"].strip()
    buyer_rows = import_may10_enrichment.parse_buyer_targets(workbook_path)
    for r in buyer_rows:
        assert r["company"] and r["company"].strip()
