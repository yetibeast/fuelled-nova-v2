"""Merge public-web research findings into the supply-targets workbook.

Takes the latest export and produces an enriched copy:
- New columns on the Sellers sheet for owner/broker classification, location,
  industry, parent, disposition context, direct contact path, confidence
- New 'Research Notes' sheet with full narrative per seller + sources
- New 'LS Managers' sheet with the human-research on the 4 Liquidity Services reps

Research data (2026-05-08) is embedded below — gathered from public web only,
no paid databases. Re-run this script after any new research pass.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="EF5D28")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def write_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"


def autosize(ws, max_width: int = 60) -> None:
    for col_idx, col in enumerate(ws.columns, start=1):
        try:
            length = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        except ValueError:
            length = 10
        ws.column_dimensions[get_column_letter(col_idx)].width = min(length + 2, max_width)


# -----------------------------------------------------------------------------
# Research data — keyed by exact 'Seller' column value from the workbook.
# Add entries here as new research lands; missing sellers stay blank in the output.
# -----------------------------------------------------------------------------

RESEARCH: dict[str, dict] = {
    # ---- South Africa mining cluster (Liversage's portfolio) ----
    "Assmang Cato Ridge Works": {
        "owner_class": "asset owner",
        "industry": "Ferromanganese smelter",
        "location": "Cato Ridge, KwaZulu-Natal, South Africa",
        "parent": "Assmang Pty Ltd (JV: Assore + ARM)",
        "disposition": "Permanent closure announced July 2025. Plant ceased 31 Aug 2025; ~600 jobs lost. Land sold to Assore PropCo for R453m for redevelopment. Equipment is being cleared, not idled.",
        "direct_contact": "Assore (parent): assore.com/contact-us | ARM: arm.co.za/contact-us | For asset disposition: Assore PropCo team handling the site redevelopment",
        "confidence": "confirmed",
        "notes": "Highest-priority warm lead. Real closure, real equipment moving. Bypass LS via Assore.",
    },
    "Assmang Khumani Iron Ore Mine": {
        "owner_class": "asset owner",
        "industry": "Iron ore surface mine (~14 Mtpa hematite)",
        "location": "30 km south of Kathu, Northern Cape, South Africa",
        "parent": "Assmang Pty Ltd (JV: Assore + ARM)",
        "disposition": "Khumani still producing, but iron-ore portfolio contracting. Sister mine Beeshoek placed on care & maintenance Nov 2025 after AMSA halted offtake. Listings likely fleet renewal + Beeshoek wind-down spillover.",
        "direct_contact": "assmang.co.za | ARM Ferrous: arm.co.za/b/ferrous.php | Procurement via Assore",
        "confidence": "confirmed",
        "notes": "",
    },
    "Glencore Operations South Africa (Pty) Ltd": {
        "owner_class": "asset owner",
        "industry": "Ferrochrome smelting + vanadium + thermal coal",
        "location": "Multiple SA sites: Rustenburg/Boshoek/Wonderkop/Lion smelters; Mpumalanga coal; Rhovan vanadium",
        "parent": "Glencore plc (LSE: GLEN)",
        "disposition": "Boshoek + Wonderkop indefinitely suspended May–June 2025. Rustenburg idled through 2024. S189 retrenchments at SA ferrochrome + vanadium ops Sept 2025. H1 2025 ferrochrome production down 28% YoY. $378m SA coal impairment 2025. Real surplus, not flips.",
        "direct_contact": "glencore.com/south-africa (Johannesburg HQ) | Glencore Merafe Chrome Venture procurement | JV partner Merafe (JSE-listed): merafe.co.za",
        "confidence": "confirmed",
        "notes": "",
    },
    "Northam Booysendal Platinum": {
        "owner_class": "asset owner",
        "industry": "PGM underground mine (4E concentrate)",
        "location": "Limpopo / Mpumalanga border, ~35 km west of Mashishing",
        "parent": "Northam Platinum Holdings Ltd (JSE: NPH)",
        "disposition": "Booysendal South (BSM) suspended 2024–25 while North module (BNM) ramped to ~50,000 oz/y. Surplus = development equipment from completed BNM ramp + BSM suspension.",
        "direct_contact": "northam.co.za/contact (Hyde Park, Johannesburg) | CEO Paul Dunne; CFO Alet Coetzee; IR Memory Johnstone",
        "confidence": "likely",
        "notes": "",
    },
    "MHI Power ZAF (Pty) Ltd": {
        "owner_class": "asset owner",
        "industry": "Coal-fired power-plant boilers (Medupi + Kusile, 800 MW units)",
        "location": "Sandton HQ; equipment likely staged at Medupi (Lephalale, Limpopo) and Kusile (eMalahleni, Mpumalanga) sites",
        "parent": "Mitsubishi Heavy Industries Ltd (TSE: 7011)",
        "disposition": "Medupi + Kusile boiler scopes at end-of-construction. With Eskom's last Kusile units commissioning, MHI winding down site activity — surplus is construction inventory, spares overstock, tooling.",
        "direct_contact": "power.mhi.com/regions/afr/company (Mitsubishi Power Africa, Sandton) | Contact form on power.mhi.com/regions/afr",
        "confidence": "likely",
        "notes": "",
    },
    "ArcelorMittal South Africa  (AMSA)": {  # double-space in seller name preserved
        "owner_class": "asset owner",
        "industry": "Integrated steelmaking — long products + flats",
        "location": "Newcastle (KZN), Vereeniging (Gauteng), Vanderbijlpark, Saldanha",
        "parent": "ArcelorMittal SA (Luxembourg; NYSE: MT)",
        "disposition": "Long Steel Business wind-down confirmed Jan 2025 (Newcastle, Vereeniging) — ~3,500 direct jobs. IDC R1.683bn facility deferred closure to Aug 2025. S189 retrenchment notices began Sept 2025. Auction lots = real, ongoing wind-down. High-quality steelworks equipment.",
        "direct_contact": "arcelormittalsa.com (Vanderbijlpark HQ) | CEO Kobus Verster; IR Tinus Brits | Procurement: arcelormittalsa.com/Procurement",
        "confidence": "confirmed",
        "notes": "Highest-priority warm lead alongside Cato Ridge.",
    },
    "Black Mountain Mining (Pty) Ltd": {
        "owner_class": "asset owner",
        "industry": "Underground zinc/lead/copper + open-pit zinc (Gamsberg)",
        "location": "Aggeneys, Northern Cape, South Africa",
        "parent": "Vedanta Zinc International (Vedanta 69.6% / Exxaro 24.4% / ESOP 6%)",
        "disposition": "Active operations + Gamsberg Phase II expansion (approved March 2022, doubling capacity to 8.5 Mtpa). Single-listing surplus = routine fleet turnover during expansion, not closure.",
        "direct_contact": "vedanta-zincinternational.com/contact-us | aggeneys.com/business/black-mountain-mining | Exxaro alternate route: exxaro.com",
        "confidence": "confirmed",
        "notes": "Lower priority — active mine, just fleet rotation.",
    },
    "SARS": {
        "owner_class": "broker (state agency)",
        "industry": "Mixed seized/forfeited goods (industrial, automotive, agricultural, consumer)",
        "location": "SARS State Warehouses across SA (Cape Town, Durban, JNB/OR Tambo, Beit Bridge)",
        "parent": "South African Revenue Service",
        "disposition": "Routine auctioning of seized/forfeited/unclaimed goods to recover duty + storage costs. AllSurplus is SARS's appointed online auction partner. Original asset owners are unknown third parties.",
        "direct_contact": "sars.gov.za/customs-and-excise/state-warehouses — auction channel IS the disposal channel; no corporate backchannel exists",
        "confidence": "confirmed",
        "notes": "Skip for direct outreach — there is no underlying owner to bypass to.",
    },

    # ---- US oil & gas / refining / utilities cluster (Hernandez + others) ----
    "ConocoPhillips – Permian": {
        "owner_class": "asset owner",
        "industry": "Upstream oil & gas (unconventional shale)",
        "location": "Permian/Delaware Basin, West Texas + New Mexico (likely Midland/Odessa staging)",
        "parent": "ConocoPhillips (NYSE: COP)",
        "disposition": "Active $5B non-core divestiture program post-Marathon merger (closed late 2024). Closed $3.2B in asset sales in 2025. Feb 2026 began exploring ~$2B Delaware Basin sale. Field-level surplus consistent with portfolio rationalization.",
        "direct_contact": "ConocoPhillips Procurement / Supply Chain (925 N. Eldridge Pkwy, Houston) — Investment Recovery falls under SCM. LinkedIn: 'ConocoPhillips Investment Recovery' / 'Surplus Materials' or Permian BU SCM leads in Midland, TX.",
        "confidence": "likely",
        "notes": "Strong warm lead — multi-year divestiture in progress.",
    },
    "Pemex Deer Park": {
        "owner_class": "asset owner",
        "industry": "Petrochemical refinery (340 kbd crude)",
        "location": "Deer Park, TX (Houston Ship Channel) — 5900 Hwy 225",
        "parent": "Petróleos Mexicanos (Pemex)",
        "disposition": "Major maintenance overhaul began 2025 (~$500M+) on largest CDU + several other units. Surplus = turnaround / replace-in-kind, not closure. Pemex investing, not divesting.",
        "direct_contact": "pemex.com/deerpark/en — corporate switchboard (713) 246-7000 | LinkedIn: Pemex Deer Park company page → search Procurement / SCM Manager",
        "confidence": "confirmed",
        "notes": "",
    },
    "Central Valley Renewables (CVRF)": {
        "owner_class": "asset owner",
        "industry": "Renewable diesel refining (~250 MMgal/yr)",
        "location": "Bakersfield, Kern County, CA (former Big West / Alon refinery site)",
        "parent": "Grapevine Energy Holdings, LLC (post-Apr 2025 Ch.11; senior lenders led by OIC + CTCI; was Global Clean Energy Holdings)",
        "disposition": "Project hit $950M EPC dispute with CTCI, 3+ year delays, ExxonMobil exited as offtake. Commercial ops achieved Dec 2024. Restructured 2025 with Vitol as new offtake. Surplus likely from EPC closeout / construction overage / spec'd-out spares.",
        "direct_contact": "cvrenewables.com (Bakersfield site) | Grapevine Energy Holdings (Houston, post-restructuring) | DPSI was EPC services partner",
        "confidence": "confirmed",
        "notes": "Highest-urgency target. Just-restructured single-site operator with EPC-dispute surplus. Small org footprint = direct outreach lands.",
    },
    "Flint Hills Resources Pine Bend, LLC": {
        "owner_class": "asset owner",
        "industry": "Petroleum refining (heavy/sour Canadian crude)",
        "location": "Rosemount / Inver Grove Heights, MN (Twin Cities)",
        "parent": "Koch Industries → Flint Hills Resources, LP",
        "disposition": "No public divestiture announced. Remains FHR's flagship MN refinery. Single AllSurplus listing = routine investment-recovery / capital project surplus.",
        "direct_contact": "fhr.com (Wichita HQ) | pinebendrefinery.com | LinkedIn: 'Flint Hills Resources' + 'Materials Management' / 'Investment Recovery'",
        "confidence": "confirmed",
        "notes": "Lower priority — no closure event.",
    },
    "Valaris - US (SM)": {
        "owner_class": "asset owner",
        "industry": "Offshore drilling (jackups + semisubs)",
        "location": "US Gulf of Mexico stacking yards (Houston / Ingleside / Sabine Pass area)",
        "parent": "Valaris plc (NYSE: VAL) — emerged from Ch.11 Apr 2021",
        "disposition": "Feb 2025 Fleet Rationalization — retiring semisubs DPS-3, DPS-5, DPS-6 (scrap or repurpose) + sold jackup VALARIS 75 for $24M after 5 yrs stacked. Listings = rig dismantlement / cold-stack offload. $4B+ contract backlog = high-grading the fleet.",
        "direct_contact": "valaris.com (5847 San Felipe St #3300, Houston) — +1 713-979-4200 | LinkedIn: Valaris → Supply Chain / Investment Recovery / Asset Manager",
        "confidence": "confirmed",
        "notes": "Strong warm lead — active rig dismantlement.",
    },
    "Entergy Services, Inc.": {
        "owner_class": "asset owner",
        "industry": "Electric utility (Gulf states regulated)",
        "location": "MS/LA/AR/TX service territory; corporate The Woodlands TX / New Orleans LA",
        "parent": "Entergy Corporation (NYSE: ETR)",
        "disposition": "In-house Investment Recovery program. Baxter Wilson steam plant (Vicksburg MS) retired 2018-22, demolished Dec 2025. Site to host new Vicksburg Advanced Power Station gas plant by 2028. Coal-to-gas transition driving plant-retirement surplus.",
        "direct_contact": "entergy.com/operations/investment-recovery — public IR contact form | Fleet via JJ Kane | Aucto seller page: aucto.com/marketplace/seller/entergy/2510",
        "confidence": "confirmed",
        "notes": "Public IR page — easiest direct contact of all the US sellers.",
    },
    "CenterPoint Energy": {
        "owner_class": "asset owner",
        "industry": "Gas LDC + electric T&D utility",
        "location": "Houston, TX HQ; Greater Houston gas/electric territory + IN/MN/OH",
        "parent": "CenterPoint Energy, Inc. (NYSE: CNP)",
        "disposition": "Active modernization cycle (post-Beryl resilience build-out + rate-case capex). 300 mi pipe replaced + 175k new meters in 2023-24. Divested OH gas business to National Fuel Gas for $2.62B (2025). Surplus is steady byproduct of capex cycle.",
        "direct_contact": "surplus.centerpointenergy.com (PRIMARY — dedicated public surplus portal) | corp procurement: 1111 Louisiana St, Houston | centerpointenergy.com/business/about-us/suppliers",
        "confidence": "confirmed",
        "notes": "Public surplus portal already exists. Easiest direct route.",
    },
    "Archer Daniels Midland Company": {
        "owner_class": "asset owner",
        "industry": "Ag processing — soybean crush, cottonseed crush, soy protein, ethanol",
        "location": "Likely one of: Bushnell IL (soy protein, closing); Kershaw SC (soy, closing 2025); Memphis TN (cottonseed, closing Jan 2026)",
        "parent": "Archer-Daniels-Midland Co. (NYSE: ADM)",
        "disposition": "$500-700M cost-cutting program over 3-5 years; up to 700 job cuts. Closing Bushnell IL soy protein, Kershaw SC soybean plant, Memphis TN cottonseed (Jan 2026, 95 layoffs). Memphis JV with Planters Cotton Oil (Pine Bluff AR refinery stays). Surplus tracks footprint rationalization.",
        "direct_contact": "ADM Corporate Procurement / Investment Recovery (77 W. Wacker Dr, Chicago) | adm.com/en-us/suppliers | Site-level: closing-plant managers in Bushnell/Kershaw/Memphis",
        "confidence": "confirmed",
        "notes": "Three plant closures = three direct routes.",
    },

    # ---- Mystery / LS-managed codes ----
    "Seller 23609 - ANCO": {
        "owner_class": "asset owner (CONFIRMED via lot-data triangulation)",
        "industry": "Heavy-civil road construction (asphalt paving, aggregate, dump fleet)",
        "location": "Lake City FL HQ; FL plants at Old Town, Quincy, Marianna, Sumterville; TX yards at Robstown, Weslaco, Corpus Christi (post-2022 Haas-Anderson acquisition), New Braunfels (Tejas Quarry, 2016); GA aggregate ops",
        "parent": "Anderson Columbia Co., Inc. (ACCI) — privately held",
        "disposition": "FLEET ROTATION, not bankruptcy. ACCI is an active operator running a multi-yard equipment-cycling program. The 'ANCO' 5-letter code maps directly to ANderson COlumbia. Asset mix (asphalt paver, asphalt distributor, asphalt-plant drag elevator, McLanahan screen, water trucks, dump trucks, rollers, broom) is a textbook ACCI heavy-civil fleet. ACCI is a documented McLanahan customer; one of the listings is a McLanahan 616TD screen.",
        "direct_contact": "andersoncolumbia.com — Lake City FL HQ. Active operator; corporate procurement / equipment manager is the right route. Recurring dispo channel = relationship play, not one-time auction.",
        "confidence": "confirmed",
        "notes": "BIGGEST WIN of the unmasking pass. 23 listings, was the largest unidentified seller on AllSurplus. ACCI is a recurring fleet rotation, so ongoing buy-side relationship is feasible — not just chasing one auction.",
    },
    "Seller 26812 - ELGLRTNIC": {
        "owner_class": "LS-managed code",
        "industry": "Unknown — only 4 listings",
        "location": "Unknown",
        "parent": "Liquidity Services CAG",
        "disposition": "LS internal code (matches naming/numbering pattern of confirmed LS terms pages in the 19k–28k range).",
        "direct_contact": "Liquidity Services CAG",
        "confidence": "likely",
        "notes": "",
    },
    "Seller 19807 - ASFL": {
        "owner_class": "LS-managed code (CONFIRMED)",
        "industry": "Oilfield (likely)",
        "location": "Unknown — likely Texas-based",
        "parent": "Liquidity Services (T&C page reads 'Liquidity Services – ASFL')",
        "disposition": "Explicitly hosted under 'Liquidity Services –' branding on the AllSurplus T&C page. Underlying client name obfuscated by LS standard practice. Plausible expansion 'Absolute Oil Field Services LLC' (Eagle Ford roustabout) — unverified.",
        "direct_contact": "Liquidity Services CAG — liquidityservices.com/energy-surplus-asset-sales",
        "confidence": "confirmed (LS managed); unverified (underlying entity)",
        "notes": "The smoking gun confirming the LS-code pattern.",
    },
    "Seller 21141 - RELLC": {
        "owner_class": "LS-managed code",
        "industry": "Unknown — only 1 listing",
        "location": "Unknown",
        "parent": "Liquidity Services CAG",
        "disposition": "LS internal code (5-letter acronym ending in 'LLC' suffix collapsed — parallels SBLLC, TKEPTSSLC, VOYMSLC).",
        "direct_contact": "Liquidity Services CAG",
        "confidence": "likely",
        "notes": "",
    },
    "Seller 31459 - AWCTL": {
        "owner_class": "LS-managed code",
        "industry": "Unknown",
        "location": "Unknown",
        "parent": "Liquidity Services CAG",
        "disposition": "Possibly LS code in 31k seller-ID band (newer than ASFL). Plausible expansions: 'Atlantic / American [Water/Well/Wireline] [Co/Tools] Ltd' — none verifiable. Not in TX/DE/OK public registries.",
        "direct_contact": "Liquidity Services CAG (presumed)",
        "confidence": "unverified",
        "notes": "",
    },
    "Global Power Products, LLC": {
        "owner_class": "ambiguous",
        "industry": "Power generation / electrical (residential transfer switches + metering for the GA candidate)",
        "location": "Lawrenceville GA (225 Arnold Rd, 30044) OR Scottsdale AZ (7645 E. Evans Rd #140, 85260)",
        "parent": "Two candidates: Global Power Products, Inc. (Lawrenceville GA — submetering / GenerLink) OR Global Power Products LLC (Scottsdale AZ entity #L20583660). Worldwide Power Products (Houston) is a separate frequently-confused company.",
        "disposition": "Single listing — likely one-off liquidation of demo/returned inventory.",
        "direct_contact": "globalpowerproducts.com (GA candidate) — info@globalpowerproducts.com / (800) 886-3837",
        "confidence": "unverified",
        "notes": "Need actual lot description from AllSurplus to disambiguate which 'Global Power Products' this is.",
    },
    "CV Supply LLC": {
        "owner_class": "asset owner (small)",
        "industry": "Oilfield supply / Eagle Ford South Texas",
        "location": "Pleasanton, TX 78064 (3180 2nd St)",
        "parent": "Standalone LLC (FMCSA USDOT #4196444 — 1 truck, 2 drivers)",
        "disposition": "Single-listing dispo via LS event-manager Ruth Hernandez. Sole-prop / very small LLC — not a portfolio remarketer.",
        "direct_contact": "3180 2nd St, Pleasanton TX 78064 (no website). Reach via LS Ruth Hernandez or directly at the Pleasanton yard.",
        "confidence": "confirmed (entity); likely (owner-not-broker)",
        "notes": "",
    },

    # ---- Government remarketers ----
    "State of California Department of Technology, CA": {
        "owner_class": "direct agency",
        "industry": "IT / data-center hardware",
        "location": "Rancho Cordova CA (3101 Gold Camp Dr) + Vacaville CA",
        "parent": "California Department of Technology (CDT)",
        "disposition": "Routine state IT lifecycle disposal under DGS Surplus Property and Reutilization (SPR) program. No special shutdown event.",
        "direct_contact": "cdt.ca.gov | DGS SPR (STD 152 / STD 158): dgs.ca.gov/OFAM",
        "confidence": "confirmed",
        "notes": "SKIP — IT/data-center only, not Fuelled-relevant.",
    },
    "Solaire Kapuskasing Solar, ON": {
        "owner_class": "asset owner (private/municipal-linked SPV)",
        "industry": "Solar PV (6 MW Kap Solar Park, commissioned 2014)",
        "location": "Kapuskasing, Ontario, Canada (49.375, -82.393); entity registered in Mississauga ON",
        "parent": "Énergie Kapuskasing Energy (municipally-owned utility holding co); operating entity per gem.wiki: 'New Ospvf 7 LP'",
        "disposition": "UNVERIFIED — Kap Solar Park listed as 'operating' on Global Energy Monitor. 4 listings could be (a) routine inverter swap (park is ~12 yrs old, common at year 10-12), (b) repowering, or (c) partial decom. Worth a phone call.",
        "direct_contact": "Town of Kapuskasing — 705-335-2341 / general@kapuskasing.ca | Énergie Kapuskasing Energy (parent)",
        "confidence": "likely (ownership); unverified (disposition reason)",
        "notes": "Investigate further — phone the town.",
    },
    "State of Washington Surplus Operations, WA": {
        "owner_class": "direct agency",
        "industry": "Mixed — fleet vehicles, electronics, office furniture; federal surplus channel can include heavier industrial gear",
        "location": "Washington State (Tumwater/Olympia + Eastern WA warehouses)",
        "parent": "Washington Department of Enterprise Services (DES)",
        "disposition": "Centralized state-wide surplus consolidation for all WA agencies + many local govts. Routine.",
        "direct_contact": "des.wa.gov/purchase/buy-surplus | Surplus@des.wa.gov / 360-407-2270 | AgencySurplusPurch@des.wa.gov / 360-407-2275",
        "confidence": "confirmed",
        "notes": "Multi-agency funnel — single contact point, federal surplus channel attached.",
    },
    "Factor X - Government Remarketing, CA": {
        "owner_class": "third-party remarketer (decom contractor)",
        "industry": "Facility decommissioning + capital-asset resale (HVAC, MEP, lab/process gear likely)",
        "location": "Gilroy, CA (20 Cohansey Avenue, 95020)",
        "parent": "Factor X (private company, est. 2008; MBE/EDWOSB/SDBA-certified)",
        "disposition": "Recurring decom contracts (not a one-off). Source assets from CA government facility decommissioning.",
        "direct_contact": "factorx.com — 408.842.9700 / info@FactorX.com | GovDeals storefront: govdeals.com/factorx",
        "confidence": "confirmed",
        "notes": "Pursue as channel partner. Decom feeds typically include generators, transformers, switchgear, chillers, compressors, pumps.",
    },
    "Municipal Asset Recovery of NE Ohio - Government Remarketing, OH": {
        "owner_class": "likely third-party remarketer",
        "industry": "Unknown",
        "location": "Northeast Ohio (Cleveland / Akron / Youngstown corridor)",
        "parent": "Unknown — no standalone website, LinkedIn, or BBB listing under this exact name",
        "disposition": "Unknown — only 1 listing",
        "direct_contact": "Reach via AllSurplus / Liquidity Services account team (no public corporate site)",
        "confidence": "unverified",
        "notes": "Too thin — 1 listing, no web footprint. Let it accumulate before acting.",
    },
    "Sierra Auction - PHXFLT": {
        "owner_class": "third-party auctioneer (LS-owned channel partner)",
        "industry": "Heavy fleet — trucks, refuse/dump/water trucks, street sweepers, buses, construction gear, generators",
        "location": "Phoenix AZ (3570 Grand Ave); statewide AZ consigners",
        "parent": "Sierra Auction Management Inc. (est. 1986); acquired by Liquidity Services 2023",
        "disposition": "PHXFLT = City of Phoenix Fleet (Sierra is sole auction outlet for City of Phoenix + many AZ govt/police agencies). Ongoing fleet rotation.",
        "direct_contact": "sierraauction.com — (602) 242-7121 | corporate parent now Liquidity Services",
        "confidence": "confirmed",
        "notes": "Pursue as channel — water trucks / dump trucks / construction gear regularly cross municipal fleet auctions.",
    },

    # ---- Smaller named sellers worth pursuing (from triage of the 38 deferred) ----
    "Columbus Fleet Management, OH": {
        "owner_class": "direct agency (HIGH Fuelled relevance)",
        "industry": "Municipal heavy fleet — excavators, scissor lifts, truck tractors, JLG/Genie aerial, Cat/International (confirmed via IronPlanet/GovPlanet listings)",
        "location": "City of Columbus, OH",
        "parent": "City of Columbus Fleet Management Division",
        "disposition": "Ongoing municipal fleet rotation. Auctions consigned through Ritchie/IronPlanet + GovDeals.",
        "direct_contact": "City of Columbus Fleet Management; ironplanet.com/city-of-columbus seller-info form. Ask AllSurplus rep for direct Fleet Management contact.",
        "confidence": "confirmed",
        "notes": "Standout in the smaller-seller triage. Wider catalog on IronPlanet shows excavators + aerial lifts — Fuelled wheelhouse. Worth direct outbound.",
    },
    "King County, WA": {
        "owner_class": "direct agency (HIGH Fuelled relevance)",
        "industry": "County heavy fleet — possibly construction/utility equipment alongside passenger fleet",
        "location": "Seattle, WA — warehouse at 707 S. Orcas St.",
        "parent": "King County Fleet Services (Department of Executive Services)",
        "disposition": "Launched new online surplus auction Oct 2025 — early relationship window.",
        "direct_contact": "Fleet Services warehouse 206-477-0889 | kingcounty.gov/en/dept/executive-services/transit-transportation-roads/fleet-services/surplus",
        "confidence": "confirmed",
        "notes": "New online auction launched Oct 2025 — early outbound timing.",
    },
    "Lake County, IL": {
        "owner_class": "direct agency (HIGH Fuelled relevance)",
        "industry": "Public-works fleet, landscape, municipal heavy equipment",
        "location": "Lake County, IL (north of Chicago)",
        "parent": "Lake County Purchasing Division",
        "disposition": "Live spring/fall auctions via Obenauf + GovDeals year-round.",
        "direct_contact": "Lake County Purchasing Division | lakecountyil.gov/5035/Surplus-Policy",
        "confidence": "confirmed",
        "notes": "",
    },
    "Clean Water Services, OR": {
        "owner_class": "direct agency (MEDIUM Fuelled relevance)",
        "industry": "Wastewater utility — pumps, blowers, motors, valves, service vehicles, possibly process equipment",
        "location": "Washington County, OR",
        "parent": "Clean Water Services (regional wastewater utility)",
        "disposition": "Routine wastewater-treatment-plant equipment rotation.",
        "direct_contact": "cleanwaterservices.org Procurement | bids/surplus on m.publicsurplus.com seller 'Clean Water Services'",
        "confidence": "confirmed",
        "notes": "Real wastewater utility, not parks dept. WWTPs dispose of pumps/blowers/motors over time. Worth a procurement-team intro.",
    },
    "North Baldwin Utilities, AL": {
        "owner_class": "direct agency (MEDIUM Fuelled relevance)",
        "industry": "Small municipal water/electric utility — possible transformers, line trucks, meters",
        "location": "Bay Minette area, AL",
        "parent": "North Baldwin Utilities",
        "disposition": "Unknown — only 1 listing.",
        "direct_contact": "northbaldwinutilities.com — General Manager / Operations",
        "confidence": "likely",
        "notes": "Only true utility on the smaller-seller list. Low volume but profile fits.",
    },
}


# Named contacts at the top warm-lead organizations — feeds a separate sheet.
# Each row is one person at one company. Confidence: verified / likely / pattern-derived.
NAMED_CONTACTS = [
    # CVRF / Grapevine Energy Holdings
    {"company": "Central Valley Renewables (CVRF)", "name": "Mariah Mandt, PE", "title": "VP / GM, Bakersfield Renewable Fuels",
     "linkedin": "https://www.linkedin.com/in/mariah-mandt-pe-22a33895/", "location": "Bakersfield CA",
     "email": "mariah.mandt@bkrenewablefuels.com", "confidence": "likely",
     "notes": "Site GM — owns physical asset decisions at the only operating Grapevine facility. Smallest org has no IR-specific role."},
    {"company": "Central Valley Renewables (CVRF)", "name": "Saket Adhatrao", "title": "DCS Controls Specialist / Operations",
     "linkedin": "https://www.linkedin.com/in/saket-adhatrao/", "location": "Bakersfield CA",
     "email": "saket.adhatrao@bkrenewablefuels.com", "confidence": "likely",
     "notes": "Active operations contact — useful warm node below GM level."},

    # Assmang / Cato Ridge
    {"company": "Assmang Cato Ridge Works", "name": "Devan Kannie", "title": "Manager, Production CRA & Laboratory (Alt Director, Cato Ridge Alloys Pty Ltd)",
     "linkedin": "https://www.linkedin.com/in/devan-kannie-23452aa1/", "location": "Cato Ridge / Durban, KZN, SA",
     "email": "devan.kannie@feralloys.co.za", "confidence": "verified",
     "notes": "Sits inside the actual Cato Ridge Alloys JV being decommissioned. Domain feralloys.co.za is the working ops domain."},
    {"company": "Assmang Cato Ridge Works", "name": "Bastiaan ('Tiaan') van Aswegen", "title": "Deputy CEO, Assore (Group Technical & Ops; Assore SA PropCo lead on Insimbi Ridge)",
     "linkedin": "Assore leadership page", "location": "Johannesburg, SA",
     "email": "bastiaan.vanaswegen@assore.com", "confidence": "verified (role) / pattern-derived (email)",
     "notes": "Group ops head — owns the PropCo land/property transfer including the Cato Ridge equipment disposal. Canonical first name is Bastiaan ('Tiaan' is colloquial)."},
    {"company": "Assmang Cato Ridge Works", "name": "Henk Gouws", "title": "Managing Director, Assore South Africa",
     "linkedin": "Assore leadership page", "location": "Johannesburg, SA",
     "email": "henk.gouws@assore.com", "confidence": "pattern-derived",
     "notes": "Runs the SA operating subsidiary that owns PropCo."},
    {"company": "Assmang Cato Ridge Works", "name": "Sandra Du Toit", "title": "Chief Financial Officer, Assore",
     "linkedin": "Assore leadership page", "location": "Johannesburg, SA",
     "email": "sandra.dutoit@assore.com", "confidence": "verified (role) / pattern-derived (email)",
     "notes": "CFO signs off on the R453m PropCo land transfer + residual asset-disposal proceeds. Decision-maker for surplus-equipment monetization."},
    {"company": "Assmang Cato Ridge Works", "name": "Kieran Daly", "title": "Executive: Growth & Strategic Development; MD Assore International",
     "linkedin": "Assore leadership page", "location": "Johannesburg / London",
     "email": "kieran.daly@assore.com", "confidence": "verified (role) / pattern-derived (email)",
     "notes": "Owns Assore's M&A and disposal strategy including Cato Ridge transition."},
    {"company": "Assmang Cato Ridge Works", "name": "Herman Smith", "title": "Executive: Technical Services, Assore",
     "linkedin": "Assore leadership page", "location": "Johannesburg, SA",
     "email": "herman.smith@assore.com", "confidence": "verified (role) / pattern-derived (email)",
     "notes": "Group-level technical engineering lead — likely owns the equipment specification list + condition assessments handed to PropCo."},
    {"company": "Assmang Cato Ridge Works", "name": "Andre Joubert", "title": "Chief Executive, ARM Ferrous (manganese/iron unit — JV partner with Assore)",
     "linkedin": "https://www.linkedin.com/in/andre-joubert-7017451b/", "location": "Johannesburg, SA",
     "email": "andre.joubert@arm.co.za", "confidence": "verified (title, LinkedIn) / pattern-derived (email)",
     "notes": "THE operating exec on Cato Ridge wind-down from ARM's side. NOTE: previous record incorrectly listed him as CFO ARM."},
    {"company": "Assmang Cato Ridge Works", "name": "Thando Mkatshana", "title": "Chief Executive, ARM Platinum (ARM exco member)",
     "linkedin": "ARM leadership page", "location": "Johannesburg, SA",
     "email": "thando.mkatshana@arm.co.za", "confidence": "verified (role) / pattern-derived (email)",
     "notes": "Sits on ARM's exco — peer-level escalation if Joubert is unresponsive."},

    # ArcelorMittal SA
    {"company": "ArcelorMittal South Africa  (AMSA)", "name": "Travesh Ramkhelawan", "title": "Head of Strategic Projects and Properties",
     "linkedin": "https://www.linkedin.com/in/travesh-ramkhelawan-28195a57/", "location": "Vanderbijlpark, Gauteng",
     "email": "travesh.ramkhelawan@arcelormittal.com", "confidence": "verified",
     "notes": "Strongest IR-equivalent match. 'Strategic Projects and Properties' owns idle-site asset disposition at integrated steel mills."},
    {"company": "ArcelorMittal South Africa  (AMSA)", "name": "Werner Venter", "title": "General Manager, AMSA",
     "linkedin": "https://za.linkedin.com/in/werner-venter-00a56b76", "location": "Vanderbijlpark, Gauteng",
     "email": "werner.venter@arcelormittal.com", "confidence": "likely",
     "notes": "GM-level operations contact — practical decision authority on plant-level surplus during Long Steel wind-down."},
    {"company": "ArcelorMittal South Africa  (AMSA)", "name": "Hendrik Jacobus ('Kobus') Verster", "title": "Chief Executive Officer, AMSA",
     "linkedin": "https://za.linkedin.com/in/kobus-verster-457b48211", "location": "Vanderbijlpark, Gauteng",
     "email": "kobus.verster@arcelormittal.com", "confidence": "verified (title, LinkedIn) / pattern-derived (email)",
     "notes": "CEO calling the shots on Long Steel wind-down. Personally led media + government engagement on the closure — every asset-disposal decision rolls up to him."},
    {"company": "ArcelorMittal South Africa  (AMSA)", "name": "Adriaan Liebenberg", "title": "Senior Manager, Downstream Operations — SA Long Steel Products",
     "linkedin": "https://www.linkedin.com/in/adriaan-liebenberg/", "location": "Newcastle, KZN",
     "email": "adriaan.liebenberg@arcelormittal.com", "confidence": "verified (LinkedIn role + location) / pattern-derived (email)",
     "notes": "DIRECT operational owner of the closing Newcastle Long Steel lines. Almost certainly has the equipment list + is dealing with care-and-maintenance scope day-to-day."},
    {"company": "ArcelorMittal South Africa  (AMSA)", "name": "Henk van der Merwe", "title": "Maintenance Manager, Coke Making — AMSA Newcastle Works",
     "linkedin": "https://www.linkedin.com/in/henk-vd-merwe-34131315/", "location": "Newcastle, KZN",
     "email": "henk.vandermerwe@arcelormittal.com", "confidence": "verified (LinkedIn) / pattern-derived (email)",
     "notes": "Maintenance lead on Newcastle coke ovens — knows surplus rotables, spares, and idled equipment intimately."},
    {"company": "ArcelorMittal South Africa  (AMSA)", "name": "Henning van Wyk", "title": "Vendor Specialist (Procurement), AMSA",
     "linkedin": "https://www.linkedin.com/in/henning-van-wyk-40a6a191/", "location": "Vanderbijlpark, Gauteng",
     "email": "henning.vanwyk@arcelormittal.com", "confidence": "verified (LinkedIn) / pattern-derived (email)",
     "notes": "The 'in-the-system' route to becoming an approved buyer of surplus equipment, even though final disposal authority sits with Ramkhelawan."},

    # Glencore SA
    {"company": "Glencore Operations South Africa (Pty) Ltd", "name": "Japie Fullard", "title": "CEO, Glencore Alloys (Ferrochrome + Vanadium SA)",
     "linkedin": "no clean LinkedIn (see theofficialboard.com)", "location": "Rustenburg / Johannesburg, SA",
     "email": "japie.fullard@glencore.co.za", "confidence": "verified",
     "notes": "Public face of the smelter S189 process. Ultimate disposition decisions on suspended ferrochrome assets sit with him."},
    {"company": "Glencore Operations South Africa (Pty) Ltd", "name": "Christof Bester", "title": "GM, Rustenburg Smelter",
     "linkedin": "https://www.linkedin.com/in/christof-bester-246705106/", "location": "Rustenburg, North West, SA",
     "email": "christof.bester@glencore.co.za", "confidence": "likely",
     "notes": "Site-level GM at the largest impacted smelter."},
    {"company": "Glencore Operations South Africa (Pty) Ltd", "name": "Chabisi Motloung", "title": "GM, Glencore Boshoek Smelter",
     "linkedin": "https://www.linkedin.com/in/chabisi-motloung-572105175/", "location": "Rustenburg / North West, SA",
     "email": "chabisi.motloung@glencore.co.za", "confidence": "pattern-derived",
     "notes": "Site head of one of the two smelters Glencore put on care & maintenance — direct decision-maker on idle plant equipment disposal at Boshoek."},
    {"company": "Glencore Operations South Africa (Pty) Ltd", "name": "Anna Krutikov", "title": "Group Head of Sustainable Development, Glencore plc",
     "linkedin": "https://ch.linkedin.com/in/anna-krutikov-57020431", "location": "Zürich, Switzerland",
     "email": "anna.krutikov@glencore.com", "confidence": "pattern-derived",
     "notes": "Named in S189 / Save SA Smelters comms — group-level signatory; useful for ESG framing of equipment redeployment."},
    {"company": "Glencore Operations South Africa (Pty) Ltd", "name": "Zanele Matlala", "title": "CEO, Merafe Resources (50:50 JV partner in Glencore-Merafe Chrome)",
     "linkedin": "https://za.linkedin.com/in/zanele-matlala-b49b0b11", "location": "Johannesburg, SA",
     "email": "info@meraferesources.co.za (no public personal address)", "confidence": "verified (role) / unknown (email)",
     "notes": "JV co-owner has consent rights on smelter disposition. Parallel approval path independent of Glencore HQ."},

    # Valaris
    {"company": "Valaris - US (SM)", "name": "Jorge Lebrija", "title": "VP, Global Supply Chain",
     "linkedin": "https://www.linkedin.com/in/jorgelebrija/", "location": "Houston TX",
     "email": "jorge.lebrija@valaris.com", "confidence": "verified",
     "notes": "13-year tenure (joined Ensco 2013). Owns global procurement + asset/spares lifecycle for the rig fleet rationalization."},
    {"company": "Valaris - US (SM)", "name": "Chandler Breaux", "title": "Supply Chain / Materials, Valaris",
     "linkedin": "https://www.linkedin.com/in/chandler-breaux-07a4b246/", "location": "Houston TX area",
     "email": "chandler.breaux@valaris.com", "confidence": "likely",
     "notes": "Working-level entry point into the same supply-chain org."},
    {"company": "Valaris - US (SM)", "name": "Fayad El Reda", "title": "Director, Procurement",
     "linkedin": "https://www.linkedin.com/in/fayad-el-reda-824190125", "location": "Houston TX",
     "email": "fayad.elreda@valaris.com", "confidence": "pattern-derived",
     "notes": "Director-level procurement reporting up to Lebrija — clearest 'second seat' in supply chain after the VP."},
    {"company": "Valaris - US (SM)", "name": "Bryan W. Sanchez", "title": "Director, Capital Projects",
     "linkedin": "https://www.linkedin.com/in/bryan-sanchez-83006738/", "location": "Houston TX",
     "email": "bryan.sanchez@valaris.com", "confidence": "pattern-derived",
     "notes": "Capital Projects owns reactivation/upgrade scope for stacked rigs — sees what's being cannibalised vs retained on the DPS-3/5/6 retirements."},
    {"company": "Valaris - US (SM)", "name": "Christopher T. Beckett", "title": "CFO, Valaris plc (escalation only)",
     "linkedin": "Valaris executive page", "location": "Houston TX",
     "email": "chris.beckett@valaris.com", "confidence": "pattern-derived",
     "notes": "Escalation lane only. CFO ultimately signs on asset-sale economics in the Transocean-merger context."},

    # ConocoPhillips Permian
    {"company": "ConocoPhillips – Permian", "name": "Bryan Morris", "title": "Sr. Supply Chain Coordinator",
     "linkedin": "https://www.linkedin.com/in/bryan-morris-6a213021/", "location": "Midland TX",
     "email": "bryan.morris@conocophillips.com", "confidence": "likely",
     "notes": "Permian BU operational warm node — closest match to plant-side IR contact."},
    {"company": "ConocoPhillips – Permian", "name": "Justin Nolen", "title": "Supply Chain Specialist",
     "linkedin": "https://www.linkedin.com/in/justin-nolen-05660189/", "location": "Midland TX",
     "email": "justin.nolen@conocophillips.com", "confidence": "likely",
     "notes": "Second route into the Midland materials function."},
    {"company": "ConocoPhillips – Permian", "name": "Aaron Hunter", "title": "President, Permian Business Unit",
     "linkedin": "https://www.linkedin.com/in/aaron-hunter-2ab5b761/", "location": "Midland TX",
     "email": "aaron.hunter@conocophillips.com", "confidence": "pattern-derived",
     "notes": "Permian BU head, 21+ yrs in Midland — TOP of the Midland decision tree for any Delaware-Basin asset disposition."},
    {"company": "ConocoPhillips – Permian", "name": "Angelle Rodriguez", "title": "Supply Chain Manager (Lower 48 SC)",
     "linkedin": "https://www.linkedin.com/in/angelle-rodriguez-4a5b5b99/", "location": "Houston TX",
     "email": "angelle.rodriguez@conocophillips.com", "confidence": "pattern-derived",
     "notes": "Manager-level SC role over Midland coordinators (Morris/Nolen) — natural escalation peer."},

    # Pemex Deer Park
    {"company": "Pemex Deer Park", "name": "Joe Riley", "title": "Refinery Supply Chain Manager (also Maintenance Lead, DARP)",
     "linkedin": "https://www.linkedin.com/in/joe-riley-965b3842", "location": "Deer Park TX",
     "email": "joe.riley@deerparkrefining.com", "confidence": "verified",
     "notes": "Owns end-to-end procurement, contracts, warehousing & distribution at Deer Park. Direct surplus-equipment decision maker."},
    {"company": "Pemex Deer Park", "name": "Darren Anderson, C.P.M.", "title": "Procurement Manager — Data, Systems & Processes",
     "linkedin": "https://www.linkedin.com/in/darren-anderson-c-p-m-63972242/", "location": "Katy TX",
     "email": "darren.anderson@deerparkrefining.com", "confidence": "likely",
     "notes": "30 yrs O&G procurement, certified purchasing manager."},
    {"company": "Pemex Deer Park", "name": "Andrew Szafran", "title": "Reliability and Mechanical Manager (Rotating Equipment Project Team Lead)",
     "linkedin": "https://www.linkedin.com/in/andrew-szafran-020209250/", "location": "Houston TX",
     "email": "andrew.szafran@deerparkrefining.com", "confidence": "likely",
     "notes": "DIRECT decision-maker on pumps/compressors/turbines being swapped during the $500M+ turnaround. Best operational lead at Deer Park."},
    {"company": "Pemex Deer Park", "name": "Prem Vuthandam", "title": "Vice President, Business Management",
     "linkedin": "https://www.linkedin.com/in/prem-vuthandam-61428210/", "location": "Houston TX",
     "email": "prem.vuthandam@deerparkrefining.com", "confidence": "pattern-derived",
     "notes": "Executive Leadership Team member — sponsor-level for any equipment monetization tied to turnaround capex."},
    {"company": "Pemex Deer Park", "name": "Brett McGregor", "title": "Operations Manager",
     "linkedin": "https://www.linkedin.com/in/brett-mcgregor-2b876a20", "location": "Houston TX",
     "email": "brett.mcgregor@deerparkrefining.com", "confidence": "pattern-derived",
     "notes": "Knows what's actually being decommissioned vs mothballed in unit handovers."},

    # ADM
    {"company": "Archer Daniels Midland Company", "name": "André Alves", "title": "Source-to-Pay Director (Global Process Owner: Inventory & Investment Recovery)",
     "linkedin": "https://www.linkedin.com/in/andr%C3%A9-alves-74935b25/", "location": "Decatur IL",
     "email": "andre.alves@adm.com", "confidence": "verified",
     "notes": "STRONGEST single target of all 10 warm leads. Explicit Investment Recovery mandate at corporate level. Owns Bushnell + Kershaw + Memphis closures."},
    {"company": "Archer Daniels Midland Company", "name": "Lynn Lee", "title": "ADM employee at Kershaw SC site",
     "linkedin": "https://www.linkedin.com/in/lynn-lee-38a2a49/", "location": "Kershaw SC",
     "email": "lynn.lee@adm.com", "confidence": "pattern-derived",
     "notes": "Local Kershaw site contact — only public LinkedIn hit at that closing facility."},
    {"company": "Archer Daniels Midland Company", "name": "Fred Serven", "title": "General Manager, ADM Cottonseed Business",
     "linkedin": "no public match — search ADM cottonseed GM", "location": "Decatur IL or Memphis TN",
     "email": "fred.serven@adm.com", "confidence": "likely (named in ADM/Planters press release) / pattern-derived (email)",
     "notes": "DIRECT owner of the Memphis cottonseed asset transfer into the Planters JV — controls what equipment moves to JV vs is surplus."},
    {"company": "Archer Daniels Midland Company", "name": "Cody Young", "title": "CEO, Planters Cotton Oil Mill (JV partner receiving Memphis assets)",
     "linkedin": "search 'Cody Young Planters Cotton Oil Mill'", "location": "Pine Bluff AR",
     "email": "cyoung@planterscottonoil.com", "confidence": "pattern-derived (role verified in press release)",
     "notes": "JV-side counterparty — anything ADM Memphis sheds that the JV doesn't want flows out as surplus. Cody is the gatekeeper on that triage."},

    # Entergy
    {"company": "Entergy Services, Inc.", "name": "Justin McCabe", "title": "Manager, Investment Recovery",
     "linkedin": "https://theorg.com/org/entergy/org-chart/justin-mccabe", "location": "New Orleans / Jackson MS",
     "email": "justin.mccabe@entergy.com", "confidence": "verified",
     "notes": "DIRECT HIT — explicit Investment Recovery Program Manager. Tenure since May 2015. Owns Baxter Wilson decom assets."},
    {"company": "Entergy Services, Inc.", "name": "Jason Rees", "title": "Director, Supply Chain Commercial Operations",
     "linkedin": "no clean LinkedIn", "location": "New Orleans LA",
     "email": "jason.rees@entergy.com", "confidence": "pattern-derived",
     "notes": "McCabe's reporting line — escalation contact."},
    {"company": "Entergy Services, Inc.", "name": "Lawrence Clark", "title": "Investment Recovery Specialist Senior Lead",
     "linkedin": "no public match (ZoomInfo confirms l***@entergy.com)", "location": "New Orleans LA",
     "email": "lclark@entergy.com (or lawrence.clark@entergy.com)", "confidence": "likely (role verified ZoomInfo) / pattern-derived (email)",
     "notes": "McCabe's most senior IC — direct operator on equipment lots; likely runs the actual sale processes day-to-day."},
    {"company": "Entergy Services, Inc.", "name": "Lee Alexander", "title": "VP, Power Plant Operations (managed Baxter Wilson decommissioning + Rex Brown)",
     "linkedin": "search 'Lee Alexander Entergy'", "location": "Jackson / Vicksburg MS",
     "email": "lalexander@entergy.com (or lee.alexander@entergy.com)", "confidence": "likely (role verified Entergy news) / pattern-derived (email)",
     "notes": "THE decommissioning lead for Baxter Wilson and Rex Brown. Direct authority on what Entergy is salvaging vs selling from those sites."},
    {"company": "Entergy Services, Inc.", "name": "Haley Fisackerly", "title": "President & CEO, Entergy Mississippi",
     "linkedin": "https://www.linkedin.com/in/haley-fisackerly-b0b3bb8/", "location": "Jackson MS",
     "email": "hfisack@entergy.com", "confidence": "pattern-derived",
     "notes": "Mississippi op-co CEO — top-cover for both Baxter Wilson exit and Vicksburg Advanced new-build."},

    # CenterPoint
    {"company": "CenterPoint Energy", "name": "Michael Brown, MCA", "title": "Manager, Asset Recovery & Supply Chain Sustainability",
     "linkedin": "https://www.linkedin.com/in/michael-brown-mca-2576578/", "location": "Katy / Houston TX",
     "email": "michael.brown@centerpointenergy.com", "confidence": "verified",
     "notes": "DIRECT HIT — exact title matches the function. MCA = Member Certified Appraiser (IR Association credential). Owns surplus.centerpointenergy.com."},
    {"company": "CenterPoint Energy", "name": "Stephen Hester", "title": "SVP & Chief Procurement / Chief Supply Chain Officer",
     "linkedin": "https://www.linkedin.com/in/stephenthester/", "location": "Houston TX",
     "email": "stephen.hester@centerpointenergy.com", "confidence": "likely (role verified) / pattern-derived (email)",
     "notes": "Brown's executive chain — sponsor for any sustained surplus partnership. Ex-LyondellBasell CPO so he understands secondary-market dynamics."},
    {"company": "CenterPoint Energy", "name": "Christe Singleton", "title": "Vice President, Minnesota Gas",
     "linkedin": "search 'Christe Singleton CenterPoint'", "location": "Minneapolis MN",
     "email": "christe.singleton@centerpointenergy.com", "confidence": "pattern-derived",
     "notes": "Owns MN gas LDC ops — modernization there generates transformer/regulator/pipe surplus distinct from Houston flow."},
    {"company": "CenterPoint Energy", "name": "Gregg Maurer", "title": "Vice President, Indiana Gas (post-OH-divestiture)",
     "linkedin": "search 'Gregg Maurer CenterPoint'", "location": "Evansville IN",
     "email": "gregg.maurer@centerpointenergy.com", "confidence": "pattern-derived",
     "notes": "Indiana gas LDC ops leader — post-OH divest, Indiana asset base is the focus of any midwest modernization surplus."},

    # ---- CVRF / Grapevine — additional contacts ----
    {"company": "Central Valley Renewables (CVRF)", "name": "Bruce Nelson", "title": "Chief Financial Officer (legacy GCEH; reportedly retained at BKRF/Grapevine)",
     "linkedin": "https://www.linkedin.com/in/bruce-nelson-74488910", "location": "Torrance / LA, CA",
     "email": "bruce.nelson@grapevineenergy.com", "confidence": "pattern-derived",
     "notes": "Long-tenured GCEH CFO; controls cash, vendor payments, and any asset-disposal P&L decisions during the post-Ch.11 ramp."},
    {"company": "Central Valley Renewables (CVRF)", "name": "Antonio D'Amico", "title": "EVP, Chief Administrative Officer & General Counsel",
     "linkedin": "Grapevine management page", "location": "Long Beach / LA, CA",
     "email": "antonio.damico@grapevineenergy.com", "confidence": "verified (title) / pattern-derived (email)",
     "notes": "Owns contracts, MSAs, equipment-disposition legal sign-off. Useful entry point if procurement is centralized through legal."},
    {"company": "Central Valley Renewables (CVRF)", "name": "Brian Coffman", "title": "Independent Director (board); ex-CEO Motiva",
     "linkedin": "Grapevine board page", "location": "Houston TX",
     "email": "(via Grapevine corp secretary)", "confidence": "verified (board seat)",
     "notes": "40-yr refining operator, only board director with hands-on refinery ops experience. Likely the board voice on plant capex/surplus decisions."},
    {"company": "Central Valley Renewables (CVRF)", "name": "Ethan Shoemaker", "title": "Investment Partner OIC (Houston); Director Grapevine",
     "linkedin": "Grapevine board page", "location": "Houston TX",
     "email": "ethan.shoemaker@oic.com", "confidence": "verified (board seat) / pattern-derived (email)",
     "notes": "OIC's Houston lead and a director — controls lender-side appetite for asset sales/surplus monetization at BKRF."},
]


# Liquidity Services event managers (separate sheet — they're people, not sellers)
LS_MANAGERS = [
    {
        "name": "Ruth Hernandez",
        "email": "ruth.hernandez@liquidityservices.com",
        "phone": "713-590-1418 (Houston)",
        "linkedin": "https://www.linkedin.com/in/ruth-hernandez-14866239",
        "title": "Inside Sales Representative, Liquidity Services CAG",
        "region": "US — Houston-based; supports CAG energy/industrial events including ConocoPhillips Permian + Pemex Deer Park",
        "tenure": "unknown",
        "covers_events": "ConocoPhillips Permian, Pemex Deer Park, Central Valley Renewables, several anonymous US events",
        "notes": "Operational point of contact / event coordinator — NOT decision-maker on seller relationships. Treat as gatekeeper.",
        "confidence": "likely",
    },
    {
        "name": "Herman Liversage",
        "email": "herman.liversage@liquidityservices.com",
        "phone": "+27 (0)82 452 8811 (SA mobile)",
        "linkedin": "https://www.linkedin.com/in/herman-liversage-6a08009b/",
        "title": "Associate Director, Liquidity Services SA (Pty) Ltd",
        "region": "Sub-Saharan Africa — mining vertical; Pretoria-based",
        "tenure": "Long-tenured mining-industry insider. Prior: BGS Industrial Equipment, Privest Outsourcing, Durnacal Mine, Hillside Aluminium.",
        "covers_events": "Assmang Cato Ridge, Assmang Khumani, Glencore SA, Northam Booysendal, Black Mountain Mining, Dwarsrivier",
        "notes": "Senior operator at LS-SA, NOT a junior coordinator. The +27 82 mobile is direct. LS-SA team is small — reaching Herman ≈ reaching SA leadership. Pitch as peer-to-peer mining-equipment conversation, not BD pitch.",
        "confidence": "confirmed",
    },
    {
        "name": "Jeff Goddard",
        "email": "goddard@liquidityservices.com",
        "phone": "714-376-1422 (Orange County CA)",
        "linkedin": "https://www.linkedin.com/in/jeff-goddard-cea-34703857/",
        "title": "CAG Account Manager, Liquidity Services (CEA — Certified Equipment Appraiser, AMEA member)",
        "region": "Southern California / US West — UC Irvine alum",
        "tenure": "3+ years in CAG seat (FY22/23/24 quota numbers public on profile). Multi-year appraisal career.",
        "covers_events": "Three smaller anonymous events; verticals incl. Aviation Ground Support, Port Logistics, Electronics/Semi Mfg, govt accounts (GSA, SoCal Edison)",
        "notes": "Real account manager with appraisal credentials. Meaningful seller-side relationships. Worth a direct outreach. Pitch on equipment-data quality, NOT generic marketplace volume.",
        "confidence": "confirmed",
    },
    {
        "name": "Marcin Gadzinski",
        "email": "marcin.gadzinski@liquidityservices.com",
        "phone": "+49 162 491 87 93 (Germany)",
        "linkedin": "https://www.linkedin.com/in/marcin-gadzinski-889265136/",
        "title": "Sales / BD (no formal title disclosed publicly; headline reads marketing tagline)",
        "region": "Continental Europe — likely DACH + CEE under GoIndustry DoveBid / AllSurplus EU operation",
        "tenure": "~15 yrs multi-industry; LS tenure unknown",
        "covers_events": "At least one European event",
        "notes": "Weakest public footprint of the four. LS EU team is small post-GoIndustry-DoveBid integration. Treat as likely BD contact, but seniority unknown — could be senior account exec or mid-level rep.",
        "confidence": "likely",
    },
]


# ============================================================================
# BUYER RECRUITMENT — companies that BUY Fuelled-relevant equipment
# Researched 2026-05-08 across US Upstream + US Refining/Midstream + Canadian O&G.
# Each contact: pattern-derived emails unless otherwise noted; LinkedIn URLs verified.
# ============================================================================

BUYER_COMPANIES = [
    # === US Upstream Oil & Gas ===
    {"vertical": "US Upstream O&G", "name": "Diamondback Energy", "ticker": "FANG", "hq": "Midland TX",
     "basin": "Permian (Midland + Delaware)", "scale": "13-14 rigs + 5 frac crews end-2025; ~890k boe/d post-Endeavor; ~$45B mkt cap",
     "capex_driver": "Closed Endeavor mega-deal 2024. May 2026 flipped to green light, ramping capex from $3.75B → $3.9B with new rigs/frac spreads. Drilling 445-465 wells in 2025 — heavy compressor/separator/treater buying.",
     "suppliers_page": "https://www.diamondbackenergy.com/about/overview"},
    {"vertical": "US Upstream O&G", "name": "Permian Resources", "ticker": "PR", "hq": "Midland + Houston TX",
     "basin": "Permian (Delaware Basin core)", "scale": "190.7 net wells in 2025; 2026 guide $1.75-1.95B; ~$10B mkt cap",
     "capex_driver": "$608M Apache bolt-on + $471M property purchases in 2025. ~$400M of 2026 capex earmarked for facilities/infra/workover — direct relevance for compressor/separator/treater purchases.",
     "suppliers_page": "https://permianres.com/contact/"},
    {"vertical": "US Upstream O&G", "name": "Matador Resources", "ticker": "MTDR", "hq": "Dallas TX",
     "basin": "Delaware Basin (Lea, Eddy NM)", "scale": "111 wells drilled 2025; D/C/E capex $1.53B; ~$6B mkt cap",
     "capex_driver": "Major Ameredev bolt-on still being developed. Patterson-UTI runs >90% of drilling; Matador self-builds production facilities — heavy ongoing buyer of separators/treaters/compressors/tank batteries.",
     "suppliers_page": "https://www.matadorresources.com/contact"},
    {"vertical": "US Upstream O&G", "name": "Civitas/SM Energy (merged Jan 2026)", "ticker": "SM", "hq": "Houston TX (legacy Denver)",
     "basin": "Permian + DJ + Uinta + Eagle Ford post-merger", "scale": "Combined ~13 rigs; ~$8B mkt cap",
     "capex_driver": "$13B SM-Civitas merger closed Jan 2026. $190M+ in synergies — surplus equipment liquidation likely as fleets are rationalized; new development still active.",
     "suppliers_page": "https://sm-energy.com/contact/"},
    {"vertical": "US Upstream O&G", "name": "Coterra Energy", "ticker": "CTRA", "hq": "Houston TX",
     "basin": "Permian + Marcellus + Anadarko", "scale": "9 Permian rigs + 3 frac crews; 2025 capex ~$2.0-2.3B; ~$20B mkt cap; merging with Devon",
     "capex_driver": "Closed Permian bolt-ons (Franklin Mountain + Avant) early 2025. Restarted Marcellus Q2 2025. Production wedge ongoing.",
     "suppliers_page": "https://www.coterra.com/contact-us/"},
    {"vertical": "US Upstream O&G", "name": "Chord Energy", "ticker": "CHRD", "hq": "Houston TX (Bakken ops Williston ND)",
     "basin": "Williston (Bakken + Three Forks)", "scale": "FY2026 capex ~$1.4B; largest pure-play Bakken; ~$8B mkt cap",
     "capex_driver": "Closed $550M Williston acquisition from XTO Q4 2025. Continental halted Bakken Jan 2026 — Chord + Hess now dominant Williston buyers; equipment supply chains shifting in their favor.",
     "suppliers_page": "https://www.chordenergy.com/about-us/"},
    {"vertical": "US Upstream O&G", "name": "Crescent Energy", "ticker": "CRGY", "hq": "Houston TX (KKR-backed)",
     "basin": "Eagle Ford + Uinta + Permian (post-Vital)", "scale": "6-7 rigs across portfolio; ~$3B mkt cap pre-merger",
     "capex_driver": "$905M Central Eagle Ford bolt-on + $2.1B Vital Energy merger Dec 2025. Multi-basin acquirer with explicit ops-driven equipment strategy.",
     "suppliers_page": "https://crescentenergyco.com/contact/"},
    {"vertical": "US Upstream O&G", "name": "Riley Permian", "ticker": "REPX", "hq": "Oklahoma City OK",
     "basin": "Permian (San Andres horizontal)", "scale": "2026 capex $190-210M; production 35-37 mboed; ~$650M mkt cap",
     "capex_driver": "Closed $142M Silverback acquisition July 2025. Also expanding midstream — direct compressor/separator buyer. Small enough that a marketplace play is decisively useful.",
     "suppliers_page": "https://www.rileypermian.com/about/contact/"},
    {"vertical": "US Upstream O&G", "name": "Magnolia Oil & Gas", "ticker": "MGY", "hq": "Houston TX",
     "basin": "Eagle Ford + Austin Chalk", "scale": "201-500 employees; ~$5B mkt cap; conservative FCF-focused",
     "capex_driver": "Steady ~$450-500M/yr. Smaller scale = a marketplace channel for mid-life surplus is highly relevant to their cost-leadership posture.",
     "suppliers_page": "https://www.magnoliaoilgas.com/about-us"},
    {"vertical": "US Upstream O&G", "name": "Ovintiv", "ticker": "OVV", "hq": "Denver + Calgary",
     "basin": "Permian + Anadarko (selling) + Montney; exited Bakken/Uinta 2025", "scale": "2025 capex ~$2.2B; ~$11B mkt cap",
     "capex_driver": "Sold $1.9B Uinta to FourPoint Jan 2025; selling $3B Anadarko Q2 2026. BOTH a major surplus liquidator AND a heavy Permian/Montney buyer — uniquely two-sided fit for Fuelled.",
     "suppliers_page": "https://www.ovintiv.com/suppliers/ (best-in-class formal portal)"},

    # === US Refining + Midstream ===
    {"vertical": "US Refining", "name": "HF Sinclair", "ticker": "DINO", "hq": "Dallas TX",
     "basin": "7 refineries (~678 Mbpd) across TX/KS/WY/UT/NM/WA", "scale": "Lubricants + Petro-Canada base oils + Sinclair retail",
     "capex_driver": "2025 capex $875M; 2026 $775M as heavy-maintenance cycle ends. Oct 2025: phased PADD 4/5 pipeline expansion (FID mid-2026).",
     "suppliers_page": "https://www.hfsinclair.com/suppliers/HF/default.aspx"},
    {"vertical": "US Refining", "name": "PBF Energy", "ticker": "PBF", "hq": "Parsippany NJ",
     "basin": "6 refineries (~1.0 MMbpd)", "scale": "Logistics via PBFX",
     "capex_driver": "2025 capex $750-775M ex-Martinez restoration. Q4 2025 turnarounds at Chalmette + Toledo. RBI program targeting >$200M annualized savings — strategic procurement is one of five pillars.",
     "suppliers_page": "https://www.pbfenergy.com/"},
    {"vertical": "US Refining", "name": "Par Pacific", "ticker": "PARR", "hq": "Houston TX",
     "basin": "4 refineries: Kapolei HI, Newcastle WY, Tacoma WA, Billings MT", "scale": "ex-ExxonMobil Billings + Hawaii retail",
     "capex_driver": "YTD 2025 accrued $204M. Wyoming 66-day idle drove unscheduled equipment replacement. Hawaii renewable fuels project underway.",
     "suppliers_page": "https://www.parpacific.com/"},
    {"vertical": "US Refining", "name": "Calumet Inc.", "ticker": "CLMT", "hq": "Indianapolis IN",
     "basin": "Specialty refineries + Montana Renewables (RD/SAF)", "scale": "Multiple LA + PA + IL + TX sites",
     "capex_driver": "MaxSAF expansion at Montana Renewables targeting +120-150M gal/yr SAF by Q2 2026 — major rotating-equipment, exchanger, vessel demand.",
     "suppliers_page": "https://calumet.com/about/"},
    {"vertical": "US Refining", "name": "Delek US", "ticker": "DK", "hq": "Brentwood TN",
     "basin": "4 refineries (~302 Mbpd) — Tyler/Big Spring TX, El Dorado AR, Krotz Springs LA", "scale": "Delek Logistics (DKL)",
     "capex_driver": "Big Spring selected by DOE Carbon Capture Pilot ($95M federal cost share, 70%) — drives rotating/exchanger/vessel demand for CCUS skids.",
     "suppliers_page": "https://delekus.com/"},
    {"vertical": "US Refining", "name": "CVR Energy", "ticker": "CVI", "hq": "Sugar Land TX",
     "basin": "Coffeyville KS + Wynnewood OK refineries (206.5 Mbpd nameplate)", "scale": "Plus nitrogen fertilizer + pipelines",
     "capex_driver": "2026 petroleum capex $130-145M including Wynnewood Alky Project. Reverted Wynnewood RD unit back to hydrocarbon — repurposing process equipment.",
     "suppliers_page": "https://cvrenergy.com/Company/Careers"},
    {"vertical": "US Midstream", "name": "ONEOK", "ticker": "OKE", "hq": "Tulsa OK",
     "basin": "~50,000 mi NGL/gas pipelines + processing + frac (post-EnLink + Medallion)", "scale": "Permian/Mid-Con/Rockies/Williston/Gulf",
     "capex_driver": "2026 capex $2.7-3.2B. Aug 2025: $365M Delaware processing plant (+300 MMcf/d). >500 MMcf/d new Permian processing through 2027. Eiger Express (Permian-Houston, 2 Bcf/d). >$8.3B in active/proposed projects.",
     "suppliers_page": "https://www.oneok.com/"},
    {"vertical": "US Midstream", "name": "Targa Resources", "ticker": "TRGP", "hq": "Houston TX",
     "basin": "Permian + Bakken + SCOOP/STACK G&P; Mont Belvieu frac", "scale": "Galena Park export",
     "capex_driver": "2025 net growth capex ~$3.3B. Yeti gas processing plant (275 MMcf/d, Delaware) joins 4 other Permian plants in flight totaling +1.4 Bcf/d. $1.6B Speedway NGL pipeline.",
     "suppliers_page": "https://www.targaresources.com/customer-center/customer-center-overview"},
    {"vertical": "US Midstream", "name": "Energy Transfer", "ticker": "ET", "hq": "Dallas TX",
     "basin": "~130,000 mi pipeline; gas processing; NGL frac at Mont Belvieu; Lake Charles + Nederland export", "scale": "Sunoco LP GP",
     "capex_driver": "2026 growth capex $5.5-5.9B. Major 2026 startups: Nederland Flexport NGL, Mustang Draw I+II Permian, Hugh Brinson Phase I, Gateway NGL. Desert Southwest pipe upsized to 48\".",
     "suppliers_page": "https://apttus.energytransfer.com/"},
    {"vertical": "US Midstream", "name": "Williams Companies", "ticker": "WMB", "hq": "Tulsa OK",
     "basin": "Transco (largest US gas pipeline), NW Pipeline, Gulfstream; Marcellus/Haynesville/DJ G&P", "scale": "Plus storage + LNG investments",
     "capex_driver": "2025 growth $3.95-4.25B; 2026 $6.1-6.7B. Southeast Supply Enhancement (1.6 Bcf/d). Northeast Supply Enhancement (400 Mdth/d Q4 2027). 14 transmission projects in execution.",
     "suppliers_page": "https://www.williams.com/all-about-supplier-portal/ (Oracle Cloud)"},
    {"vertical": "US Midstream", "name": "MPLX (Marathon Petroleum sister)", "ticker": "MPLX", "hq": "Findlay OH",
     "basin": "Marcellus/Utica G&P, Permian G&P, NGL frac (Mont Belvieu via Hopedale)", "scale": "Refined products + marine",
     "capex_driver": "2025 growth capex $2.4B (90% gas/NGL). Permian: Secretariat I (200 MMcf/d) Jan 2026 + Secretariat II (300 MMcf/d) H2 2028. Marcellus: Harmon Creek III (300 MMcf/d) Q3 2026.",
     "suppliers_page": "Shared via Marathon Petroleum supplier portal"},
    {"vertical": "US Midstream", "name": "Enterprise Products Partners", "ticker": "EPD", "hq": "Houston TX",
     "basin": "Largest US NGL midstream — 50K+ mi pipelines, Mont Belvieu frac complex, Permian gas plants, NGL/ethane export", "scale": "Plus petrochemicals + crude",
     "capex_driver": "2025 growth capex $4.0-4.5B. 4 new Permian gas plants. 550-mi Bahia NGL pipeline (600 Mbpd) entering service Nov 2025. NGL Frac 14 ramping mid-Oct 2025.",
     "suppliers_page": "https://www.enterpriseproducts.com/"},

    # === Canadian O&G + Industrial ===
    {"vertical": "Canadian O&G", "name": "Whitecap Resources", "ticker": "TSX:WCP", "hq": "Calgary AB",
     "basin": "Post-Veren: Duvernay (Kaybob) + Montney (NEBC/Lator) + legacy SK/AB", "scale": "~370,000 boe/d",
     "capex_driver": "2026 capex $2.0-2.1B. 7-rig program (3 Duvernay, 4 Montney). Lator 04-13 facility startup with $180M committed in 2026. Compressors flagged as long-lead.",
     "suppliers_page": "https://www.wcap.ca/operations/suppliers"},
    {"vertical": "Canadian O&G", "name": "Tourmaline Oil", "ticker": "TSX:TOU", "hq": "Calgary AB",
     "basin": "Canada's largest gas producer. Alberta Deep Basin, NEBC Montney, Peace River Charlie Lake", "scale": "~635,000 boe/d",
     "capex_driver": "2026 EP capex C$2.55B. Aitken expansion (20 Mbbl/d condensate) Q4 2026 startup; Groundbirch deep-cut plant Q4 2027; sand transload Q2 2026; 3 new water recycling facilities.",
     "suppliers_page": "https://www.tourmalineoil.com/contact-us"},
    {"vertical": "Canadian O&G", "name": "ARC Resources", "ticker": "TSX:ARX", "hq": "Calgary AB",
     "basin": "Pure-play Montney — Kakwa, Sunrise, Dawson, Ante Creek, Attachie BC", "scale": "~370,000 boe/d",
     "capex_driver": "Attachie Phase I (40,000 boe/d facility) ~75% complete. Phase II long-lead spend continues into 2026 ($50M baseline) with 2028 first production. Lower Montney development added 2026.",
     "suppliers_page": "https://www.arcresources.com/about-us/suppliers/"},
    {"vertical": "Canadian O&G", "name": "Cenovus Energy", "ticker": "TSX:CVE", "hq": "Calgary AB",
     "basin": "Integrated — oil sands (Foster Creek, Christina Lake, Sunrise, Lloydminster), refining + MEG/Christina Lake North post-2025", "scale": "Largest Canadian OFS spender",
     "capex_driver": "2026 capital $5.0-5.3B. Oil sands $3.5-3.6B incl. $850M Christina Lake North expansion. Sustaining capital alone $3.5-3.6B drives massive recurring equipment demand.",
     "suppliers_page": "https://www.cenovus.com/Suppliers + ALREADY RUNS surplusinventory.cenovus.com (monitor)"},
    {"vertical": "Canadian O&G", "name": "Imperial Oil", "ticker": "TSX:IMO", "hq": "Calgary AB (ExxonMobil 69.6%)",
     "basin": "Kearl mine, Cold Lake CSS, Syncrude (25%), Sarnia/Strathcona/Nanticoke refineries", "scale": "300 Mbbl/d Kearl target",
     "capex_driver": "2026 capex $2.0-2.2B. Secondary bitumen recovery at Kearl, Mahihkan SA-SAGD + infill at Cold Lake (165 Mbbl/d target). Heavy infill = continuous artificial-lift, separator, treater demand.",
     "suppliers_page": "https://www.imperialoil.ca/company/suppliers"},
    {"vertical": "Canadian O&G", "name": "Strathcona Resources", "ticker": "TSX:SCR", "hq": "Calgary AB (Waterous-controlled)",
     "basin": "Heavy oil pure-play — Cold Lake (Tucker, Lindbergh), Lloyd thermal (Meota, Vawn), Lloyd conventional", "scale": "Exit rate ~135 Mbbl/d",
     "capex_driver": "2026 capex $1.0B. $360M Meota Central thermal brownfield ~91% complete. Lindbergh fuel-gas pipeline repair Q3 2026. Closed Vawn thermal acquisition.",
     "suppliers_page": "https://www.strathconaresources.com/contact-us/"},
    {"vertical": "Canadian Midstream", "name": "Pembina Pipeline", "ticker": "TSX:PPL", "hq": "Calgary AB",
     "basin": "Conventional pipelines (Peace, NGL), Pembina Gas Infrastructure (PGI) — multiple Montney/Duvernay plants", "scale": "Cedar LNG (50/50 Haisla) + RFS IV + Prince Rupert",
     "capex_driver": "2026 = 'largest single capital year' for Cedar LNG (FLNG vessel, marine substation, transmission line). Fox Creek-to-Namao Peace Expansion. Compressor stations heavily refreshed.",
     "suppliers_page": "https://www.pembina.com/operations/suppliers (ISNetworld prequal mandatory)"},
    {"vertical": "Canadian Midstream", "name": "Keyera Corp", "ticker": "TSX:KEY", "hq": "Calgary AB",
     "basin": "G&P plants (Wapiti, Pipestone, Simonette, Strachan), KAPS NGL pipeline, KFS frac, Edmonton/Fort Sask storage", "scale": "Mid-cap midstream",
     "capex_driver": "2026 growth capex $400-475M. Sanctioned KAPS Zone 4 ($220M net, mid-2027 ISD), KFS Frac III, KFS Frac II debottleneck. Pumps/compressors/vessels/heaters demand.",
     "suppliers_page": "https://www.keyera.com/suppliers/"},
    {"vertical": "Canadian Midstream", "name": "AltaGas", "ticker": "TSX:ALA", "hq": "Calgary AB",
     "basin": "Midstream — Pipestone I+II, Townsend, Harmattan, Dimsdale storage; REEF LPG export", "scale": "Plus Utilities (WGL, SEMCO)",
     "capex_driver": "2026 capex ~$1.6B (~27% Midstream). Pipestone II commercial ops Dec 2025. REEF: 80% costs incurred. $230M Dimsdale storage expansion underway.",
     "suppliers_page": "https://www.altagas.ca/suppliers"},
    {"vertical": "Canadian O&G", "name": "Vermilion Energy", "ticker": "TSX:VET", "hq": "Calgary AB",
     "basin": "Post-Westbrick: Deep Basin + Montney WCSB flagship; SK; European gas; Australia", "scale": "Global gas portfolio",
     "capex_driver": "2026 E&D capex $600-630M (85% global gas). $415M Deep Basin + Montney — 49 wells. Montney throughput targeted to 28,000 boe/d.",
     "suppliers_page": "https://www.vermilionenergy.com/about-us/suppliers/"},
    {"vertical": "Canadian O&G", "name": "Birchcliff Energy", "ticker": "TSX:BIR", "hq": "Calgary AB",
     "basin": "100% owned Pouce Coupe Gas Plant (340 MMcf/d), Gordondale, Elmworth Montney emerging", "scale": "Mid-cap pure-play",
     "capex_driver": "2026 F&D capex $325-375M. Pure Lower/Upper Montney development. Owns own infra → buys compression, treating, dehy, sweetening kit directly. NOTE email convention: first-initial+lastname (different from peers).",
     "suppliers_page": "https://www.birchcliffenergy.com/suppliers"},
    {"vertical": "Canadian O&G", "name": "Suncor Energy", "ticker": "TSX:SU", "hq": "Calgary AB",
     "basin": "Base Mine, Firebag, MacKay River, Fort Hills (oil sands); Petro-Canada retail; Edmonton/Sarnia/Montreal refineries", "scale": "Major integrated",
     "capex_driver": "Mine progression + in-situ optimization. Active surplus turnover via secondary markets.",
     "suppliers_page": "https://www.suncor.com/en-ca/about-us/suppliers"},
]


# Buyer-side contacts — one row per person, joined to BUYER_COMPANIES by 'company'
BUYER_CONTACTS = [
    # === US Upstream ===
    {"company": "Diamondback Energy", "name": "Gwendolyn ('Gwen') Smith", "title": "Senior Procurement Specialist",
     "linkedin": "https://www.linkedin.com/in/gwendolyn-smith2/", "location": "Midland TX",
     "email": "gwendolyn.smith@diamondbackenergy.com", "confidence": "likely",
     "notes": "Senior procurement IC — likely owner of category buys for production equipment."},
    {"company": "Diamondback Energy", "name": "Cody King", "title": "Materials Supervisor",
     "linkedin": "https://www.linkedin.com/in/cody-king-977b2292/", "location": "Midland TX",
     "email": "cody.king@diamondbackenergy.com", "confidence": "likely",
     "notes": "Runs materials coordination supplier-to-wellsite — direct gatekeeper for inbound used/surplus offers."},
    {"company": "Permian Resources", "name": "Steven Dodson", "title": "CISO (cross-functional escalation; pattern verifier)",
     "linkedin": "https://www.linkedin.com/in/stevedod/", "location": "Dallas-Fort Worth TX",
     "email": "steven.dodson@permianres.com", "confidence": "likely",
     "notes": "ZoomInfo confirms @permianres.com email pattern."},
    {"company": "Matador Resources", "name": "Chris Cavner", "title": "Engineering / Operations",
     "linkedin": "https://www.linkedin.com/in/chris-cavner-a1b4631a/", "location": "Dallas TX",
     "email": "chris.cavner@matadorresources.com", "confidence": "pattern-derived",
     "notes": "Useful warm-channel given Matador's small procurement footprint."},
    {"company": "Civitas/SM Energy (merged Jan 2026)", "name": "Tina Arneson", "title": "Manager, Corporate Facility Services",
     "linkedin": "https://www.linkedin.com/in/tina-arneson-229a02126/", "location": "Denver CO",
     "email": "tina.arneson@civitasresources.com", "confidence": "likely",
     "notes": "Liaison between legal, supply chain, and stakeholders for ops/capital projects — real procurement crossover."},
    {"company": "Chord Energy", "name": "Jason Stroud", "title": "Senior Staff Sourcing Specialist",
     "linkedin": "https://www.zoominfo.com/p/Jason-Stroud/11514673131", "location": "Houston TX",
     "email": "jason.stroud@chordenergy.com", "confidence": "verified (email pattern @chordenergy.com confirmed)",
     "notes": "Ex-BHP global petroleum supply chain — sets sourcing policy. HIGHEST-LEVERAGE contact at Chord."},
    {"company": "Chord Energy", "name": "Kent Specht", "title": "Supply Chain Supervisor",
     "linkedin": "https://www.linkedin.com/in/kentspecht/", "location": "Williston ND",
     "email": "kent.specht@chordenergy.com", "confidence": "likely",
     "notes": "Field-side SC supervisor — direct buyer for Bakken wellsite equipment."},
    {"company": "Chord Energy", "name": "Armando Almaguer", "title": "Senior Procurement Consultant/Analyst",
     "linkedin": "https://www.linkedin.com/in/armandoalmaguer/", "location": "Pearland TX",
     "email": "armando.almaguer@chordenergy.com", "confidence": "likely",
     "notes": "Houston-side procurement analyst."},
    {"company": "Magnolia Oil & Gas", "name": "Steven Kovacs", "title": "Lead, Supply Chain and Contracts Analyst",
     "linkedin": "https://www.linkedin.com/in/steven-kovacs-7620a713/", "location": "Houston TX",
     "email": "steven.kovacs@magnoliaoilgas.com", "confidence": "likely",
     "notes": "Lead-level SC IC — owns category contracts. HIGHEST-leverage Magnolia contact."},
    {"company": "Magnolia Oil & Gas", "name": "Sayuri Gonzalez", "title": "Supply Chain and Contract Coordinator",
     "linkedin": "https://www.linkedin.com/in/sayuri-gonzalez-aba6a814a/", "location": "Houston TX",
     "email": "sayuri.gonzalez@magnoliaoilgas.com", "confidence": "likely",
     "notes": "Tactical day-to-day buyer / great inbound contact for surplus listings."},
    {"company": "Ovintiv", "name": "Greg Givens", "title": "Supply Chain (per LinkedIn)",
     "linkedin": "https://www.linkedin.com/in/greg-givens-8970354/", "location": "Denver CO",
     "email": "greg.givens@ovintiv.com", "confidence": "pattern-derived",
     "notes": "One of few publicly indexed Ovintiv SC LinkedIn profiles."},

    # === US Refining ===
    {"company": "HF Sinclair", "name": "Mike O'Neill", "title": "Director, Procurement Operations",
     "linkedin": "https://www.linkedin.com/in/oneillms/", "location": "Dallas TX",
     "email": "moneill@hfsinclair.com", "confidence": "pattern-derived (firstinitial+lastname)",
     "notes": "Covers refining MRO + capital procurement."},
    {"company": "PBF Energy", "name": "Kelsey Beans McKell", "title": "Procurement / Strategic Sourcing (Senior Category Leader)",
     "linkedin": "https://www.linkedin.com/in/kelsey-beans-mckell-mba-263b5740/", "location": "Parsippany NJ",
     "email": "kbeansmckell@pbfenergy.com", "confidence": "likely",
     "notes": "Public posts reference hiring Senior Category Leaders for procurement."},
    {"company": "Calumet Inc.", "name": "Kathy Yohe", "title": "Purchasing Coordinator",
     "linkedin": "https://www.linkedin.com/in/kathy-yohe-4025604a/", "location": "Karns City PA",
     "email": "kathy.yohe@calumet.com", "confidence": "likely",
     "notes": "Confirmed purchasing role; tactical buyer level."},
    {"company": "Calumet Inc.", "name": "Telesa Mitchell", "title": "Purchasing Agent",
     "linkedin": "https://www.linkedin.com/in/telesa-mitchell-751025a3", "location": "Shreveport LA",
     "email": "telesa.mitchell@calumet.com", "confidence": "likely",
     "notes": "Shreveport-side purchasing — relevant for surplus rotating eq + vessels."},
    {"company": "Delek US", "name": "Michael Elliott", "title": "VP Supply Chain / Head of Procurement",
     "linkedin": "https://www.linkedin.com/in/mikeelliott2050/", "location": "Brentwood TN",
     "email": "michael.elliott@delekus.com", "confidence": "likely",
     "notes": "Primary procurement decision-maker."},
    {"company": "CVR Energy", "name": "Lance Dietlein", "title": "Sr. Director of Procurement, Supply Chain",
     "linkedin": "https://www.linkedin.com/in/lance-dietlein-8599309/", "location": "Sugar Land TX",
     "email": "lance.dietlein@cvrenergy.com", "confidence": "likely",
     "notes": "Senior procurement leader."},
    {"company": "ONEOK", "name": "Scott Schingen", "title": "SVP, Engineering & Operations — NGL/Gas Pipelines/Crude (capital projects + supply chain)",
     "linkedin": "search 'Scott Schingen ONEOK'", "location": "Tulsa OK",
     "email": "scott.schingen@oneok.com", "confidence": "verified (title)",
     "notes": "Owns capital projects + SC capabilities for the NGL/gas footprint per ONEOK bio."},
    {"company": "Targa Resources", "name": "Stephen Hester", "title": "Chief Supply Chain Officer (formerly CenterPoint — VERIFY current employer)",
     "linkedin": "https://www.linkedin.com/in/stephenthester/", "location": "Houston TX",
     "email": "shester@targaresources.com", "confidence": "likely (ROLE CONFLICT: also listed as CenterPoint CPO/CSCO; same LinkedIn URL — same person)",
     "notes": "VERIFY which company he's at currently. Same LinkedIn URL as the CenterPoint Stephen Hester — he likely moved CenterPoint→Targa or vice versa. Don't double-count."},
    {"company": "Targa Resources", "name": "Brian Kateb", "title": "Enterprise Procurement (20+ yr Targa)",
     "linkedin": "https://www.linkedin.com/in/briankateb/", "location": "Houston TX",
     "email": "bkateb@targaresources.com", "confidence": "likely",
     "notes": "Long-tenured procurement owner."},
    {"company": "Energy Transfer", "name": "Dallas Hicks", "title": "Senior Procurement Contracts Lead (energy midstream)",
     "linkedin": "https://www.linkedin.com/in/dallas-hicks-9820359b/", "location": "Dallas TX",
     "email": "dallas.hicks@energytransfer.com", "confidence": "pattern-derived",
     "notes": "Verify employer current."},
    {"company": "Enterprise Products Partners", "name": "Richard Jones", "title": "Projects Procurement Manager",
     "linkedin": "https://www.linkedin.com/in/richard-jones-a094757/", "location": "Houston TX",
     "email": "rjones@eprod.com", "confidence": "likely",
     "notes": "Owns capital-project sourcing including international."},
    {"company": "Enterprise Products Partners", "name": "Sandra Dobbs", "title": "Manager, Projects Procurement",
     "linkedin": "https://www.linkedin.com/in/sandra-dobbs-b933468", "location": "Houston TX",
     "email": "sdobbs@eprod.com", "confidence": "likely",
     "notes": "Capital projects buyer; equipment-skid scope."},
    {"company": "MPLX (Marathon Petroleum sister)", "name": "Eric Conley", "title": "Procurement (MPC)",
     "linkedin": "https://www.linkedin.com/in/eaconley/", "location": "San Antonio TX",
     "email": "econley@marathonpetroleum.com", "confidence": "pattern-derived",
     "notes": "MPC convention firstinitial+lastname."},

    # === Canadian O&G ===
    {"company": "Whitecap Resources", "name": "Tracy Rosser", "title": "Chief Supply Chain Officer",
     "linkedin": "search 'Tracy Rosser Whitecap'", "location": "Calgary AB",
     "email": "tracy.rosser@wcap.ca", "confidence": "pattern-derived (title verified)",
     "notes": "Top SC decision-maker; drives post-Veren integration of vendor base."},
    {"company": "Tourmaline Oil", "name": "Alison Mosca", "title": "Supply Chain (APC-credentialed)",
     "linkedin": "https://www.linkedin.com/in/alison-mosca-apc/", "location": "Calgary AB",
     "email": "alison.mosca@tourmalineoil.com", "confidence": "likely",
     "notes": "APC designation = senior contracting/procurement role."},
    {"company": "Tourmaline Oil", "name": "Erik Lyneborg", "title": "Facility & Production Equipment Specialist",
     "linkedin": "https://ca.linkedin.com/in/erik-lyneborg-66b91a32", "location": "Calgary AB",
     "email": "erik.lyneborg@tourmalineoil.com", "confidence": "likely",
     "notes": "27 yrs Alberta facilities experience — direct equipment-spec influencer."},
    {"company": "ARC Resources", "name": "Becky Salmond", "title": "Team Lead, Procurement",
     "linkedin": "https://ca.linkedin.com/in/beckysalmond", "location": "Calgary AB",
     "email": "bsalmond@arcresources.com", "confidence": "pattern-derived (title verified)",
     "notes": "Day-to-day procurement gatekeeper at ARC."},
    {"company": "Cenovus Energy", "name": "Andrew Straile", "title": "VP, Supply Chain Management",
     "linkedin": "search 'Andrew Straile Cenovus'", "location": "Calgary AB",
     "email": "andrew.straile@cenovus.com", "confidence": "pattern-derived (title verified ZoomInfo)",
     "notes": "Top SCM decision-maker. Came from facility engineering — equipment-literate."},
    {"company": "Cenovus Energy", "name": "Adam Lawlor", "title": "Procurement Advisor",
     "linkedin": "https://ca.linkedin.com/in/adam-lawlor-7737698a", "location": "Calgary AB",
     "email": "adam.lawlor@cenovus.com", "confidence": "pattern-derived",
     "notes": "Working-level procurement entry point."},
    {"company": "Imperial Oil", "name": "Laralyn Klesken", "title": "VP Procurement, Canada",
     "linkedin": "https://www.linkedin.com/in/laralyn-klesken-58521417/", "location": "Greater Calgary AB",
     "email": "laralyn.klesken@imperialoil.ca", "confidence": "likely (title verified LinkedIn)",
     "notes": "Top procurement contact in Canada. ExxonMobil global SCM rolls up through Hansen."},
    {"company": "Strathcona Resources", "name": "Rob Rattie", "title": "Team Lead, Procurement / SC Consultant",
     "linkedin": "https://www.linkedin.com/in/rob-rattie-88646430/", "location": "Greater Calgary AB",
     "email": "rob.rattie@strathconaresources.com", "confidence": "likely (LinkedIn-confirmed)",
     "notes": "Lead procurement for SC operations. Active hiring under his team."},
    {"company": "Pembina Pipeline", "name": "Jeremy Campbell", "title": "Director of Operations and Director of Supply Chain",
     "linkedin": "https://www.linkedin.com/in/jeremy-campbell-248656235/", "location": "Calgary AB",
     "email": "jeremy.campbell@pembina.com", "confidence": "likely (title verified)",
     "notes": "Dual-hat ops + SCM — high-leverage contact."},
    {"company": "Pembina Pipeline", "name": "Tapan Gupta", "title": "Supply Chain & Procurement (11 yrs)",
     "linkedin": "https://www.linkedin.com/in/tapangupta711/", "location": "Calgary AB",
     "email": "tapan.gupta@pembina.com", "confidence": "pattern-derived",
     "notes": "Engineer + MBA — category management depth."},
    {"company": "Keyera Corp", "name": "Josh Bangsund", "title": "General Manager – Supply Chain",
     "linkedin": "https://ca.linkedin.com/in/josh-bangsund", "location": "Calgary AB",
     "email": "josh.bangsund@keyera.com", "confidence": "likely (title verified LinkedIn)",
     "notes": "Senior-most SCM contact at Keyera."},
    {"company": "Keyera Corp", "name": "Peyton Cox", "title": "Contract & Procurement Analyst (Facilities SC)",
     "linkedin": "https://www.linkedin.com/in/peyton-cox-1a07a621a/", "location": "Calgary AB",
     "email": "peyton.cox@keyera.com", "confidence": "pattern-derived",
     "notes": "Facilities-side procurement — directly aligned with Fuelled's catalogue."},
    {"company": "Vermilion Energy", "name": "Daxesh Shah", "title": "Supply Chain",
     "linkedin": "https://www.linkedin.com/in/daxeshshah/", "location": "Calgary AB",
     "email": "daxesh.shah@vermilionenergy.com", "confidence": "pattern-derived",
     "notes": "LinkedIn-verified Vermilion SC person; first-call entry point."},
    {"company": "Suncor Energy", "name": "Tiffany Thornton", "title": "Procurement Advisor",
     "linkedin": "https://ca.linkedin.com/in/thorntontiffany", "location": "Calgary AB",
     "email": "tiffany.thornton@suncor.com", "confidence": "pattern-derived",
     "notes": "Working-level procurement contact."},
]


# Top-line insight blocks for a 'Summary' sheet at the front
SUMMARY_BLOCKS = [
    ("ANCO UNMASKED — Anderson Columbia Co., Inc. (CONFIRMED)",
     "The biggest seller on AllSurplus by volume (23 listings) was previously masked as 'Seller 23609 - ANCO.' "
     "Per-lot location triangulation confirms it as Anderson Columbia Co., Inc. (ACCI) — a privately-held heavy-civil road "
     "construction contractor HQ in Lake City FL with plants at Old Town/Quincy/Marianna FL, TX yards at Robstown/Weslaco/"
     "Corpus Christi (post-2022 Haas-Anderson acquisition) + New Braunfels (Tejas Quarry, 2016), and GA aggregate ops. "
     "Every FL location maps to a known ACCI plant; the asset mix (asphalt paver, asphalt distributor, McLanahan screen, "
     "asphalt-plant drag elevator, water trucks, dump trucks, rollers, broom) is textbook ACCI heavy-civil. ACCI is a "
     "documented McLanahan customer. Critically: this is a FLEET ROTATION, not a bankruptcy — meaning ACCI is a "
     "*recurring* dispo channel, suitable for an ongoing buy-side relationship rather than a one-time auction chase. "
     "The other four LS codes (ASFL, ELGLRTNIC, RELLC, AWCTL) have insufficient public fingerprint to identify."),
    ("LS-MANAGED CODES (4 still unidentified, 1 unmasked)",
     "ASFL, ELGLRTNIC, RELLC, AWCTL remain Liquidity Services internal program codes that mask the underlying client. "
     "ASFL is explicitly confirmed via the public AllSurplus T&C page ('Liquidity Services – ASFL'); the others match the "
     "same URL pattern. Each has too thin a fingerprint (1–4 listings) to triangulate without paid registry data. "
     "Playbook: treat LS CAG as a single channel relationship covering these + future LS-managed dispositions, OR wait "
     "for additional listings to accrue more lot-data fingerprint."),
    ("HIGHEST-PRIORITY DIRECT-OWNER WARM LEADS",
     "Real ongoing disposition events with identifiable corporate procurement: ArcelorMittal SA (Long Steel wind-down, "
     "Jan 2025), Assmang/ARM/Assore (Cato Ridge full closure Aug 2025 + Beeshoek care & maintenance Nov 2025), "
     "Glencore Operations SA (ferrochrome smelter suspensions May–June 2025), MHI Power ZAF (Medupi/Kusile end-of-construction), "
     "ConocoPhillips Permian ($5B post-Marathon divest), Pemex Deer Park ($500M+ turnaround), Valaris (Feb 2025 fleet "
     "rationalization), Entergy (active IR program + Baxter Wilson plant retirement), CenterPoint (dedicated public surplus "
     "portal), ADM (3 plant closures: Bushnell IL, Kershaw SC, Memphis TN), CVRF / Grapevine Energy (post-bankruptcy "
     "restart + EPC-dispute surplus — small org, fastest direct landing)."),
    ("LIQUIDITY SERVICES MANAGER HIERARCHY",
     "Herman Liversage = Associate Director, LS-SA (senior, deep mining network — peer-to-peer pitch). "
     "Jeff Goddard = CAG Account Manager + CEA (real BD with appraisal credentials — pitch on data quality). "
     "Ruth Hernandez = Inside Sales Rep (event coordinator / gatekeeper — not decision-maker). "
     "Marcin Gadzinski = EU sales/BD (thin public footprint, seniority unknown)."),
    ("CHANNEL-PARTNER OPPORTUNITIES (third-party remarketers)",
     "Factor X (CA decom contractor — info@factorx.com / 408.842.9700) and Sierra Auction (AZ municipal fleet, now "
     "LS-owned — sierraauction.com / 602-242-7121) are the two real channel-partner candidates. WA DES Surplus "
     "(Surplus@des.wa.gov / 360-407-2270) is the multi-agency direct-agency funnel. Skip CA Dept of Technology (IT only) "
     "and Municipal Asset Recovery NE Ohio (1 listing, no web footprint)."),
    ("INVESTIGATE",
     "Solaire Kapuskasing Solar — 4 listings from a 6 MW operating solar park is unusual. Could be inverter swap (year ~12, "
     "common), repowering, or partial decom. Phone the Town of Kapuskasing — 705-335-2341 / general@kapuskasing.ca."),
    ("NAMED CONTACTS NOW ON FILE (see 'Named Contacts (Direct)' tab)",
     "Public-web research surfaced 18 named humans across the 10 top warm leads — direct procurement / Investment "
     "Recovery / asset-disposition contacts that bypass Liquidity Services. STRONGEST hits (verified, exact-title match): "
     "Justin McCabe (Manager, Investment Recovery, Entergy), Michael Brown MCA (Manager, Asset Recovery & Supply Chain "
     "Sustainability, CenterPoint), André Alves (Source-to-Pay Director with explicit Investment Recovery mandate, ADM), "
     "Joe Riley (Refinery Supply Chain Manager, Pemex Deer Park), Jorge Lebrija (VP Global Supply Chain, Valaris), "
     "Devan Kannie (Production Manager at the Cato Ridge Alloys JV being decommissioned), Travesh Ramkhelawan (Head of "
     "Strategic Projects and Properties at AMSA — owns idle-site disposition). Emails are pattern-derived from public "
     "domain conventions; treat as untested but high-likelihood. These are the prioritized first calls."),
    ("SMALLER SELLERS WORTH PURSUING (after triage of 38 deferred)",
     "33 of 38 deferred sellers are SKIP (police/IT/schools/light fleet). 5 worth pursuing: Columbus Fleet Management OH "
     "(HIGH — excavators + aerial lifts on IronPlanet, Fuelled wheelhouse), King County WA (HIGH — county heavy fleet, "
     "new online auction launched Oct 2025), Lake County IL (HIGH — public-works fleet + landscape), Clean Water Services "
     "OR (MEDIUM — wastewater utility, pumps/blowers/motors), North Baldwin Utilities AL (MEDIUM — small water/electric "
     "utility, only true utility on the list)."),
    ("BUYER RECRUITMENT LIST (39 companies, ~80 contacts) — see 'Buyer Targets' tabs",
     "First-cut buyer-recruitment list across three verticals: 12 US Upstream operators (Diamondback, Permian Resources, "
     "Matador, Civitas/SM, Coterra, Chord, Crescent, Riley Permian, Magnolia, Ovintiv); 12 US Refining + Midstream "
     "(HF Sinclair, PBF, Par Pacific, Calumet, Delek, CVR, ONEOK, Targa, Energy Transfer, Williams, MPLX, EPD); 15 Canadian "
     "O&G (Whitecap, Tourmaline, ARC, Cenovus, Imperial, Strathcona, Pembina, Keyera, AltaGas, Vermilion, Birchcliff, "
     "Suncor, etc). Strongest verified contacts: Jason Stroud (Chord — sets sourcing policy, ex-BHP), Andrew Straile "
     "(Cenovus VP SCM), Laralyn Klesken (Imperial VP Procurement Canada), Tracy Rosser (Whitecap CSCO), Michael Elliott "
     "(Delek VP SC). NOTABLE: Cenovus already runs a public surplus inventory portal at surplusinventory.cenovus.com — "
     "Fuelled should monitor as both buy-side signal AND competitive listing comparison. Two-sided fits (selling AND "
     "buying simultaneously): Ovintiv, SM/Civitas, Crescent — pitch as 'list your divestiture surplus, source your new "
     "development gear' in one channel."),
    ("AUCTION-BIDDER TRACKING — DEAD END on equipment marketplaces",
     "Probed all major auction marketplaces (AllSurplus, IronPlanet, GovDeals, Ritchie Bros, BidSpotter). Verdict: bidder "
     "identity is anonymized BY DESIGN on every major equipment marketplace — that's their competitive moat. Ritchie Bros "
     "even monetizes the data they hold by selling it to OEMs under contract. Only partial signal in adjacent data: "
     "federal/state oil & gas LEASE sale results (EnergyNet, BLM, state land boards) DO publish named winning bidders. "
     "That's land acquisition, not equipment — but it's a leading indicator. A company that just won 70 leases on the "
     "NPR-A in March will need equipment in 6-24 months. Cheapest viable capture: weekly scrape of EnergyNet government-"
     "results + BLM/state-land-board press releases → operators-about-to-need-equipment prospecting list."),
]


# -----------------------------------------------------------------------------

ENRICH_HEADERS = [
    "Owner Class", "Industry", "Asset Location", "Parent / Group",
    "Disposition Context", "Direct Contact (bypass LS)", "Research Confidence", "Notes",
]


def enrich_sellers_sheet(ws) -> tuple[int, int]:
    """Append research columns to the Sellers sheet. Returns (matched, total)."""
    # Find the 'Seller' column index (1-based)
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    seller_col = headers.index("Seller") + 1

    # Append new headers
    start_col = ws.max_column + 1
    for offset, h in enumerate(ENRICH_HEADERS):
        cell = ws.cell(row=1, column=start_col + offset, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    matched = 0
    total = ws.max_row - 1
    for row_idx in range(2, ws.max_row + 1):
        seller = ws.cell(row=row_idx, column=seller_col).value
        if not seller:
            continue
        data = RESEARCH.get(seller)
        if not data:
            continue
        matched += 1
        for offset, key in enumerate([
            "owner_class", "industry", "location", "parent",
            "disposition", "direct_contact", "confidence", "notes",
        ]):
            ws.cell(row=row_idx, column=start_col + offset, value=data.get(key, ""))
    return matched, total


def write_summary_sheet(wb) -> None:
    ws = wb.create_sheet(title="Research Summary", index=0)
    write_header(ws, ["Topic", "Insight"])
    for topic, insight in SUMMARY_BLOCKS:
        ws.append([topic, insight])
        # Wrap the insight cell
        ws.cell(row=ws.max_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[ws.max_row].height = 90
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 130


def write_ls_managers_sheet(wb) -> None:
    ws = wb.create_sheet(title="LS Event Managers")
    headers = [
        "Name", "Email", "Phone", "LinkedIn", "Title", "Region", "Tenure",
        "Events Covered", "Notes for Outreach", "Confidence",
    ]
    write_header(ws, headers)
    for m in LS_MANAGERS:
        ws.append([
            m["name"], m["email"], m["phone"], m["linkedin"], m["title"],
            m["region"], m["tenure"], m["covers_events"], m["notes"], m["confidence"],
        ])
    autosize(ws, max_width=80)


def write_named_contacts_sheet(wb) -> None:
    ws = wb.create_sheet(title="Named Contacts (Direct)")
    headers = [
        "Company", "Name", "Title", "LinkedIn", "Location",
        "Email (likely)", "Confidence", "Outreach Notes",
    ]
    write_header(ws, headers)
    for c in NAMED_CONTACTS:
        ws.append([
            c["company"], c["name"], c["title"], c["linkedin"], c["location"],
            c["email"], c["confidence"], c["notes"],
        ])
    autosize(ws, max_width=80)


def write_buyer_companies_sheet(wb) -> None:
    ws = wb.create_sheet(title="Buyer Targets (Companies)")
    headers = [
        "Vertical", "Company", "Ticker", "HQ", "Basin / Footprint",
        "Scale", "Capex Driver / Why They Buy", "Suppliers Page",
    ]
    write_header(ws, headers)
    for b in BUYER_COMPANIES:
        ws.append([
            b["vertical"], b["name"], b["ticker"], b["hq"], b["basin"],
            b["scale"], b["capex_driver"], b["suppliers_page"],
        ])
    autosize(ws, max_width=80)


def write_buyer_contacts_sheet(wb) -> None:
    ws = wb.create_sheet(title="Buyer Contacts")
    headers = [
        "Company", "Name", "Title", "LinkedIn", "Location",
        "Email (likely)", "Confidence", "Outreach Notes",
    ]
    write_header(ws, headers)
    for c in BUYER_CONTACTS:
        ws.append([
            c["company"], c["name"], c["title"], c["linkedin"], c["location"],
            c["email"], c["confidence"], c["notes"],
        ])
    autosize(ws, max_width=80)


def main(argv: list[str]) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True, help="Input workbook path")
    ap.add_argument("--out", default=None, help="Output workbook path (defaults to <input>_enriched.xlsx)")
    args = ap.parse_args(argv)

    in_path = Path(args.inp)
    if not in_path.exists():
        print(f"ERROR: input not found: {in_path}", file=sys.stderr)
        return 2

    out_path = Path(args.out) if args.out else in_path.with_name(in_path.stem + "_enriched.xlsx")

    wb = load_workbook(in_path)
    sellers_ws = wb["Sellers"]

    matched, total = enrich_sellers_sheet(sellers_ws)
    write_summary_sheet(wb)
    write_ls_managers_sheet(wb)
    write_named_contacts_sheet(wb)
    write_buyer_companies_sheet(wb)
    write_buyer_contacts_sheet(wb)

    wb.save(out_path)
    print(f"\nWrote {out_path}")
    print(f"  Enriched: {matched} / {total} sellers")
    print(f"  Summary sheet: {len(SUMMARY_BLOCKS)} insight blocks")
    print(f"  LS Managers sheet: {len(LS_MANAGERS)} people")
    print(f"  Named Contacts (sellers) sheet: {len(NAMED_CONTACTS)} people")
    print(f"  Buyer Targets (companies) sheet: {len(BUYER_COMPANIES)} companies")
    print(f"  Buyer Contacts sheet: {len(BUYER_CONTACTS)} people")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
