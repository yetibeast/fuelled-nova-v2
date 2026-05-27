"""Import the May 10 supply-targets workbook into seller_contact_enrichment
+ buyer_targets.

Mirrors the dealer_contacts pattern (2026-05-13_create_dealer_contacts.sql +
seed_dealer_contacts.sql) but reads from an xlsx instead of inline INSERT
statements, so the team can refresh the enrichment by editing the workbook
and re-running this script.

Idempotent: ON CONFLICT (...) DO UPDATE SET ... so re-runs reflect xlsx
edits without producing duplicate rows.

Usage:
    PYTHONPATH=backend python3 backend/scripts/import_may10_enrichment.py [--db-url ...] [path/to/xlsx]

Default xlsx path: docs/supply-targets/supply_targets_enriched_2026-05-10.xlsx
Env: DATABASE_URL (or pass --db-url).
"""
from __future__ import annotations

import argparse
import asyncio
import os
import sys
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from sqlalchemy import text
from sqlalchemy.ext.asyncio import create_async_engine


DEFAULT_XLSX = "docs/supply-targets/supply_targets_enriched_2026-05-10.xlsx"


# ── Cell normalization ──────────────────────────────────────────────────


def _clean(v: Any) -> str | None:
    """Coerce openpyxl cell value → trimmed str or None.

    Treats whitespace-only strings as None so the unique indexes treat
    blank emails as missing rather than equal.
    """
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _row_dict(headers: list[str], row: Iterable[Any]) -> dict[str, Any]:
    return {h: _clean(v) for h, v in zip(headers, row)}


def _load_sheet(xlsx_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers_raw = next(rows_iter)
    except StopIteration:
        return []
    headers = [str(h) if h is not None else "" for h in headers_raw]
    out: list[dict[str, Any]] = []
    for raw in rows_iter:
        d = _row_dict(headers, raw)
        # Drop fully-empty rows
        if not any(v for v in d.values()):
            continue
        out.append(d)
    return out


# ── Parsers ─────────────────────────────────────────────────────────────


def parse_named_contacts(xlsx_path: Path) -> list[dict[str, Any]]:
    """`Named Contacts (Direct)` sheet → seller_contact_enrichment rows.

    Source is left NULL so the enrichment applies to every source for
    that seller_name. Rows with no Company or no Name are skipped.
    """
    raw = _load_sheet(xlsx_path, "Named Contacts (Direct)")
    out: list[dict[str, Any]] = []
    for r in raw:
        company = r.get("Company")
        name = r.get("Name")
        if not company or not name:
            continue
        out.append({
            "seller_name": company,
            "source": None,
            "contact_name": name,
            "contact_title": r.get("Title"),
            "contact_email": r.get("Email (likely)"),
            "contact_phone": None,
            "contact_linkedin": r.get("LinkedIn"),
            "contact_confidence": r.get("Confidence"),
            "location": r.get("Location"),
            "outreach_notes": r.get("Outreach Notes"),
        })
    return out


def parse_ls_event_managers(xlsx_path: Path) -> list[dict[str, Any]]:
    """`LS Event Managers` sheet → seller_contact_enrichment rows.

    These are Liquidity Services people who cover AllSurplus events, so we
    scope them to source='allsurplus' and bucket them under a canonical
    seller_name='Liquidity Services'. Outreach notes get an explicit
    "[LS Event Manager — covers events: X, Y]" prefix so reviewers can
    distinguish them from named-contact rows.
    """
    raw = _load_sheet(xlsx_path, "LS Event Managers")
    out: list[dict[str, Any]] = []
    for r in raw:
        name = r.get("Name")
        if not name:
            continue
        events = r.get("Events Covered") or "(unspecified)"
        notes_orig = r.get("Notes for Outreach") or ""
        prefix = f"[LS Event Manager — covers events: {events}]"
        notes_combined = (prefix + " " + notes_orig).strip() if notes_orig else prefix
        out.append({
            "seller_name": "Liquidity Services",
            "source": "allsurplus",
            "contact_name": name,
            "contact_title": r.get("Title"),
            "contact_email": r.get("Email"),
            "contact_phone": r.get("Phone"),
            "contact_linkedin": r.get("LinkedIn"),
            "contact_confidence": r.get("Confidence"),
            "location": r.get("Region"),
            "outreach_notes": notes_combined,
        })
    return out


def parse_buyer_targets(xlsx_path: Path) -> list[dict[str, Any]]:
    """`Buyer Targets (Companies)` + `Buyer Contacts` → buyer_targets rows.

    Each buyer-contact row carries the company-level metadata. Companies
    without any contacts get a single row with contact fields NULL so
    they still surface in the buyers.csv export.
    """
    companies = _load_sheet(xlsx_path, "Buyer Targets (Companies)")
    contacts = _load_sheet(xlsx_path, "Buyer Contacts")

    by_company: dict[str, dict[str, Any]] = {}
    for c in companies:
        company_name = c.get("Company")
        if not company_name:
            continue
        by_company[company_name] = {
            "vertical": c.get("Vertical"),
            "company": company_name,
            "ticker": c.get("Ticker"),
            "hq": c.get("HQ"),
            "basin": c.get("Basin / Footprint"),
            "scale": c.get("Scale"),
            "capex_driver": c.get("Capex Driver / Why They Buy"),
            "suppliers_page": c.get("Suppliers Page"),
        }

    matched: set[str] = set()
    out: list[dict[str, Any]] = []
    for ct in contacts:
        company_name = ct.get("Company")
        name = ct.get("Name")
        if not company_name or not name:
            continue
        meta = by_company.get(company_name, {
            "vertical": None,
            "company": company_name,
            "ticker": None,
            "hq": None,
            "basin": None,
            "scale": None,
            "capex_driver": None,
            "suppliers_page": None,
        })
        out.append({
            **meta,
            "contact_name": name,
            "contact_title": ct.get("Title"),
            "contact_email": ct.get("Email (likely)"),
            "contact_linkedin": ct.get("LinkedIn"),
            "contact_confidence": ct.get("Confidence"),
            "location": ct.get("Location"),
            "outreach_notes": ct.get("Outreach Notes"),
        })
        matched.add(company_name)

    # Companies with no contact match → one company-only row.
    for company_name, meta in by_company.items():
        if company_name in matched:
            continue
        out.append({
            **meta,
            "contact_name": None,
            "contact_title": None,
            "contact_email": None,
            "contact_linkedin": None,
            "contact_confidence": None,
            "location": None,
            "outreach_notes": None,
        })
    return out


# ── DB upserts ──────────────────────────────────────────────────────────


_UPSERT_SELLER_SQL = text("""
    INSERT INTO seller_contact_enrichment (
        seller_name, source, contact_name, contact_title,
        contact_email, contact_phone, contact_linkedin,
        contact_confidence, location, outreach_notes,
        enrichment_source, imported_at
    ) VALUES (
        :seller_name, :source, :contact_name, :contact_title,
        :contact_email, :contact_phone, :contact_linkedin,
        :contact_confidence, :location, :outreach_notes,
        'may10_workbook', NOW()
    )
    ON CONFLICT (seller_name, source, contact_email) DO UPDATE SET
        contact_name       = EXCLUDED.contact_name,
        contact_title      = EXCLUDED.contact_title,
        contact_phone      = EXCLUDED.contact_phone,
        contact_linkedin   = EXCLUDED.contact_linkedin,
        contact_confidence = EXCLUDED.contact_confidence,
        location           = EXCLUDED.location,
        outreach_notes     = EXCLUDED.outreach_notes,
        imported_at        = NOW()
""")


_UPSERT_BUYER_SQL = text("""
    INSERT INTO buyer_targets (
        vertical, company, ticker, hq, basin, scale,
        capex_driver, suppliers_page,
        contact_name, contact_title, contact_email,
        contact_linkedin, contact_confidence,
        location, outreach_notes,
        enrichment_source, imported_at
    ) VALUES (
        :vertical, :company, :ticker, :hq, :basin, :scale,
        :capex_driver, :suppliers_page,
        :contact_name, :contact_title, :contact_email,
        :contact_linkedin, :contact_confidence,
        :location, :outreach_notes,
        'may10_workbook', NOW()
    )
    ON CONFLICT (company, contact_email) DO UPDATE SET
        vertical           = EXCLUDED.vertical,
        ticker             = EXCLUDED.ticker,
        hq                 = EXCLUDED.hq,
        basin              = EXCLUDED.basin,
        scale              = EXCLUDED.scale,
        capex_driver       = EXCLUDED.capex_driver,
        suppliers_page     = EXCLUDED.suppliers_page,
        contact_name       = EXCLUDED.contact_name,
        contact_title      = EXCLUDED.contact_title,
        contact_linkedin   = EXCLUDED.contact_linkedin,
        contact_confidence = EXCLUDED.contact_confidence,
        location           = EXCLUDED.location,
        outreach_notes     = EXCLUDED.outreach_notes,
        imported_at        = NOW()
""")


async def _upsert_rows(engine, sql, rows: list[dict[str, Any]]) -> int:
    if not rows:
        return 0
    async with engine.begin() as conn:
        for r in rows:
            await conn.execute(sql, r)
    return len(rows)


async def _run(xlsx_path: Path, db_url: str | None) -> int:
    seller_named = parse_named_contacts(xlsx_path)
    seller_ls = parse_ls_event_managers(xlsx_path)
    buyer_rows = parse_buyer_targets(xlsx_path)

    print(f"Parsed from {xlsx_path}:")
    print(f"  Named Contacts (Direct):        {len(seller_named):>4} rows")
    print(f"  LS Event Managers (allsurplus): {len(seller_ls):>4} rows")
    print(f"  Buyer Targets + Contacts:       {len(buyer_rows):>4} rows")

    if not db_url:
        print("\n--db-url / DATABASE_URL not set — dry-run only. No writes.", file=sys.stderr)
        return 0

    url = db_url.replace("postgresql://", "postgresql+asyncpg://", 1)
    engine = create_async_engine(url)
    try:
        n_seller = await _upsert_rows(engine, _UPSERT_SELLER_SQL, seller_named + seller_ls)
        n_buyer = await _upsert_rows(engine, _UPSERT_BUYER_SQL, buyer_rows)
        print(f"\nUpserted {n_seller} seller_contact_enrichment rows.")
        print(f"Upserted {n_buyer} buyer_targets rows.")
    finally:
        await engine.dispose()
    return 0


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("xlsx", nargs="?", default=DEFAULT_XLSX,
                   help=f"Path to workbook (default: {DEFAULT_XLSX})")
    p.add_argument("--db-url", default=os.environ.get("DATABASE_URL"),
                   help="Postgres URL. Defaults to $DATABASE_URL. Omit for dry-run.")
    args = p.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERROR: workbook not found: {xlsx_path}", file=sys.stderr)
        return 2

    return asyncio.run(_run(xlsx_path, args.db_url))


if __name__ == "__main__":
    sys.exit(main())
