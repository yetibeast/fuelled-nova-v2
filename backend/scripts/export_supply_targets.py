"""Export the supply-targets list as a multi-sheet Excel workbook.

Sheet 1 (Sellers): one row per seller — listing count, totals, contact info.
Sheet 2 (Listings): every individual listing with full seller attribution + URL.
Sheet 3+ (per-source): one Listings sheet per requested source so Mark can
filter by marketplace without learning Excel filters.

Run from repo root:
    python3 backend/scripts/export_supply_targets.py \\
        --source allsurplus --source kijiji \\
        [--api https://api.fuellednova.com] [--out PATH]

Hits prod via the admin endpoint, so it picks up whatever the live DB has.
"""
from __future__ import annotations

import argparse
import os
import sys
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import jwt
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def mint_admin_token(jwt_secret: str) -> str:
    return jwt.encode(
        {"sub": "supply-targets-export", "role": "admin", "exp": int(time.time()) + 3600},
        jwt_secret, algorithm="HS256",
    )


def fetch(api_base: str, token: str, path: str, **params) -> list[dict]:
    r = requests.get(
        f"{api_base.rstrip('/')}{path}",
        headers={"Authorization": f"Bearer {token}"},
        params=params, timeout=120,
    )
    r.raise_for_status()
    return r.json()


def autosize(ws):
    for col_idx, col in enumerate(ws.columns, start=1):
        try:
            length = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        except ValueError:
            length = 10
        ws.column_dimensions[get_column_letter(col_idx)].width = min(length + 2, 60)


HEADER_FILL = PatternFill("solid", fgColor="EF5D28")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def write_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"


SELLERS_HEADERS = [
    "Rank", "Source", "Seller", "Account Type", "Anonymous?",
    "Listings", "Consignments", "Events", "Priced",
    "Total Asking (USD)", "Total Current Bid (USD)",
    "Contact Name", "Contact Email", "Contact Phone",
    "Other Assets URL", "Last Seen", "Seller Key (internal)",
]

LISTINGS_HEADERS = [
    "Source", "Seller", "Account Type",
    "Title", "Category", "Make", "Model", "Year", "Condition",
    "Asking Price", "Current Bid", "Currency", "Location",
    "Event Title", "Contact Name", "Contact Email", "Contact Phone",
    "URL", "Last Seen", "Listing ID",
]


def write_sellers_sheet(wb: Workbook, sellers: list[dict]) -> None:
    ws = wb.active
    ws.title = "Sellers"
    write_header(ws, SELLERS_HEADERS)

    ranked = sorted(sellers, key=lambda x: (-x["listing_count"], -(x.get("total_asking") or 0)))
    for idx, s in enumerate(ranked, start=1):
        ws.append([
            idx,
            s.get("source", "-"),
            s.get("seller_name") or f"(anonymous: {s['seller_key']})",
            s.get("account_type") or "-",
            "yes" if s.get("is_anonymous") else "no",
            s["listing_count"],
            s.get("consignment_count", 1),
            s.get("event_count", 0),
            s.get("priced_count", 0),
            s.get("total_asking"),
            s.get("total_current_bid"),
            s.get("contact_name") or "",
            s.get("contact_email") or "",
            s.get("contact_phone") or "",
            s.get("other_assets_url") or "",
            (s.get("last_seen") or "").split("T")[0],
            s["seller_key"],
        ])
    autosize(ws)


def write_listings_sheet(wb: Workbook, name: str, listings: list[dict]) -> None:
    ws = wb.create_sheet(title=name)
    write_header(ws, LISTINGS_HEADERS)
    for L in listings:
        ws.append([
            L.get("source", "-"),
            L.get("seller_name") or f"(anonymous: {L.get('seller_key')})",
            L.get("seller_account_type") or "-",
            L.get("title") or "",
            L.get("category") or "",
            L.get("make") or "",
            L.get("model") or "",
            L.get("year") or "",
            L.get("condition") or "",
            L.get("asking_price"),
            L.get("current_bid"),
            L.get("currency") or "",
            L.get("location") or "",
            L.get("event_title") or "",
            L.get("contact_name") or "",
            L.get("contact_email") or "",
            L.get("contact_phone") or "",
            L.get("url") or "",
            (L.get("last_seen") or "").split("T")[0],
            L.get("id") or "",
        ])
    autosize(ws)


def main(argv: list[str]) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", action="append", default=[],
                    help="Marketplace to include (repeat for multiple). Default: allsurplus + kijiji.")
    ap.add_argument("--api", default="https://api.fuellednova.com")
    ap.add_argument("--secret", default=os.environ.get("JWT_SECRET", ""),
                    help="JWT signing secret (or set JWT_SECRET env var)")
    ap.add_argument("--min-listings", type=int, default=1,
                    help="Drop sellers with fewer than this many listings")
    ap.add_argument("--out", default=None)
    args = ap.parse_args(argv)

    sources = args.source or ["allsurplus", "kijiji"]

    if not args.secret:
        print("ERROR: provide --secret or set JWT_SECRET", file=sys.stderr)
        return 2

    token = mint_admin_token(args.secret)

    all_sellers: list[dict] = []
    listings_by_source: dict[str, list[dict]] = {}

    for source in sources:
        print(f"\n=== {source} ===")
        print("Fetching seller aggregate ...")
        sellers = fetch(args.api, token, "/api/admin/supply-targets",
                        source=source, min_listings=args.min_listings, limit=5000)
        print(f"  {len(sellers)} sellers")
        all_sellers.extend(sellers)

        print("Fetching per-listing rows ...")
        listings = fetch(args.api, token, "/api/admin/supply-targets/listings",
                         source=source, limit=50000)
        print(f"  {len(listings)} listings")
        listings_by_source[source] = listings

    out_path = Path(args.out) if args.out else Path(
        f"docs/supply-targets/supply_targets_{'_'.join(sources)}_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    write_sellers_sheet(wb, all_sellers)

    # Combined listings sheet (everything in one place)
    combined: list[dict] = []
    for source in sources:
        combined.extend(listings_by_source[source])
    write_listings_sheet(wb, "Listings (all)", combined)

    # Per-source sheets — only when more than one source was requested
    if len(sources) > 1:
        for source in sources:
            write_listings_sheet(wb, f"Listings ({source})", listings_by_source[source])

    wb.save(out_path)
    print(f"\nWrote {out_path}")
    print(f"  Sellers:   {len(all_sellers)}")
    print(f"  Listings:  {len(combined)}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
